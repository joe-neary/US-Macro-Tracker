#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
US Economic Indicators Tracker  v8
────────────────────────────────────
Grabs US macro data from FRED, rips PMI numbers straight out of PDFs,
and spits out a polished Excel dashboard. That's basically it.

Drop the PDFs in, run the script, open the xlsx. Done.

Run: py -3 economic_tracker.py
"""
import sys
sys.dont_write_bytecode = True   # keeps the project folder clean (no __pycache__)

import os, re, json, time, requests, pdfplumber, openpyxl, csv
from openpyxl.styles      import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments    import Comment
from openpyxl.utils       import get_column_letter
from openpyxl.chart       import LineChart, BarChart, Reference
from openpyxl.chart.axis  import ChartLines
from datetime             import datetime

import pandas as pd
import numpy as np
try:
    from statsmodels.tsa.statespace.dynamic_factor import DynamicFactor as _DFM
    _HAS_STATSMODELS = True
except ImportError:
    _HAS_STATSMODELS = False

from scipy.stats import multivariate_normal as _mvn
_HAS_HMM = True     # built-in implementation - no external dependency


# ── CONFIG ────────────────────────────────────────────────────────────────────

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))   # keeps the project portable

# Load .env file if present (keeps secrets out of source code)
_env_path = os.path.join(BASE_DIR, ".env")
if os.path.isfile(_env_path):
    with open(_env_path) as _ef:
        for _line in _ef:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())

# FRED API key - get a free key at https://fred.stlouisfed.org/docs/api/api_key.html
# Set via environment variable or place in a .env file in the project root.
FRED_API_KEY  = os.environ.get("FRED_API_KEY", "")
DATA_DIR      = os.path.join(BASE_DIR, "data")
PDF_FOLDER    = os.path.join(DATA_DIR, "pdfs")
CSV_FOLDER    = os.path.join(DATA_DIR, "csv")
IMPORT_FOLDER = os.path.join(DATA_DIR, "import")
OUTPUT_DIR    = os.path.join(BASE_DIR, "output")
ISM_FOLDER    = PDF_FOLDER                                  # legacy alias used by PDF parsing
MANUAL_FILE   = os.path.join(DATA_DIR, "ISM_Manual_Input.xlsx")
OUTPUT_FILE   = os.path.join(OUTPUT_DIR, "US_Economic_Tracker.xlsx")
TEMPLATE_FILE = os.path.join(OUTPUT_DIR, "US_Economic_Tracker_TEMPLATE.xlsx")
def _find_csv(folder, pattern, exclude=None):
    """Find a CSV file in folder whose name contains the given pattern (case-insensitive).
    Optionally exclude files whose name contains the exclude string."""
    if not os.path.isdir(folder):
        return None
    for f in sorted(os.listdir(folder)):
        fl = f.lower()
        if fl.endswith(".csv") and pattern in fl:
            if exclude and exclude in fl:
                continue
            return os.path.join(folder, f)
    return None

# Default filenames used when creating new CSVs from scratch
_CSV_DEFAULTS = {
    "ism_mfg":      "ism_manufacturing_pmi.csv",
    "ism_svc":      "ism_non_manufacturing_pmi.csv",
    "chicago_pmi":  "chicago_pmi_mni_business_barometer.csv",
}
# Dynamic discovery: find CSVs by keyword pattern, fall back to default name
CSV_ISM_MFG   = _find_csv(CSV_FOLDER, "manufactur", exclude="non") or os.path.join(CSV_FOLDER, _CSV_DEFAULTS["ism_mfg"])
CSV_ISM_SVC   = (_find_csv(CSV_FOLDER, "non_manufactur") or _find_csv(CSV_FOLDER, "non-manufactur")
                 or _find_csv(CSV_FOLDER, "nonmanufactur") or _find_csv(CSV_FOLDER, "services")
                 or os.path.join(CSV_FOLDER, _CSV_DEFAULTS["ism_svc"]))
CSV_CHICAGO   = (_find_csv(CSV_FOLDER, "chicago") or _find_csv(CSV_FOLDER, "barometer")
                 or _find_csv(CSV_FOLDER, "mni") or os.path.join(CSV_FOLDER, _CSV_DEFAULTS["chicago_pmi"]))
CSV_SPG_SVC   = (_find_csv(CSV_FOLDER, "spglobal") or _find_csv(CSV_FOLDER, "sp_services")
                 or _find_csv(CSV_FOLDER, "sp_global") or _find_csv(CSV_FOLDER, "markit")
                 or os.path.join(CSV_FOLDER, "sp_global_services_pmi.csv"))
CSV_PMI_MAP   = {"ism_mfg": CSV_ISM_MFG, "ism_svc": CSV_ISM_SVC, "chicago_pmi": CSV_CHICAGO, "spg_svc": CSV_SPG_SVC}
CACHE_FILE    = os.path.join(BASE_DIR, ".fred_cache.json")
CACHE_MAX_AGE = 72000  # seconds (20 hours) - cache is valid for one trading day
FONT_NAME     = "Calibri"

# Gets flipped to True in main() when a template is detected.
# In template mode the build functions update values only - your custom
# formatting, colours, column widths etc. are left completely untouched.
_TEMPLATE_MODE = False
_OFFLINE_MODE  = False


# ── COLOURS - edit colorz!!! ──────────────────────────────────────────────────
# (NOT the traffic-light signal ones at the top - those are sacred and reserved
#  exclusively for the expansionary / neutral / contractionary indicators)
C = {
    # Signal colours - green / amber / red. Don't touch these.
    "exp":       "1E8449",
    "neu":       "D4870A",
    "con":       "B03A2E",

    # Section header colours - safe to tweak here or just edit the template
    "jm":        "8A2BD0",   # medium purple  - Job Market
    "inf":       "A957E8",   # lilac purple   - Inflation
    "ea":        "C98CFF",   # light purple   - Economic Activities

    # UI chrome
    "title":     "640394",
    "sub":       "640394",
    "col_hdr":   "640394",
    "chart_hdr": "640394",

    # Row backgrounds
    "white":     "FFFFFF",
    "row0":      "FFFFFF",
    "row1":      "F2F3F4",
    "thresh_bg": "EEF2F7",
    "thresh_fc": "1A252F",

    # Pending cells - shows when a PDF hasn't been dropped in yet
    "input":     "FEF9E7",
    "input_fc":  "6E4800",

    # Text
    "dgray":     "4A4A4A",
    "lgray":     "E8EAED",
    "note_bg":   "F8F9FA",

    # Chart line colours - one per series, edit freely
    "ch_ur":     "2980B9",
    "ch_ic":     "E67E22",
    "ch_nfp":    "27AE60",
    "ch_cpi":    "C0392B",
    "ch_ccpi":   "8E44AD",
    "ch_ppi":    "7F8C8D",
    "ch_imfg":   "1ABC9C",
    "ch_isvc":   "F39C12",
    "ch_chi":    "2C3E50",
    "ch_cc":     "3498DB",
    "ch_ff":     "E74C3C",
    "ch_crb":    "D35400",   # burnt-orange - commodity index
    "ch_yc":     "1A5276",   # navy   - yield curve spread (10Y-2Y)
    # Positioning chart colours
    "ch_vix":    "C0392B",
    "ch_nfci":   "2980B9",
    "ch_fsi":    "8E44AD",
    "ch_lend":   "E67E22",

    # Leading Indicators category + chart colours
    "li":        "0E6655",   # dark teal - Leading Indicators
    "ch_hy":     "E74C3C",   # red - HY credit spread
    "ch_ig":     "CB4335",   # dark red - IG credit spread
    "ch_t3m":    "117864",   # dark teal - 10Y−3M yield curve
    "ch_bp":     "229954",   # green - building permits
    "ch_jo":     "2E86C1",   # blue - JOLTS openings
    "ch_jq":     "1F618D",   # teal - JOLTS quits
    "ch_be":     "D68910",   # gold - breakeven inflation
    "ch_cfnai":  "6C3483",   # purple - CFNAI
    "ch_epu":    "C62828",   # dark maroon - policy uncertainty
    "ch_adp":    "16A085",   # sea green - ADP employment
    "ch_ahe":    "45B39D",   # jade green - average hourly earnings
    "ch_spg":    "AF601A",   # burnt sienna - S&P US Global Services PMI
    "ch_pce":    "9B59B6",   # amethyst - core PCE
    "ch_ip":     "5D6D7E",   # slate - industrial production
    "ch_ret":    "F1948A",   # salmon - retail sales
    "ch_imp":    "D4AC0D",   # gold - import prices
    "ch_dxy":    "1A5276",   # navy - trade-weighted dollar
    "ch_hs":     "48C9B0",   # mint - housing starts

    # Misc UI
    "warn":      "F39C12",   # amber warning (HMM transition risk)
    "ret_pos":   "52BE80",   # light green for positive returns 0-5%
    "link":      "1155CC",   # blue hyperlink text
    "stabilise": "2471A3",   # blue - cautiously positive forecast
}

CAT_COLOR = {
    "Job Market":          C["jm"],
    "Inflation":           C["inf"],
    "Economic Activities": C["ea"],
    "Leading Indicators":  C["li"],
}
CAT_TAB = {
    "Job Market":          "7A1DB8",
    "Inflation":           "A957E8",
    "Economic Activities": "C98CFF",
    "Leading Indicators":  "148F77",
}

MONTH_MAP = dict(
    january="01",  february="02", march="03",    april="04",
    may="05",      june="06",     july="07",     august="08",
    september="09",october="10",  november="11", december="12",
)

# Keys whose values come from PDFs rather than FRED
PDF_KEYS = {"ism_mfg", "ism_svc", "chicago_pmi"}

# Y-axis number format per metric - used in charts
CHART_Y_FMT = {
    "unemployment_rate": "0.0",
    "initial_claims":    "#,##0",
    "nonfarm_payrolls":  '+#,##0;-#,##0',
    "cpi_yoy":           "0.00",
    "core_cpi_yoy":      "0.00",
    "ppi_mom":           "0.00",
    "ism_mfg":           "0.0",
    "ism_svc":           "0.0",
    "chicago_pmi":       "0.0",
    "consumer_conf":     "0.0",
    "fedfunds":          "0.00",
    "crb_yoy":           "0.00",
    "vix":               "0.0",
    "nfci":              "0.000",
    "stlfsi":            "0.000",
    "lending":           "0.0",
    "yield_curve":       "0.00",
    # Leading indicators
    "hy_spread":         "0.0",
    "ig_spread":         "0.0",
    "t10y3m":            "0.00",
    "building_permits":  "+#,##0.0;-#,##0.0",
    "jolts_openings":    "+#,##0.0;-#,##0.0",
    "jolts_quits":       "0.0",
    "breakeven_10y":     "0.00",
    "cfnai":             "0.00",
    "epu":               "0.0",
}


# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
# Shortcuts so I don't have to type PatternFill(...) every five seconds

_fill  = lambda h: PatternFill(start_color=h, end_color=h, fill_type="solid")
_font  = lambda bold=False, color="000000", size=10, italic=False: \
             Font(bold=bold, color=color, size=size, italic=italic, name=FONT_NAME)
_align = lambda h="center", v="center", wrap=False: \
             Alignment(horizontal=h, vertical=v, wrap_text=wrap)
_S     = lambda s: Side(style=s)

def _border_all(s="thin"):
    b = _S(s)
    return Border(left=b, right=b, top=b, bottom=b)

def box_border(ws, r1, c1, r2, c2, weight="medium"):
    if _TEMPLATE_MODE: return
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = Border(
                left   = _S(weight) if c == c1 else _S("thin"),
                right  = _S(weight) if c == c2 else _S("thin"),
                top    = _S(weight) if r == r1 else _S("thin"),
                bottom = _S(weight) if r == r2 else _S("thin"),
            )

def W(ws, row, col, value=None, bg=None, fc="000000", bold=False, size=10,
      italic=False, h="center", v="center", wrap=False, fmt=None, bs="thin"):
    c = ws.cell(row=row, column=col, value=value)
    if not _TEMPLATE_MODE:
        c.font      = _font(bold=bold, color=fc, size=size, italic=italic)
        c.alignment = _align(h=h, v=v, wrap=wrap)
        if bg:  c.fill = _fill(bg)
        c.border = _border_all(bs)
    if fmt: c.number_format = fmt
    return c

def M(ws, r1, c1, r2, c2, value=None, bg=None, fc="000000",
      bold=False, size=10, h="center", v="center", wrap=False, italic=False):
    if not _TEMPLATE_MODE:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if not _TEMPLATE_MODE:
        c.font      = _font(bold=bold, color=fc, size=size, italic=italic)
        c.alignment = _align(h=h, v=v, wrap=wrap)
        if bg: c.fill = _fill(bg)
        box_border(ws, r1, c1, r2, c2)
    return c

def signal_M(ws, r1, c1, r2, c2, value=None, bg=None, fc="000000",
             bold=False, size=10, h="center", v="center", italic=False):
    """Like M() but always applies styling - signal colour IS the data, can't skip it."""
    if not _TEMPLATE_MODE:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font      = _font(bold=bold, color=fc, size=size, italic=italic)
    c.alignment = _align(h=h, v=v)
    if bg: c.fill = _fill(bg)
    if not _TEMPLATE_MODE:
        box_border(ws, r1, c1, r2, c2)
    return c

def set_widths(ws, widths):
    if _TEMPLATE_MODE: return   # preserve the user's column widths
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def rh(ws, row, height):
    if _TEMPLATE_MODE: return   # preserve the user's row heights
    ws.row_dimensions[row].height = height

def p_bg(p):
    return {
        "expansionary": C["exp"],  "neutral":   C["neu"],  "contractionary": C["con"],
        "mixed_con":    C["neu"],  "mixed_exp": C["neu"],  "mixed":          C["neu"],
    }.get(p, C["neu"])

def p_lbl(p):
    return {
        "expansionary": "EXPANSIONARY",  "neutral": "NEUTRAL",  "contractionary": "CONTRACTIONARY",
        "mixed_con":    "MIXED (CONTRACTIONARY)",
        "mixed_exp":    "MIXED (EXPANSIONARY)",
        "mixed":        "MIXED",
    }.get(p, "NEUTRAL")

def p_arr(p):
    return {
        "expansionary": "▲  GROWTH",  "neutral": "◆  STABLE",  "contractionary": "▼  CAUTION",
        "mixed_con": "◆  MIXED",  "mixed_exp": "◆  MIXED",  "mixed": "◆  MIXED",
    }.get(p, "◆  STABLE")

def pressure_cell(ws, row, col, p, pending=False, size=9):
    if pending:
        c = ws.cell(row=row, column=col, value="PENDING")
        c.fill = _fill(C["input"]); c.font = _font(bold=True, color=C["input_fc"], size=size)
        c.alignment = _align(); c.border = _border_all()
        return c
    c = ws.cell(row=row, column=col, value=p_lbl(p))
    c.fill = _fill(p_bg(p)); c.font = _font(bold=True, color="FFFFFF", size=size)
    c.alignment = _align(); c.border = _border_all()
    return c

def signal_cell(ws, row, col, p, pending=False, size=9):
    if pending:
        c = ws.cell(row=row, column=col, value="? PENDING")
        c.fill = _fill(C["input"]); c.font = _font(bold=True, color=C["input_fc"], size=size)
        c.alignment = _align(); c.border = _border_all()
        return c
    c = ws.cell(row=row, column=col, value=p_arr(p))
    c.fill = _fill(p_bg(p)); c.font = _font(bold=True, color="FFFFFF", size=size)
    c.alignment = _align(); c.border = _border_all()
    return c


# ── FRED HELPERS ──────────────────────────────────────────────────────────────

_FRED_MEM = {}   # in-memory: series_id -> (max_limit_fetched, result_list)

def _load_disk_cache():
    try:
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _save_disk_cache(cache):
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f)
    except Exception as e:
        print(f"  [WARN] Cache save failed: {e}")

def _fred(series_id, limit=26):
    if not FRED_API_KEY and not _OFFLINE_MODE:
        print("  [ERROR] No FRED API key set. Get a free key at https://fred.stlouisfed.org/docs/api/api_key.html")
        print("          Set it via: set FRED_API_KEY=your_key_here (Windows) or export FRED_API_KEY=your_key_here (Mac/Linux)")
        return []

    # 1. In-memory cache (same-run dedup: larger fetch serves smaller requests)
    if series_id in _FRED_MEM:
        cached_limit, cached_data = _FRED_MEM[series_id]
        if cached_limit >= limit:
            return cached_data[:limit]

    # 2. Persistent disk cache - keyed by "SERIES|limit"
    cache_key = f"{series_id}|{limit}"
    disk = _load_disk_cache()
    entry = disk.get(cache_key)
    if entry:
        age = (datetime.now() - datetime.fromisoformat(entry["ts"])).total_seconds()
        if age < CACHE_MAX_AGE or _OFFLINE_MODE:
            tag = "CACHE" if age < CACHE_MAX_AGE else "OFFLINE"
            print(f"  [{tag}] {series_id} ({age/3600:.0f}h old)")
            data = [(d, v) for d, v in entry["data"]]
            if series_id not in _FRED_MEM or limit > _FRED_MEM[series_id][0]:
                _FRED_MEM[series_id] = (limit, data)
            return data

    # 3. Offline mode - no network allowed
    if _OFFLINE_MODE:
        print(f"  [OFFLINE] No cache for {series_id} - skipping")
        return []

    # 4. Live fetch from FRED API
    url = "https://api.stlouisfed.org/fred/series/observations"
    try:
        time.sleep(0.15)   # throttle: smooth network traffic, avoid burst
        r = requests.get(url, params=dict(
            series_id=series_id, api_key=FRED_API_KEY,
            file_type="json", sort_order="desc", limit=limit), timeout=20)
        r.raise_for_status()
        result = [(o["date"], float(o["value"]))
                  for o in r.json().get("observations", []) if o["value"] != "."]
    except Exception as e:
        # Network failed - try stale cache as fallback
        if entry:
            print(f"  [WARN] FRED {series_id}: {e} - using stale cache")
            return [(d, v) for d, v in entry["data"]]
        print(f"  [WARN] FRED {series_id}: {e}"); return []

    # 5. Update both caches
    if series_id not in _FRED_MEM or limit > _FRED_MEM[series_id][0]:
        _FRED_MEM[series_id] = (limit, result)
    disk[cache_key] = {"ts": datetime.now().isoformat(), "data": result}
    _save_disk_cache(disk)
    return result


def _fred_vintage(series_id, as_of_date, limit=26):
    """Fetch FRED data as it was known on a specific date (real-time vintage).
    as_of_date: 'YYYY-MM-DD' string. Returns data available as of that date."""
    if not FRED_API_KEY:
        return []
    url = "https://api.stlouisfed.org/fred/series/observations"
    try:
        time.sleep(0.2)
        r = requests.get(url, params=dict(
            series_id=series_id, api_key=FRED_API_KEY,
            file_type="json", sort_order="desc", limit=limit,
            realtime_start=as_of_date, realtime_end=as_of_date), timeout=20)
        r.raise_for_status()
        return [(o["date"], float(o["value"]))
                for o in r.json().get("observations", []) if o["value"] != "."]
    except Exception as e:
        return []

def fetch_latest(sid):
    raw = _fred(sid, 14)
    return (None, []) if not raw else (raw[0][1], raw[:12])

def fetch_yoy(sid):
    raw = _fred(sid, 26)
    if not raw: return None, []
    # Use date-based YoY matching - robust to mid-series gaps in FRED data
    by_ym = {r[0][:7]: r[1] for r in raw}
    hist  = []
    for date, val in raw:
        ym     = date[:7]
        yr_ago = f"{int(ym[:4]) - 1}{ym[4:]}"
        if yr_ago in by_ym and len(hist) < 12:
            hist.append((date, round((val - by_ym[yr_ago]) / by_ym[yr_ago] * 100, 2)))
    if not hist:
        print(f"  [WARN] {sid}: insufficient history for YoY - only {len(raw)} observations")
        return None, []
    return hist[0][1], hist

def fetch_mom_pct(sid):
    raw = _fred(sid, 14)
    if len(raw) < 2: return None, []
    hist = [(raw[i][0], round((raw[i][1] - raw[i+1][1]) / raw[i+1][1] * 100, 2))
            for i in range(min(12, len(raw) - 1))]
    return round((raw[0][1] - raw[1][1]) / raw[1][1] * 100, 2), hist

def fetch_payrolls():
    raw = _fred("PAYEMS", 14)
    if len(raw) < 2: return None, []
    hist = [(raw[i][0], round(raw[i][1] - raw[i+1][1], 1))
            for i in range(min(12, len(raw) - 1))]
    return round(raw[0][1] - raw[1][1], 1), hist

def _scale_series(result, factor):
    """Scale a (value, history) tuple by a constant factor."""
    val, hist = result
    if val is None: return None, []
    return round(val * factor, 2), [(d, round(v * factor, 2)) for d, v in hist]

def fetch_adp():
    """ADP private payrolls MoM change (thousands). ADPMNUSNERSA is in persons, so /1000."""
    raw = _fred("ADPMNUSNERSA", 14)
    if len(raw) < 2: return None, []
    hist = [(raw[i][0], round((raw[i][1] - raw[i+1][1]) / 1000, 1))
            for i in range(min(12, len(raw) - 1))]
    return round((raw[0][1] - raw[1][1]) / 1000, 1), hist

def fetch_fedfunds(months=48):
    raw = _fred("FEDFUNDS", months + 2)
    return raw[:months] if raw else []


def fetch_daily_latest(sid, daily_limit=90):
    """Fetch a daily series, aggregate to monthly, return latest monthly + 12-month history."""
    raw = _fred(sid, daily_limit)
    if not raw: return None, []
    monthly = _daily_to_monthly_mean(raw)
    return (monthly[0][1] if monthly else None, monthly[:12])


def fetch_level_yoy(sid, limit=26):
    """Fetch a monthly level series, compute YoY % change."""
    raw = _fred(sid, limit)
    if not raw: return None, []
    by_ym = {r[0][:7]: r[1] for r in raw}
    hist = []
    for date, val in raw:
        ym = date[:7]
        yr_ago = f"{int(ym[:4]) - 1}{ym[4:]}"
        if yr_ago in by_ym and len(hist) < 12:
            hist.append((date, round((val - by_ym[yr_ago]) / by_ym[yr_ago] * 100, 2)))
    if not hist: return None, []
    return hist[0][1], hist


# ── CSV HISTORY (PMI series) ──────────────────────────────────────────────────

def _read_reuters_commodity():
    """Read Reuters Commodities Index CSV, aggregate daily to monthly, compute YoY %.
    Returns [(YYYY-MM-01, yoy_pct), ...] newest-first, or empty list."""
    csv_path = os.path.join(CSV_FOLDER, "Reuters_Commodities_Index.csv")
    if not os.path.isfile(csv_path):
        return []
    try:
        monthly = {}
        with open(csv_path, newline="", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("Name"):
                    continue
                parts = line.split(",")
                if len(parts) < 2:
                    continue
                try:
                    date_str, val_str = parts[0].strip(), parts[1].strip()
                    val = float(val_str)
                    # Parse DD/MM/YYYY
                    dd, mm, yyyy = date_str.split("/")
                    ym = f"{yyyy}-{mm}"
                    monthly.setdefault(ym, []).append(val)
                except (ValueError, IndexError):
                    continue
        # Monthly means
        monthly_means = {ym: sum(vals)/len(vals) for ym, vals in monthly.items()}
        months = sorted(monthly_means.keys())
        # YoY % change
        yoy = []
        for ym in months:
            yr_ago = f"{int(ym[:4]) - 1}{ym[4:]}"
            if yr_ago in monthly_means:
                pct = (monthly_means[ym] - monthly_means[yr_ago]) / monthly_means[yr_ago] * 100
                yoy.append((f"{ym}-01", round(pct, 2)))
        yoy.sort(key=lambda x: x[0], reverse=True)  # newest first
        return yoy
    except Exception as e:
        print(f"  [WARN] Reuters commodity CSV read failed: {e}")
        return []

def _read_pmi_csv(filepath):
    """Read year/month/value CSV → [(YYYY-MM-01, float), ...] newest-first."""
    rows = []
    try:
        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                try:
                    yr, mo, v = int(r["year"]), int(r["month"]), float(r["value"])
                    rows.append((f"{yr:04d}-{mo:02d}-01", v))
                except (ValueError, KeyError):
                    continue
    except Exception as e:
        print(f"  [WARN] CSV read failed - {os.path.basename(filepath)}: {e}")
    rows.sort(key=lambda x: x[0], reverse=True)
    return rows

def _write_pmi_csv(filepath, date_val_dict):
    """Write {YYYY-MM-01: float} dict to year/month/value CSV."""
    sorted_dates = sorted(date_val_dict.keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        f.write("year,month,value\n")
        for d in sorted_dates:
            yr, mo = int(d[:4]), int(d[5:7])
            f.write(f"{yr},{mo},{date_val_dict[d]}\n")

def update_pmi_csvs(mfg_date, mfg_val, svc_date, svc_val, chi_date=None, chi_val=None,
                    spg_date=None, spg_val=None):
    """Write new PDF-parsed PMI values into the CSV history files.
    Creates the CSV folder and files if they don't exist."""
    os.makedirs(CSV_FOLDER, exist_ok=True)
    updates = [
        (CSV_ISM_MFG,  mfg_date, mfg_val, "ISM Mfg"),
        (CSV_ISM_SVC,  svc_date, svc_val,  "ISM Svc"),
        (CSV_CHICAGO,  chi_date, chi_val,   "Chicago PMI"),
        (CSV_SPG_SVC,  spg_date, spg_val,  "S&P Services"),
    ]
    for csv_path, date, val, label in updates:
        if date is None or val is None:
            continue
        existing = {r[0]: r[1] for r in _read_pmi_csv(csv_path)}
        if date in existing and existing[date] == val:
            continue  # already up to date
        existing[date] = val
        _write_pmi_csv(csv_path, existing)
        print(f"  [CSV] {label}: updated {date[:7]} = {val}")

def import_pmi_csvs():
    """Scan data/import/ for user-provided PMI CSV files, merge with main CSVs."""
    if not os.path.isdir(IMPORT_FOLDER):
        return
    imported_any = False
    mapping = {}   # "ism_mfg" -> [csv_paths]
    for f in sorted(os.listdir(IMPORT_FOLDER)):
        if not f.lower().endswith(".csv"):
            continue
        fl = f.lower()
        if "manuf" in fl and "non" not in fl:
            key = "ism_mfg"
        elif "serv" in fl or "non-manuf" in fl or "non_manuf" in fl or "nonmanuf" in fl:
            key = "ism_svc"
        elif "chicago" in fl or "barometer" in fl or "mni" in fl:
            key = "chicago_pmi"
        elif "spglobal" in fl or "sp_global" in fl or "sp_serv" in fl or "markit" in fl:
            key = "spg_svc"
        else:
            print(f"  [SKIP] Import CSV not recognised (need manuf/serv/chicago/spglobal in name): {f}")
            continue
        mapping.setdefault(key, []).append(os.path.join(IMPORT_FOLDER, f))
    for key, paths in mapping.items():
        target_csv = CSV_PMI_MAP[key]
        existing = {r[0]: r[1] for r in _read_pmi_csv(target_csv)}
        before = len(existing)
        for p in paths:
            for date, val in _read_pmi_csv(p):
                existing[date] = val
        if len(existing) > before:
            _write_pmi_csv(target_csv, existing)
            print(f"  [IMPORT] {key}: merged {len(existing) - before} new months from Import/")
            imported_any = True
    if imported_any:
        print(f"  PMI CSVs updated - charts will use extended history")


# ── PDF PARSING ───────────────────────────────────────────────────────────────

def _pdf_text(path):
    try:
        with pdfplumber.open(path) as pdf:
            return "".join(p.extract_text() or "" for p in pdf.pages[:2])
    except Exception as e:
        print(f"  [WARN] PDF read failed - {os.path.basename(path)}: {e}")
        return ""

def _extract_report_month(txt, fpath=""):
    m = re.search(
        r'(january|february|march|april|may|june|july|august|'
        r'september|october|november|december)\s+(\d{4})',
        txt, re.IGNORECASE)
    if m:
        return f"{m.group(2)}-{MONTH_MAP[m.group(1).lower()]}-01"
    m_fn = re.match(r'(\d{4})(\d{2})\d{2}', os.path.basename(fpath))
    if m_fn:
        return f"{m_fn.group(1)}-{m_fn.group(2)}-01"
    return None

def parse_ism_pdf(path):
    txt = _pdf_text(path)
    if not txt: return None, None
    m = re.search(r'PMI[®\u00ae\u2122]?\s+at\s+(\d+\.?\d*)\s*%', txt, re.IGNORECASE)
    if not m:
        m = re.search(r'PMI[®\u00ae\u2122]?[^\n]{0,40}?at\s+(\d+\.?\d*)', txt, re.IGNORECASE)
    if not m:
        m = re.search(r'registered\s+(\d+\.?\d*)\s+percent', txt, re.IGNORECASE)
    if not m:
        m = re.search(r'PMI[^\n]{0,20}at\s+(\d+\.?\d*)', txt, re.IGNORECASE)
    pmi = float(m.group(1)) if m else None
    return _extract_report_month(txt, path), pmi

def parse_mni_pdf(path):
    txt = _pdf_text(path)
    if not txt: return None, None
    # Four fallback patterns because MNI can't decide how to write the same sentence
    m = re.search(
        r'\bto\s+(\d+\.?\d*)\s+in\s+'
        r'(?:january|february|march|april|may|june|july|august|'
        r'september|october|november|december)\b',
        txt, re.IGNORECASE)
    if not m:
        m = re.search(
            r'(?:climbed|fell|rose|dropped|slipped|edged|moved|eased)\s+'
            r'\d+\.?\d*\s+points?\s+to\s+(\d+\.?\d*)',
            txt, re.IGNORECASE)
    if not m:
        m = re.search(r'Barometer[^.]{0,80}?(\d{2}\.\d)\s*(?:percent\b|in\b)', txt, re.IGNORECASE)
    if not m:
        m = re.search(r'\bat\s+(\d{2}\.\d)\b', txt)
    val = float(m.group(1)) if m else None
    return _extract_report_month(txt, path), val

def parse_spg_pdf(path):
    """Parse S&P Global US Services PMI from press release PDF."""
    txt = _pdf_text(path)
    if not txt: return None, None
    # Pattern 1: "Services PMI ... 49.8" or "Business Activity Index ... 49.8"
    m = re.search(r'(?:Services\s+PMI|Business\s+Activity\s+Index)[^\n]{0,60}?(\d+\.?\d*)', txt, re.IGNORECASE)
    if not m:
        # Pattern 2: "recorded 49.8 in March"
        m = re.search(r'recorded\s+(\d+\.?\d*)\s+in\b', txt, re.IGNORECASE)
    if not m:
        # Pattern 3: "fell to 49.8" / "rose to 54.2"
        m = re.search(r'(?:fell|rose|climbed|dropped|slipped|edged|eased|moved)\s+to\s+(\d+\.?\d*)', txt, re.IGNORECASE)
    if not m:
        # Pattern 4: "at 49.8" with PMI-range value
        m = re.search(r'\bat\s+(\d{2}\.\d)\b', txt)
    val = float(m.group(1)) if m else None
    return _extract_report_month(txt, path), val

def parse_reports_folder():
    mfg_date = mfg_val = svc_date = svc_val = chi_date = chi_val = spg_date = spg_val = None
    if not os.path.isdir(ISM_FOLDER):
        print(f"  [WARN] Reports folder not found: {ISM_FOLDER}")
        return mfg_date, mfg_val, svc_date, svc_val, chi_date, chi_val, spg_date, spg_val

    for fname in sorted(os.listdir(ISM_FOLDER)):
        if not fname.lower().endswith(".pdf"): continue
        fpath = os.path.join(ISM_FOLDER, fname)
        fl    = fname.lower()
        is_mni = "mni" in fl or "barometer" in fl or "chi" in fl
        is_spg = "spglobal" in fl or "sp_global" in fl or "s&p" in fl or "markit" in fl
        is_mfg = (not is_mni) and (not is_spg) and ("manuf" in fl)
        is_svc = (not is_mni) and (not is_spg) and ("ser" in fl or "non-manuf" in fl
                                   or "nonmanuf" in fl or "non_manuf" in fl)
        if is_mni:
            d, v = parse_mni_pdf(fpath)
            if v is not None:
                chi_date, chi_val = d, v
                print(f"  PDF -> Chicago PMI        {v:5.1f}   ({d})")
            else:
                print(f"  [WARN] Chicago PMI extraction failed: {fname}")
        elif is_spg:
            d, v = parse_spg_pdf(fpath)
            if v is not None:
                spg_date, spg_val = d, v
                print(f"  PDF -> S&P Services PMI   {v:5.1f}   ({d})")
            else:
                print(f"  [WARN] S&P Services PMI extraction failed: {fname}")
        elif is_mfg:
            d, v = parse_ism_pdf(fpath)
            if v is not None:
                mfg_date, mfg_val = d, v
                print(f"  PDF -> ISM Mfg PMI        {v:5.1f}   ({d})")
            else:
                print(f"  [WARN] ISM Mfg PMI extraction failed: {fname}")
        elif is_svc:
            d, v = parse_ism_pdf(fpath)
            if v is not None:
                svc_date, svc_val = d, v
                print(f"  PDF -> ISM Services PMI   {v:5.1f}   ({d})")
            else:
                print(f"  [WARN] ISM Svc PMI extraction failed: {fname}")
        else:
            print(f"  [SKIP] Unrecognised PDF (rename to include manuf/serv/mni/spglobal): {fname}")

    return mfg_date, mfg_val, svc_date, svc_val, chi_date, chi_val, spg_date, spg_val


# ── ISM MANUAL INPUT WORKBOOK ─────────────────────────────────────────────────

MI_HDR_ROW  = 4
MI_DATA_ROW = 5

def _mi_col_headers():
    return ["Date (YYYY-MM-DD)", "ISM Manufacturing PMI",
            "ISM Non-Manufacturing PMI", "Chicago PMI",
            "S&P US Global Services PMI", "Source / Notes"]

def create_manual_input_if_needed():
    if os.path.exists(MANUAL_FILE): return
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Monthly PMI Data"
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 18, 22, 22, 18, 24, 44])
    rh(ws, 1, 44); rh(ws, 2, 20); rh(ws, 3, 20); rh(ws, 4, 28)
    M(ws, 1, 2, 1, 7,
      "PMI Monthly Data Log (Auto-populated from PDFs)",
      bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=14, h="left")
    M(ws, 2, 2, 2, 7,
      "This file is read automatically by economic_tracker.py on every run.  "
      "Drop the latest ISM, MNI, and S&P Global PDFs into 'Reports' and re-run.",
      bg=C["lgray"], fc=C["dgray"], size=9, italic=True, h="left")
    M(ws, 3, 2, 3, 7,
      "ISM Mfg: 1st bday  |  ISM Non-Mfg: 3rd bday  |  "
      "Chicago PMI (MNI): last bday  |  S&P Services: late month",
      bg="FFF8DC", fc=C["input_fc"], size=9, italic=True, h="left")
    for ci, hdr in enumerate(_mi_col_headers(), 2):
        W(ws, MI_HDR_ROW, ci, hdr, bg=C["col_hdr"], fc="FFFFFF",
          bold=True, size=9, bs="medium")
    wb.save(MANUAL_FILE)
    print(f"  Created: {MANUAL_FILE}")

def _mi_find_row_for_month(ws, month_str):
    for r in range(MI_DATA_ROW, ws.max_row + 1):
        cell_val = ws.cell(r, 2).value
        if cell_val and str(cell_val)[:7] == month_str[:7]:
            return r
    return None

def update_manual_input(mfg_date, mfg_val, svc_date, svc_val,
                        chi_date=None, chi_val=None, spg_date=None, spg_val=None):
    create_manual_input_if_needed()
    wb = openpyxl.load_workbook(MANUAL_FILE)
    ws = wb["Monthly PMI Data"]
    report_date = mfg_date or svc_date or chi_date or spg_date
    if not report_date:
        print("  No report date - skipping data log update"); wb.close(); return

    month_str    = report_date[:7]
    existing_row = _mi_find_row_for_month(ws, month_str)

    if existing_row:
        def _patch(col, val):
            cell = ws.cell(existing_row, col)
            if val is not None and cell.value in (None, ""):
                cell.value = val
        _patch(3, mfg_val); _patch(4, svc_val); _patch(5, chi_val); _patch(6, spg_val)
        src = ws.cell(existing_row, 7)
        if not src.value: src.value = "PDF auto-extracted"
        wb.save(MANUAL_FILE)
        print(f"  Data log patched  {month_str}  MFG={mfg_val}  SVC={svc_val}  CHI={chi_val}  SPG={spg_val}")
        return

    next_row = MI_DATA_ROW
    for row in ws.iter_rows(min_row=MI_DATA_ROW, min_col=2, max_col=2, values_only=True):
        if row[0] is None: break
        next_row += 1

    row_bg = C["row0"] if (next_row - MI_DATA_ROW) % 2 == 0 else C["row1"]
    vals   = [report_date,
              mfg_val if mfg_val is not None else "",
              svc_val if svc_val is not None else "",
              chi_val if chi_val is not None else "",
              spg_val if spg_val is not None else "",
              "PDF auto-extracted"]
    for ci, v in enumerate(vals, 2):
        c = ws.cell(row=next_row, column=ci, value=v)
        c.fill = _fill(row_bg); c.font = _font(size=9, bold=(ci == 2))
        c.alignment = _align(h="center" if ci != 7 else "left")
        c.border = _border_all()
    ws.row_dimensions[next_row].height = 20
    for r in range(MI_DATA_ROW, next_row + 1):
        stripe = C["row0"] if (r - MI_DATA_ROW) % 2 == 0 else C["row1"]
        for ci in range(2, 8):
            ws.cell(r, ci).fill = _fill(stripe)
    wb.save(MANUAL_FILE)
    print(f"  Data log updated  {month_str}  MFG={mfg_val}  SVC={svc_val}  CHI={chi_val}  SPG={spg_val}")

def read_manual_input():
    empty = {"ism_mfg": (None, []), "ism_svc": (None, []), "chicago_pmi": (None, [])}
    if not os.path.exists(MANUAL_FILE): return empty
    try:
        wb = openpyxl.load_workbook(MANUAL_FILE, data_only=True)
        ws = wb["Monthly PMI Data"]
    except Exception as e:
        print(f"  [WARN] Manual input read failed: {e}"); return empty
    rows = []
    for row in ws.iter_rows(min_row=MI_DATA_ROW, min_col=2, max_col=6, values_only=True):
        date, mfg, svc, chi, *_ = row
        if not date: continue
        rows.append((str(date)[:10], mfg, svc, chi))
    rows.sort(key=lambda x: x[0], reverse=True)
    def _series(idx):
        vals = [(r[0], float(r[idx])) for r in rows if r[idx] not in (None, "", "N/A")]
        return (vals[0][1] if vals else None), vals[:12]
    return {"ism_mfg": _series(1), "ism_svc": _series(2), "chicago_pmi": _series(3)}


# ── CLASSIFICATION ────────────────────────────────────────────────────────────
# Thresholds are conventional - widely used by economists, not plucked from thin air.
# Change them here if you want different trigger points.

RULES = {
    "unemployment_rate": [
        (lambda v: v > 6.5, "contractionary", "Weak economy - elevated unemployment, above 6.5%",          "Expansionary - Cut rates / QE"),
        (lambda v: v < 4.0, "expansionary",   "Strong economy - tight labour market, below 4%",             "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "Moderate unemployment - labour market within historical norms (4-6.5%)",  "Neutral - Hold rates"),
    ],
    "initial_claims": [
        (lambda v: v > 350, "contractionary", "Elevated layoffs signal a weakening labour market",          "Expansionary - Cut rates / QE"),
        (lambda v: v < 250, "expansionary",   "Low jobless claims reflect a healthy labour market",          "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "Claims within the normal 250-350k range",                    "Neutral - Hold rates"),
    ],
    "nonfarm_payrolls": [
        (lambda v: v > 250, "expansionary",   "Robust hiring - payrolls growing well above trend",            "Contractionary - Hike rates / QT"),
        (lambda v: v < 0,   "contractionary", "Net job losses - economy actively shedding workers",           "Expansionary - Cut rates / QE"),
        (lambda v: v < 50,  "contractionary", "Near-stall job creation - economy losing momentum",            "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Moderate job creation within the 50-250k range",              "Neutral - Hold rates"),
    ],
    "adp": [
        (lambda v: v > 200, "expansionary",   "Strong private hiring - ADP well above trend",                 "Contractionary - Hike rates / QT"),
        (lambda v: v < 0,   "contractionary", "Net private job losses - economy shedding workers",            "Expansionary - Cut rates / QE"),
        (lambda v: v < 50,  "contractionary", "Near-stall private hiring - economy losing momentum",          "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Moderate private hiring within the 50-200k range",            "Neutral - Hold rates"),
    ],
    "ahe_yoy": [
        (lambda v: v > 4.0, "contractionary", "Wage growth above 4% - wage-price spiral risk, Fed will tighten",  "Contractionary - Hike rates / QT"),
        (lambda v: v < 2.0, "contractionary", "Wage growth below 2% - labour market slack, weak consumer spending", "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Wage growth in the 2-4% range - consistent with stable inflation", "Neutral - Hold rates"),
    ],
    "cpi_yoy": [
        (lambda v: v > 2.5, "contractionary", "Inflation well above the 2% target - purchasing power eroding", "Contractionary - Hike rates / QT"),
        (lambda v: v < 1.5, "expansionary",   "Below-target inflation - room to stimulate growth",             "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Inflation near the 2% target (1.5-2.5% range)",                "Neutral - Hold rates"),
    ],
    "core_cpi_yoy": [
        (lambda v: v > 2.5, "contractionary", "Core inflation well above target - sticky price pressures",   "Contractionary - Hike rates / QT"),
        (lambda v: v < 1.8, "expansionary",   "Core inflation below target - deflationary risk",             "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Core inflation near the 2% target (1.8-2.5% range)",         "Neutral - Hold rates"),
    ],
    "core_pce_yoy": [
        (lambda v: v > 2.5, "contractionary", "Core PCE above target - the Fed's preferred gauge is elevated",  "Contractionary - Hike rates / QT"),
        (lambda v: v < 1.8, "expansionary",   "Core PCE below target - room for policy easing",                 "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Core PCE near the 2% target (1.8-2.5% range)",                  "Neutral - Hold rates"),
    ],
    "ppi_mom": [
        (lambda v: v > 0.5, "contractionary", "Elevated producer prices likely to feed through to CPI",      "Contractionary - Hike rates / QT"),
        (lambda v: v < 0.0, "neutral",         "Falling producer prices - benign if growth holds, watch for demand weakness", "Neutral - Hold rates"),
        (lambda v: True,    "neutral",         "Producer price growth contained within 0-0.5% MoM",          "Neutral - Hold rates"),
    ],
    "crb_yoy": [
        (lambda v: v > 10,  "contractionary", "Commodity prices surging - inflationary pipeline building",       "Contractionary - Hike rates / QT"),
        (lambda v: v < -10, "contractionary", "Commodity prices falling sharply - deflationary demand signal",   "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Commodity price growth within normal range (-10% to +10%)",      "Neutral - Hold rates"),
    ],
    "ism_mfg": [
        (lambda v: v > 56, "expansionary",   "Strong manufacturing expansion - reading above 56",            "Contractionary - Hike rates / QT"),
        (lambda v: v < 50, "contractionary", "Manufacturing sector in contraction - reading below 50",       "Expansionary - Cut rates / QE"),
        (lambda v: True,   "neutral",         "Moderate manufacturing expansion - reading between 50 and 56","Neutral - Hold rates"),
    ],
    "ism_svc": [
        (lambda v: v > 56, "expansionary",   "Strong services expansion - reading above 56",                 "Contractionary - Hike rates / QT"),
        (lambda v: v < 50, "contractionary", "Services sector in contraction - reading below 50",            "Expansionary - Cut rates / QE"),
        (lambda v: True,   "neutral",         "Moderate services expansion - reading between 50 and 56",     "Neutral - Hold rates"),
    ],
    "spg_svc": [
        (lambda v: v > 56, "expansionary",   "Strong services expansion - S&P US Global PMI above 56",         "Contractionary - Hike rates / QT"),
        (lambda v: v < 50, "contractionary", "Services contraction - S&P US Global PMI below 50",              "Expansionary - Cut rates / QE"),
        (lambda v: True,   "neutral",         "Moderate services expansion - S&P US Global PMI 50-56",         "Neutral - Hold rates"),
    ],
    "chicago_pmi": [
        (lambda v: v > 56, "expansionary",   "Strong Chicago-area business activity - reading above 56",     "Contractionary - Hike rates / QT"),
        (lambda v: v < 50, "contractionary", "Chicago-area business activity contracting - below 50",        "Expansionary - Cut rates / QE"),
        (lambda v: True,   "neutral",         "Moderate Chicago-area activity - reading between 50 and 56",  "Neutral - Hold rates"),
    ],
    "consumer_conf": [
        (lambda v: v > 90,  "expansionary",   "Consumer sentiment elevated - households optimistic, spending supported",  "Contractionary - Hike rates / QT"),
        (lambda v: v < 57,  "contractionary", "Consumer sentiment deeply depressed - households stressed, spending at risk", "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Consumer sentiment within the normal 65-90 range",                        "Neutral - Hold rates"),
    ],
    "indpro_mom": [
        (lambda v: v > 0.3,  "expansionary",   "Industrial output expanding - factories ramping up",            "Contractionary - Hike rates / QT"),
        (lambda v: v < -0.2, "contractionary", "Industrial output contracting - manufacturing in decline",      "Expansionary - Cut rates / QE"),
        (lambda v: True,     "neutral",         "Industrial production flat to marginally positive",             "Neutral - Hold rates"),
    ],
    "retail_mom": [
        (lambda v: v > 0.5,  "expansionary",   "Strong consumer spending - retail sales surging",               "Contractionary - Hike rates / QT"),
        (lambda v: v < -0.3, "contractionary", "Consumer spending falling - retail sales declining",            "Expansionary - Cut rates / QE"),
        (lambda v: True,     "neutral",         "Retail sales within normal monthly range",                     "Neutral - Hold rates"),
    ],
    "import_yoy": [
        (lambda v: v > 5.0,  "contractionary", "Import prices surging - tariff cost pass-through to consumers", "Contractionary - Hike rates / QT"),
        (lambda v: v < -5.0, "expansionary",   "Falling import prices - deflationary, supports purchasing power","Expansionary - Cut rates / QE"),
        (lambda v: True,     "neutral",         "Import prices within normal range",                            "Neutral - Hold rates"),
    ],
    # ── Leading Indicators ─────────────────────────────────────────────────
    # NOTE: FRED returns OAS in percentage points (e.g. 3.18 = 318bp). Thresholds below are in pct pts.
    "hy_spread": [
        (lambda v: v > 5.0, "contractionary", "HY spreads >500bp - credit stress elevated, risk-off environment",            "Expansionary - Cut rates / QE"),
        (lambda v: v < 3.5, "expansionary",   "HY spreads tight - strong risk appetite, credit freely available",            "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "HY spreads within the normal 350-500bp range",                               "Neutral - Hold rates"),
    ],
    "ig_spread": [
        (lambda v: v > 1.5, "contractionary", "IG spreads >150bp - investment-grade credit under stress",                    "Expansionary - Cut rates / QE"),
        (lambda v: v < 1.0, "expansionary",   "IG spreads tight - benign credit conditions",                                "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "IG spreads within the normal 100-150bp range",                               "Neutral - Hold rates"),
    ],
    "t10y3m": [
        (lambda v: v < 0,   "contractionary", "Yield curve inverted - historically precedes recession by 12-18 months",     "Expansionary - Cut rates / QE"),
        (lambda v: v > 1.0, "expansionary",   "Steep yield curve - growth-supportive, positive term premium",               "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "Yield curve flat to mildly positive - transition zone",                      "Neutral - Hold rates"),
    ],
    "building_permits": [
        (lambda v: v < -10, "contractionary", "Building permits falling >10% YoY - housing cycle turning down",             "Expansionary - Cut rates / QE"),
        (lambda v: v > 5,   "expansionary",   "Building permits rising >5% YoY - construction cycle expanding",             "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "Building permits growth within normal range (-10% to +5%)",                  "Neutral - Hold rates"),
    ],
    "jolts_openings": [
        (lambda v: v < -15, "contractionary", "Job openings falling >15% YoY - labour demand weakening sharply",            "Expansionary - Cut rates / QE"),
        (lambda v: v > 10,  "expansionary",   "Job openings rising >10% YoY - strong demand for workers",                   "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "Job openings growth within normal range",                                    "Neutral - Hold rates"),
    ],
    "jolts_quits": [
        (lambda v: v < 2.0, "contractionary", "Quits rate below 2% - workers lack confidence, labour market cooling",       "Expansionary - Cut rates / QE"),
        (lambda v: v > 2.5, "expansionary",   "Quits rate above 2.5% - workers confident, voluntary turnover elevated",     "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "Quits rate within the normal 2.0-2.5% range",                                "Neutral - Hold rates"),
    ],
    "breakeven_10y": [
        (lambda v: v > 2.5, "contractionary", "Breakeven inflation >2.5% - market expects above-target inflation",          "Contractionary - Hike rates / QT"),
        (lambda v: v < 1.8, "contractionary", "Breakeven inflation <1.8% - market expects below-target inflation (deflation risk)", "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Breakeven inflation near 2% target - inflation expectations anchored",       "Neutral - Hold rates"),
    ],
    "cfnai": [
        (lambda v: v < -0.7,"contractionary", "CFNAI below -0.7 - recession signal from 85-indicator composite",            "Expansionary - Cut rates / QE"),
        (lambda v: v > 0.2, "expansionary",   "CFNAI above +0.2 - economy growing above long-run trend",                   "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "CFNAI near zero - economy growing at or near trend",                         "Neutral - Hold rates"),
    ],
    "epu": [
        (lambda v: v > 300, "contractionary", "Extreme policy uncertainty - crisis-level (historically: GFC, COVID, trade wars)", "Expansionary - Cut rates / QE"),
        (lambda v: v > 200, "contractionary", "Elevated policy uncertainty - growth headwind from political/trade risk",         "Neutral - Hold rates"),
        (lambda v: v < 120, "expansionary",   "Low policy uncertainty - stable policy environment supports growth",              "Contractionary - Hike rates / QT"),
        (lambda v: True,    "neutral",         "Policy uncertainty within normal range (80-130)",                                "Neutral - Hold rates"),
    ],
    "dxy": [
        (lambda v: v > 120, "contractionary", "Strong dollar - compresses exports, tightens global financial conditions",   "Neutral - Hold rates"),
        (lambda v: v < 100, "expansionary",   "Weak dollar - supports exports and commodity prices",                       "Neutral - Hold rates"),
        (lambda v: True,    "neutral",         "Dollar within normal trade-weighted range (100-120)",                       "Neutral - Hold rates"),
    ],
    "housing_starts": [
        (lambda v: v > 10,  "expansionary",   "Housing construction accelerating - residential investment expanding",       "Contractionary - Hike rates / QT"),
        (lambda v: v < -15, "contractionary", "Housing construction falling sharply - rate-sensitive sector under stress",  "Expansionary - Cut rates / QE"),
        (lambda v: True,    "neutral",         "Housing starts within normal range",                                       "Neutral - Hold rates"),
    ],
}

METRIC_TOOLTIPS = {
    "unemployment_rate": (
        "Broadest measure of labour market health. Rising unemployment cuts consumer "
        "spending (70% of GDP) and feeds through to housing and credit.\n\n"
        "Thresholds\n"
        "< 4.0%: Below the CBO natural rate (~4.4%). Tight - wage pressure risk.\n"
        "4.0 - 6.5%: Normal cyclical range.\n"
        "> 6.5%: Bernanke's forward guidance threshold. Every reading above 6.5% "
        "since 1948 has coincided with recession.\n\n"
        "Note: Lagging - rises 3-6 months after recession starts. Misses U-6 underemployment."
    ),
    "initial_claims": (
        "Timeliest labour market indicator - reported weekly. First signal of layoffs "
        "spreading through the economy.\n\n"
        "< 250k: Healthy labour market. 250-350k: Normal churn. > 350k: Every sustained "
        "move above this since 1967 preceded recession.\n\n"
        "Watch out: Weekly noise is large. We use the 4-week moving average (IC4WSA). "
        "Seasonal adjustment can distort around holidays."
    ),
    "nonfarm_payrolls": (
        "The headline jobs number from BLS. Positive = net hiring, negative = economy "
        "shedding workers. The ~100k monthly pace is roughly breakeven for population growth.\n\n"
        "> 250k: Strong. 50-250k: Moderate. < 50k: Near-stall territory.\n\n"
        "Subject to large revisions (birth-death model). Single months can be distorted "
        "by strikes, weather, or census hiring."
    ),
    "adp": (
        "ADP processes payroll for ~25% of US private workers and reports 1-2 days before "
        "BLS. Acts as an early read on NFP direction.\n\n"
        "> 200k: Strong private hiring. 50-200k: Moderate. < 50k: Near-stall.\n\n"
        "Private sector only (no government). Can diverge from BLS due to different "
        "methodology. Revised with each release."
    ),
    "ahe_yoy": (
        "Average hourly earnings for all private-sector employees. The key link between the "
        "labour market and inflation - when wages accelerate above ~4%, firms pass costs to "
        "consumers, risking a wage-price spiral.\n\n"
        "Thresholds\n"
        "> 4.0% (Contractionary): Wage-price spiral territory. The Fed will tighten.\n"
        "2.0 - 4.0% (Neutral): Healthy wage growth consistent with the 2% inflation target.\n"
        "< 2.0% (Contractionary): Slack in the labour market, weak consumer spending ahead.\n\n"
        "Note: Released alongside NFP on the first Friday. Subject to revisions and "
        "compositional effects (if low-wage workers lose jobs, average wages rise mechanically)."
    ),
    "cpi_yoy": (
        "Primary consumer inflation gauge. Persistent above-target readings erode real wages "
        "and force the Fed to tighten.\n\n"
        "> 2.5%: Above target, hawkish. 1.5-2.5%: Near target. < 1.5%: Room to ease.\n\n"
        "Lagging - reflects prices already paid. The shelter component (35% of CPI) lags "
        "actual rents by 12+ months, which can mask turning points."
    ),
    "core_cpi_yoy": (
        "CPI ex food and energy - strips out volatile components to reveal the underlying "
        "trend. The Fed watches core more than headline for policy decisions.\n\n"
        "> 2.5%: Sticky pressures. 1.8-2.5%: Near target. < 1.8%: Deflation risk.\n\n"
        "Owner's Equivalent Rent dominates the index. Excludes energy costs which still "
        "matter for businesses and transport."
    ),
    "core_pce_yoy": (
        "The Fed's actual inflation target. Core PCE (Personal Consumption Expenditures "
        "excluding food and energy) typically runs 30-50bp below CPI because it accounts "
        "for consumer substitution and has broader coverage including employer-paid healthcare.\n\n"
        "When CPI reads 2.7%, Core PCE might read 2.3% - the policy implication is very different. "
        "This is the number the FOMC references in its statement and dot plot.\n\n"
        "Thresholds\n"
        "> 2.5%: Above target, Fed will lean hawkish.\n"
        "1.8 - 2.5%: Near the 2% target.\n"
        "< 1.8%: Below target, room to ease.\n\n"
        "Note: Released late month by BEA, ~2 weeks after CPI. Same month, different methodology."
    ),
    "ppi_mom": (
        "Producer prices - what factories pay for inputs. PPI feeds through to CPI with a "
        "3-6 month lag, making it a leading indicator of consumer inflation.\n\n"
        "> 0.5% MoM: Pipeline pressure building. 0-0.5%: Normal. < 0%: Falling input costs "
        "(benign if growth is holding, worrying if demand is collapsing).\n\n"
        "Commodity-heavy - oil swings can dominate single months."
    ),
    "crb_yoy": (
        "Broad commodity basket (IMF All-Commodity Index as CRB proxy). Early signal of "
        "global demand and inflationary pipeline pressure.\n\n"
        "> +10% YoY: Inflationary squeeze on margins. < -10%: Demand collapsing.\n\n"
        "Energy-weighted. Also reflects supply disruptions (OPEC, weather) not just demand."
    ),
    "ism_mfg": (
        "Diffusion index surveying 300+ purchasing managers. 50 = no change, above = expansion. "
        "Leads GDP by 1-2 quarters.\n\n"
        "> 56: Strong - correlates with >3% GDP growth. 50-56: Moderate expansion. "
        "< 50: Contraction - preceded 7 of last 8 recessions.\n\n"
        "Manufacturing is only ~11% of GDP now. Services matter more, but this still moves markets."
    ),
    "ism_svc": (
        "Covers the ~89% of the economy that manufacturing misses. A reading below 50 is "
        "rare and historically serious - services contraction almost always means recession.\n\n"
        "> 56: Strong. 50-56: Moderate. < 50: Contraction.\n\n"
        "Shorter history than ISM Manufacturing (since 1997). New orders sub-index is the "
        "most forward-looking component."
    ),
    "spg_svc": (
        "S&P Global's services PMI - surveys ~400 US companies. Released 1-2 days before "
        "ISM Services, so acts as a preview of the dominant sector.\n\n"
        "> 56: Strong. 50-56: Moderate. < 50: Contraction.\n\n"
        "Not on FRED - requires manual CSV or PDF update. Can diverge from ISM due to "
        "different sample and methodology. History only from 2009."
    ),
    "chicago_pmi": (
        "Regional manufacturing PMI for the Chicago/Midwest area. Released the day before "
        "ISM Mfg, so traders use it as a preview.\n\n"
        "Same 50-based diffusion index as ISM. Regional only - carries less weight than "
        "the national ISM in the overall signal."
    ),
    "consumer_conf": (
        "University of Michigan sentiment survey. Consumer spending is 70% of GDP so confidence "
        "matters, but sentiment has structurally shifted down post-COVID - don't over-read it.\n\n"
        "> 90: Strong. 57-90: Normal. < 57: Deeply depressed.\n\n"
        "Survey-based - can diverge from actual spending data (retail sales is the hard measure)."
    ),
    "hy_spread": (
        "Risk premium on junk bonds over Treasuries. Widening = investors fleeing to safety. "
        "One of the best leading indicators - widens before every recession.\n\n"
        "> 5%: Credit stress (GFC peaked at ~20%). 3.5-5%: Normal. < 3.5%: Very tight, "
        "strong risk appetite.\n\n"
        "Can stay tight for years before snapping wider. The speed of widening matters as "
        "much as the level."
    ),
    "ig_spread": (
        "Investment-grade corporate bond spreads. Less noisy than HY - filters out "
        "junk-specific risk to show broad credit health.\n\n"
        "> 1.5%: Stress. 1.0-1.5%: Normal. < 1.0%: Benign.\n\n"
        "Tends to lag HY in detecting stress but is a cleaner signal of systemic risk."
    ),
    "t10y3m": (
        "10-year minus 3-month Treasury spread. The single most reliable recession predictor - "
        "every US recession since 1960 was preceded by inversion (< 0). Zero false positives.\n\n"
        "< 0: Inverted, recession typically follows in 12-18 months. 0-1%: Flat. > 1%: Steep.\n\n"
        "Lead time is long and variable (6-24 months). Was inverted for ~2 years in 2022-24."
    ),
    "building_permits": (
        "Housing permits predict construction activity 6-12 months ahead. Housing is the most "
        "rate-sensitive sector - it turns first in every cycle.\n\n"
        "< -10% YoY: Cycle turning down (predicted 8 of 9 recessions). > +5%: Expanding.\n\n"
        "Noisy month-to-month. Regional variation is large."
    ),
    "jolts_openings": (
        "Job openings = labour demand. Falls before unemployment rises, making it a genuine "
        "leading indicator of whether employers are pulling back.\n\n"
        "< -15% YoY: Sharp decline. > +10%: Strong demand.\n\n"
        "2-month reporting lag. Some openings are 'phantom' postings that never get filled."
    ),
    "jolts_quits": (
        "Quits rate = worker confidence. People quit when they're sure they'll find something "
        "better. Falls before recessions.\n\n"
        "< 2.0%: Workers afraid to leave. 2.0-2.5%: Normal. > 2.5%: Confident workers.\n\n"
        "Same 2-month lag as JOLTS openings. Normalised sharply from the 2022 'Great Resignation' peak."
    ),
    "breakeven_10y": (
        "Market-implied inflation expectations (TIPS vs nominal Treasuries). Forward-looking "
        "unlike CPI - if markets expect inflation to persist, it usually does.\n\n"
        "Goldilocks indicator: 1.8-2.5% is anchored near target. Both extremes are bad "
        "(> 2.5% = inflation risk, < 1.8% = deflation risk).\n\n"
        "Affected by TIPS liquidity premium, so not pure inflation expectations."
    ),
    "cfnai": (
        "Weighted average of 85 monthly indicators from the Chicago Fed. The single best "
        "real-time GDP nowcast. Zero = long-run trend growth.\n\n"
        "< -0.7: Recession signal (preceded every recession since 1967). > +0.2: Above trend.\n\n"
        "Heavily revised. Monthly noise can be large - the 3-month moving average is more reliable."
    ),
    "epu": (
        "Policy uncertainty index from newspaper coverage + tax code expirations + forecaster "
        "disagreement. High uncertainty is a growth headwind - firms delay capex.\n\n"
        "> 300: Extreme (GFC, COVID, trade wars). 200-300: Elevated. < 120: Low/stable.\n\n"
        "News-driven - can spike on media attention rather than real policy change. "
        "Thresholds raised from the original Baker-Bloom-Davis levels to reflect the post-2016 normal."
    ),
    "indpro_mom": (
        "One of the 4 series the NBER uses to date recessions. Covers manufacturing, "
        "mining, and utilities - about 15% of GDP but highly cyclical.\n\n"
        "Thresholds\n"
        "> 0.3% MoM: Expanding. < -0.2% MoM: Contracting.\n\n"
        "Note: Monthly data, released ~2 weeks after month-end. Subject to revisions."
    ),
    "retail_mom": (
        "Hard spending data - what consumers actually bought, not how they feel. "
        "Consumption is 70% of GDP so this matters more than sentiment surveys.\n\n"
        "Thresholds\n"
        "> 0.5% MoM: Strong. < -0.3% MoM: Weak.\n\n"
        "Watch out: Nominal (not inflation-adjusted). A positive MoM could just be price increases."
    ),
    "import_yoy": (
        "Direct measurement of tariff cost pass-through. When tariffs hit, import prices "
        "rise and the cost flows through to CPI with a 3-6 month lag. In the current tariff "
        "environment this is one of the most informative inflation indicators.\n\n"
        "Thresholds\n"
        "> 5% YoY: Inflationary (contractionary for purchasing power).\n"
        "< -5% YoY: Deflationary (expansionary for consumers)."
    ),
    "dxy": (
        "The trade-weighted dollar index. A strong dollar tightens financial conditions "
        "globally, hurts US exporters, and compresses commodity prices. A weak dollar does "
        "the opposite.\n\n"
        "Thresholds\n"
        "> 120: Contractionary (too strong). < 100: Expansionary (weak, supports exports).\n\n"
        "Note: This is a daily series aggregated to monthly means."
    ),
    "housing_starts": (
        "Actual construction of new homes - the sharpest interest-rate-sensitive sector. "
        "Housing leads the broader economy by 6-12 months because it drives employment, "
        "materials demand, and consumer durables spending.\n\n"
        "Thresholds\n"
        "> 10% YoY: Expansionary. < -15% YoY: Contractionary.\n\n"
        "Note: Complements building permits (which lead starts by 1-2 months)."
    ),
}

def classify(key, val):
    if val is None:
        msg = ("No data - add PDF to 'Reports' folder" if key in PDF_KEYS
               else "No data - check FRED API connection or data availability")
        return "neutral", msg, "Neutral - Hold rates"
    for cond, pressure, impact, fed in RULES.get(key, []):
        if cond(val): return pressure, impact, fed
    return "neutral", "N/A", "N/A"

def overall_signal(pressures):
    known = [p for p in pressures if p in ("expansionary", "neutral", "contractionary")]
    if not known: return "neutral"
    c = {p: known.count(p) for p in ("expansionary", "neutral", "contractionary")}
    n = len(known)
    # Require a true majority (50%+) before showing a hard signal
    if c["contractionary"] / n >= 0.5: return "contractionary"
    if c["expansionary"]   / n >= 0.5: return "expansionary"
    # Neither side has a majority - return a mixed label showing the lean
    if c["contractionary"] > c["expansionary"]:  return "mixed_con"
    if c["expansionary"]   > c["contractionary"]: return "mixed_exp"
    return "mixed"


# ── WEIGHTED SIGNAL SCORING ──────────────────────────────────────────────────
# Leading indicators get 3x weight, coincident 2x, lagging 1x.
# This replaces the simple majority-vote as the primary macro assessment.

INDICATOR_WEIGHTS = {
    # Leading (3x) - predict the future
    "hy_spread": 3.0, "t10y3m": 3.0, "building_permits": 3.0,
    "cfnai": 2.5, "breakeven_10y": 2.5, "epu": 2.5,
    "jolts_openings": 2.0, "jolts_quits": 2.0, "ig_spread": 2.0, "housing_starts": 2.0, "dxy": 1.5,
    # Coincident (2x) - confirm the present
    "ism_mfg": 2.0, "ism_svc": 2.0, "spg_svc": 2.0, "initial_claims": 2.0,
    "indpro_mom": 2.0, "retail_mom": 1.5, "import_yoy": 1.5,
    "nonfarm_payrolls": 1.5, "adp": 1.5, "ahe_yoy": 1.5,
    "consumer_conf": 2.0, "chicago_pmi": 1.0,
    # Lagging (1x) - confirm the past
    "unemployment_rate": 1.0, "cpi_yoy": 1.5, "core_cpi_yoy": 1.5,
    "core_pce_yoy": 2.0, "ppi_mom": 1.0, "crb_yoy": 1.5,
}

INDICATOR_TYPE = {
    "hy_spread": "Leading", "ig_spread": "Leading", "t10y3m": "Leading",
    "building_permits": "Leading", "jolts_openings": "Leading",
    "jolts_quits": "Leading", "breakeven_10y": "Leading", "cfnai": "Leading",
    "epu": "Leading", "dxy": "Leading", "housing_starts": "Leading",
    "ism_mfg": "Coincident", "ism_svc": "Coincident", "spg_svc": "Coincident",
    "indpro_mom": "Coincident", "retail_mom": "Coincident", "import_yoy": "Coincident",
    "initial_claims": "Coincident", "nonfarm_payrolls": "Coincident", "adp": "Coincident", "ahe_yoy": "Coincident",
    "unemployment_rate": "Lagging", "cpi_yoy": "Lagging",
    "core_cpi_yoy": "Lagging", "core_pce_yoy": "Lagging", "ppi_mom": "Lagging", "crb_yoy": "Lagging",
    "consumer_conf": "Coincident", "chicago_pmi": "Coincident",
}

TYPE_COLOR = {"Leading": "0E6655", "Coincident": "2471A3", "Lagging": "7D3C98"}

# ── CONTINUOUS SCORING ───────────────────────────────────────────────────────
# Each indicator scored on [-1, +1] based on distance from neutral midpoint.
# Replaces the discrete {-1, 0, +1} mapping for weighted_signal().
# "goldilocks" indicators are contractionary at BOTH extremes (score <= 0 always).

SCORE_PARAMS = {
    # Job Market
    "unemployment_rate": {"mid": 5.25, "scale": 1.25, "invert": True},
    "initial_claims":    {"mid": 300,  "scale": 50,   "invert": True},
    "nonfarm_payrolls":  {"mid": 150,  "scale": 100,  "invert": False},
    "adp":               {"mid": 125,  "scale": 75,   "invert": False},
    "ahe_yoy":           {"mid": 3.0,  "scale": 1.0,  "goldilocks": True},
    # Inflation
    "cpi_yoy":           {"mid": 2.0,  "scale": 0.5,  "invert": True},
    "core_cpi_yoy":      {"mid": 2.15, "scale": 0.35, "invert": True},
    "core_pce_yoy":      {"mid": 2.15, "scale": 0.35, "invert": True},
    "ppi_mom":           {"mid": 0.25, "scale": 0.25, "invert": True},
    "crb_yoy":           {"mid": 0.0,  "scale": 10.0, "invert": True},
    # Economic Activities
    "ism_mfg":           {"mid": 53.0, "scale": 3.0,  "invert": False},
    "ism_svc":           {"mid": 53.0, "scale": 3.0,  "invert": False},
    "spg_svc":           {"mid": 53.0, "scale": 3.0,  "invert": False},
    "chicago_pmi":       {"mid": 53.0, "scale": 3.0,  "invert": False},
    "consumer_conf":     {"mid": 73.5, "scale": 16.5, "invert": False},
    "indpro_mom":        {"mid": 0.05, "scale": 0.25, "invert": False},
    "retail_mom":        {"mid": 0.1,  "scale": 0.4,  "invert": False},
    "import_yoy":        {"mid": 0.0,  "scale": 5.0,  "invert": True},
    # Leading Indicators
    "hy_spread":         {"mid": 4.25, "scale": 0.75, "invert": True},
    "ig_spread":         {"mid": 1.25, "scale": 0.25, "invert": True},
    "t10y3m":            {"mid": 0.5,  "scale": 0.5,  "invert": False},
    "building_permits":  {"mid": -2.5, "scale": 7.5,  "invert": False},
    "jolts_openings":    {"mid": -2.5, "scale": 12.5, "invert": False},
    "jolts_quits":       {"mid": 2.25, "scale": 0.25, "invert": False},
    "breakeven_10y":     {"mid": 2.15, "scale": 0.35, "goldilocks": True},
    "cfnai":             {"mid": -0.25,"scale": 0.45, "invert": False},
    "epu":               {"mid": 250,  "scale": 50,   "invert": True},
    "dxy":               {"mid": 110,  "scale": 10,   "invert": True},
    "housing_starts":    {"mid": -2.5, "scale": 12.5, "invert": False},
}

def continuous_score(key, val):
    """Continuous score in [-1, +1]. Captures magnitude, not just direction."""
    if val is None:
        return 0.0
    p = SCORE_PARAMS.get(key)
    if p is None:
        return 0.0
    mid, scale = p["mid"], p["scale"]
    if p.get("goldilocks"):
        s = -abs(val - mid) / scale
    elif p.get("invert"):
        s = (mid - val) / scale
    else:
        s = (val - mid) / scale
    return max(-1.0, min(1.0, s))


def weighted_signal(data):
    """
    Compute a weighted macro signal score.
    Returns: (label, score, leading_score, coincident_score, lagging_score)
    where score is in [-1, +1] (negative = contractionary, positive = expansionary).
    """
    tier_sums   = {"Leading": 0.0, "Coincident": 0.0, "Lagging": 0.0}
    tier_totals = {"Leading": 0.0, "Coincident": 0.0, "Lagging": 0.0}

    for key, w in INDICATOR_WEIGHTS.items():
        val = data.get(key, (None, []))[0]
        if val is None:
            continue
        score = continuous_score(key, val)
        tier = INDICATOR_TYPE.get(key, "Lagging")
        tier_sums[tier]   += score * w
        tier_totals[tier] += w

    tier_scores = {}
    for t in ("Leading", "Coincident", "Lagging"):
        tier_scores[t] = tier_sums[t] / tier_totals[t] if tier_totals[t] > 0 else 0.0

    total_w = sum(tier_totals.values())
    total_s = sum(tier_sums.values())
    normalized = total_s / total_w if total_w > 0 else 0.0

    if   normalized >  0.3: label = "expansionary"
    elif normalized < -0.3: label = "contractionary"
    elif normalized >  0.1: label = "mixed_exp"
    elif normalized < -0.1: label = "mixed_con"
    else:                   label = "neutral"

    return label, normalized, tier_scores["Leading"], tier_scores["Coincident"], tier_scores["Lagging"]


# ── METRIC CATALOGUE ──────────────────────────────────────────────────────────

METRICS = [
    ("unemployment_rate", "Unemployment Rate",         "Job Market",          '#,##0.0"%"',        ">6.5% Weak  |  4-6.5% Moderate  |  <4% Strong",    C["ch_ur"],   "Rate (%)"),
    ("initial_claims",    "Initial Jobless Claims (k)", "Job Market",          "#,##0.0",            ">350k Weak  |  250-350k Moderate  |  <250k Strong",  C["ch_ic"],   "Thousands"),
    ("nonfarm_payrolls",  "Nonfarm Payrolls (MoM k)",  "Job Market",          '+#,##0.0;-#,##0.0', "<50k Weak  |  50-250k Moderate  |  >250k Strong",   C["ch_nfp"],  "Thousands"),
    ("adp",               "ADP Employment (MoM k)",    "Job Market",          '+#,##0.0;-#,##0.0', "<50k Weak  |  50-200k Moderate  |  >200k Strong",   C["ch_adp"],  "Thousands"),
    ("ahe_yoy",           "Avg Hourly Earnings (YoY %)", "Job Market",        '#,##0.00"%"',        ">4% Overheating  |  2-4% Healthy  |  <2% Slack",   C["ch_ahe"],  "YoY (%)"),
    ("cpi_yoy",           "CPI (YoY %)",               "Inflation",           '#,##0.00"%"',        ">2.5% Elevated  |  1.5-2.5% On Target  |  <1.5% Low", C["ch_cpi"],  "YoY (%)"),
    ("core_cpi_yoy",      "Core CPI (YoY %)",          "Inflation",           '#,##0.00"%"',        ">2.5% Elevated  |  1.8-2.5% On Target  |  <1.8% Low", C["ch_ccpi"], "YoY (%)"),
    ("core_pce_yoy",     "Core PCE (YoY %)",          "Inflation",           '#,##0.00"%"',        ">2.5% Above Target  |  1.8-2.5% On Target  |  <1.8% Low", C["ch_pce"], "YoY (%)"),
    ("import_yoy",       "Import Price Index (YoY %)", "Inflation",          '+#,##0.0"%";-#,##0.0"%"', ">5% Elevated  |  -5 to 5% Normal  |  <-5% Falling",  C["ch_imp"], "YoY (%)"),
    ("ppi_mom",           "PPI - All Commodities (MoM %)", "Inflation",       '#,##0.00"%"',        ">0.5% Elevated  |  0-0.5% Moderate  |  <0% Falling",  C["ch_ppi"],  "MoM (%)"),
    ("crb_yoy",           "CRB Commodity Index (YoY %)", "Inflation",          '#,##0.00"%"',        ">10% Elevated  |  -10% to 10% Normal  |  <-10% Falling", C["ch_crb"],  "YoY (%)"),
    ("ism_mfg",           "ISM Manufacturing PMI",     "Economic Activities", "#,##0.0",            ">56 Strong  |  50-56 Moderate  |  <50 Contraction", C["ch_imfg"], "Index"),
    ("ism_svc",           "ISM Non-Manufacturing PMI", "Economic Activities", "#,##0.0",            ">56 Strong  |  50-56 Moderate  |  <50 Contraction", C["ch_isvc"], "Index"),
    ("spg_svc",           "S&P US Global Services PMI",   "Economic Activities", "#,##0.0",            ">56 Strong  |  50-56 Moderate  |  <50 Contraction", C["ch_spg"],  "Index"),
    ("chicago_pmi",       "Chicago PMI  (Regional)",   "Economic Activities", "#,##0.0",            ">56 Strong  |  50-56 Moderate  |  <50 Contraction", C["ch_chi"],  "Index"),
    ("indpro_mom",       "Industrial Production (MoM %)", "Economic Activities", '+#,##0.00"%"',    ">0.3% Expanding  |  Flat  |  <-0.2% Contracting",  C["ch_ip"],  "MoM (%)"),
    ("retail_mom",       "Retail Sales (MoM %)",      "Economic Activities", '+#,##0.00"%"',       ">0.5% Strong  |  Normal  |  <-0.3% Weak",          C["ch_ret"], "MoM (%)"),
    ("consumer_conf",     "Consumer Sentiment (UoM)",  "Economic Activities", "#,##0.0",            ">90 High  |  65-90 Moderate  |  <65 Low",           C["ch_cc"],   "Index"),
    # Leading Indicators
    ("hy_spread",         "HY Credit Spread (OAS %)",  "Leading Indicators",  '#,##0.00"%"',        ">5.0% Stress  |  3.5-5.0% Normal  |  <3.5% Tight",  C["ch_hy"],   "OAS (%)"),
    ("ig_spread",         "IG Credit Spread (OAS %)",  "Leading Indicators",  '#,##0.00"%"',        ">1.5% Stress  |  1.0-1.5% Normal  |  <1.0% Benign", C["ch_ig"],   "OAS (%)"),
    ("t10y3m",            "Yield Curve (10Y-3M)",      "Leading Indicators",  '+#,##0.00;-#,##0.00',  "<0 Inverted  |  0-1.0 Flat  |  >1.0 Steep",       C["ch_t3m"],  "% Spread"),
    ("building_permits",  "Building Permits (YoY %)",  "Leading Indicators",  '+#,##0.0"%";-#,##0.0"%"', "<-10% Falling  |  -10 to 5% Normal  |  >5% Rising", C["ch_bp"],   "YoY (%)"),
    ("jolts_openings",    "JOLTS Job Openings (YoY %)", "Leading Indicators", '+#,##0.0"%";-#,##0.0"%"', "<-15% Falling  |  Normal  |  >10% Rising",        C["ch_jo"],   "YoY (%)"),
    ("jolts_quits",       "JOLTS Quits Rate (%)",      "Leading Indicators",  '#,##0.0"%"',          "<2.0% Low  |  2.0-2.5% Normal  |  >2.5% High",     C["ch_jq"],   "Rate (%)"),
    ("breakeven_10y",     "10Y Breakeven Inflation",   "Leading Indicators",  '#,##0.00"%"',         "<1.8% Deflationary  |  1.8-2.5% Anchored  |  >2.5% Elevated", C["ch_be"], "Rate (%)"),
    ("cfnai",             "Chicago Fed Natl Activity",  "Leading Indicators",  '+#,##0.00;-#,##0.00', "<-0.7 Recession  |  -0.7 to 0.2 Trend  |  >0.2 Above Trend", C["ch_cfnai"], "Index"),
    ("epu",               "Policy Uncertainty (EPU)",   "Leading Indicators",  "#,##0.0",             ">300 Extreme  |  200-300 Elevated  |  <120 Low",                C["ch_epu"],   "Index"),
    ("dxy",               "Trade-Weighted Dollar",      "Leading Indicators",  "#,##0.0",             ">120 Strong (con)  |  100-120 Normal  |  <100 Weak (exp)",      C["ch_dxy"],   "Index"),
    ("housing_starts",    "Housing Starts (YoY %)",     "Leading Indicators", '+#,##0.0"%";-#,##0.0"%"', "<-15% Falling  |  Normal  |  >10% Rising",                 C["ch_hs"],    "YoY (%)"),
]
METRIC_INFO = {m[0]: m for m in METRICS}
UNIT_MAP = {
    "unemployment_rate": "%",       "initial_claims":  "k Claims",  "nonfarm_payrolls": "k MoM",  "adp": "k MoM",  "ahe_yoy": "% YoY",
    "cpi_yoy":           "% YoY",   "core_cpi_yoy":    "% YoY",  "core_pce_yoy":      "% YoY",  "import_yoy": "% YoY",  "ppi_mom": "% MoM",
    "crb_yoy":           "% YoY",
    "ism_mfg":           "Index",   "ism_svc":         "Index",   "spg_svc":           "Index",  "chicago_pmi": "Index",
    "indpro_mom":        "% MoM",   "retail_mom":      "% MoM",
    "consumer_conf":     "Index",
    # Leading indicators
    "hy_spread":         "% OAS",   "ig_spread":       "% OAS",      "t10y3m":            "% Spread",
    "building_permits":  "% YoY",   "jolts_openings":  "% YoY",   "jolts_quits":       "%",
    "breakeven_10y":     "%",       "cfnai":           "Index",
    "epu":               "Index",
    "dxy":               "Index",   "housing_starts":  "% YoY",
}


# ── ECONOMIC CALENDAR - release schedule for each metric ─────────────────────
# rule values: first_bday, third_bday, last_bday, first_friday, every_thursday,
#              mid_month, late_month, quarterly

RELEASE_SCHEDULE = {
    "ism_mfg":           {"name": "ISM Manufacturing PMI",       "rule": "first_bday",     "source": "PDF",  "priority": "HIGH"},
    "ism_svc":           {"name": "ISM Services PMI",            "rule": "third_bday",     "source": "PDF",  "priority": "HIGH"},
    "chicago_pmi":       {"name": "Chicago PMI (MNI)",           "rule": "last_bday",      "source": "PDF",  "priority": "HIGH"},
    "adp":               {"name": "ADP Employment Change",       "rule": "first_wednesday","source": "FRED", "priority": "MEDIUM"},
    "unemployment_rate": {"name": "Unemployment Rate (BLS)",     "rule": "first_friday",   "source": "FRED", "priority": "MEDIUM"},
    "nonfarm_payrolls":  {"name": "Nonfarm Payrolls (BLS)",      "rule": "first_friday",   "source": "FRED", "priority": "MEDIUM"},
    "ahe_yoy":           {"name": "Avg Hourly Earnings (BLS)",   "rule": "first_friday",   "source": "FRED", "priority": "MEDIUM"},
    "cpi_yoy":           {"name": "CPI (BLS)",                   "rule": "mid_month",      "source": "FRED", "priority": "MEDIUM"},
    "core_cpi_yoy":      {"name": "Core CPI (BLS)",              "rule": "mid_month",      "source": "FRED", "priority": "MEDIUM"},
    "ppi_mom":           {"name": "PPI (BLS)",                   "rule": "mid_month",      "source": "FRED", "priority": "MEDIUM"},
    "initial_claims":    {"name": "Initial Jobless Claims",      "rule": "every_thursday",  "source": "FRED", "priority": "LOW"},
    "spg_svc":           {"name": "S&P US Global Services PMI",     "rule": "late_month",     "source": "PDF",  "priority": "HIGH"},
    "core_pce_yoy":      {"name": "Core PCE (BEA)",              "rule": "late_month",     "source": "FRED", "priority": "MEDIUM"},
    "import_yoy":        {"name": "Import Price Index (BLS)",    "rule": "mid_month",      "source": "FRED", "priority": "LOW"},
    "indpro_mom":        {"name": "Industrial Production (Fed)", "rule": "mid_month",      "source": "FRED", "priority": "MEDIUM"},
    "retail_mom":        {"name": "Retail Sales (Census)",       "rule": "mid_month",      "source": "FRED", "priority": "MEDIUM"},
    "consumer_conf":     {"name": "UMich Consumer Sentiment",    "rule": "late_month",     "source": "FRED", "priority": "LOW"},
    "hy_spread":         {"name": "HY Credit Spread (OAS)",      "rule": "daily",          "source": "FRED", "priority": "LOW"},
    "ig_spread":         {"name": "IG Credit Spread (OAS)",      "rule": "daily",          "source": "FRED", "priority": "LOW"},
    "t10y3m":            {"name": "Yield Curve (10Y-3M)",        "rule": "daily",          "source": "FRED", "priority": "LOW"},
    "building_permits":  {"name": "Building Permits",            "rule": "mid_month",      "source": "FRED", "priority": "LOW"},
    "jolts_openings":    {"name": "JOLTS Job Openings",          "rule": "late_month",     "source": "FRED", "priority": "LOW"},
    "jolts_quits":       {"name": "JOLTS Quits Rate",            "rule": "late_month",     "source": "FRED", "priority": "LOW"},
    "breakeven_10y":     {"name": "10Y Breakeven Inflation",     "rule": "daily",          "source": "FRED", "priority": "LOW"},
    "cfnai":             {"name": "CFNAI",                       "rule": "late_month",     "source": "FRED", "priority": "LOW"},
    "epu":               {"name": "Policy Uncertainty (EPU)",    "rule": "daily",          "source": "FRED", "priority": "LOW"},
    "crb_yoy":           {"name": "CRB Commodity Index",         "rule": "mid_month",      "source": "FRED", "priority": "LOW"},
    "dxy":               {"name": "Trade-Weighted Dollar",      "rule": "daily",          "source": "FRED", "priority": "LOW"},
    "housing_starts":    {"name": "Housing Starts",             "rule": "mid_month",      "source": "FRED", "priority": "LOW"},
}

PRIORITY_ORDER = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}

def _next_release_date(rule, ref=None):
    """Compute the approximate next release date from today (or ref)."""
    from datetime import timedelta
    today = ref or datetime.now().date()
    yr, mo = today.year, today.month
    nxt_mo = mo + 1 if mo < 12 else 1
    nxt_yr = yr if mo < 12 else yr + 1

    def _nth_weekday(year, month, weekday, n):
        """Return the n-th weekday (0=Mon..4=Fri) of the given month."""
        from calendar import monthrange
        first = datetime(year, month, 1).date()
        offset = (weekday - first.weekday()) % 7
        d = first + timedelta(days=offset + 7 * (n - 1))
        return d

    def _last_bday(year, month):
        from calendar import monthrange
        last = datetime(year, month, monthrange(year, month)[1]).date()
        while last.weekday() > 4:
            last -= timedelta(days=1)
        return last

    def _nth_bday(year, month, n):
        d = datetime(year, month, 1).date()
        count = 0
        while count < n:
            if d.weekday() <= 4:
                count += 1
            if count < n:
                d += timedelta(days=1)
        return d

    def _nearest_weekday(d):
        while d.weekday() > 4:
            d += timedelta(days=1)
        return d

    if rule == "first_bday":
        d = _nth_bday(nxt_yr if mo == 12 else yr, nxt_mo, 1)
        return d if d > today else _nth_bday(nxt_yr + (1 if nxt_mo == 12 else 0), nxt_mo + 1 if nxt_mo < 12 else 1, 1)
    elif rule == "third_bday":
        d = _nth_bday(nxt_yr if mo == 12 else yr, nxt_mo, 3)
        return d if d > today else _nth_bday(nxt_yr + (1 if nxt_mo == 12 else 0), nxt_mo + 1 if nxt_mo < 12 else 1, 3)
    elif rule == "last_bday":
        d = _last_bday(yr, mo)
        return d if d > today else _last_bday(nxt_yr if mo == 12 else yr, nxt_mo)
    elif rule == "first_wednesday":
        d = _nth_weekday(nxt_yr if mo == 12 else yr, nxt_mo, 2, 1)
        return d if d > today else _nth_weekday(nxt_yr + (1 if nxt_mo == 12 else 0), nxt_mo + 1 if nxt_mo < 12 else 1, 2, 1)
    elif rule == "first_friday":
        d = _nth_weekday(nxt_yr if mo == 12 else yr, nxt_mo, 4, 1)
        return d if d > today else _nth_weekday(nxt_yr + (1 if nxt_mo == 12 else 0), nxt_mo + 1 if nxt_mo < 12 else 1, 4, 1)
    elif rule == "every_thursday":
        d = today + timedelta(days=(3 - today.weekday()) % 7 or 7)
        return d
    elif rule == "mid_month":
        d = _nearest_weekday(datetime(nxt_yr if mo == 12 else yr, nxt_mo, 13).date())
        return d if d > today else _nearest_weekday(datetime(nxt_yr + (1 if nxt_mo == 12 else 0), nxt_mo + 1 if nxt_mo < 12 else 1, 13).date())
    elif rule == "late_month":
        d = _nearest_weekday(datetime(yr, mo, 25).date()) if datetime(yr, mo, 25).date() > today else _nearest_weekday(datetime(nxt_yr if mo == 12 else yr, nxt_mo, 25).date())
        return d
    elif rule == "daily":
        d = today + timedelta(days=1)
        return _nearest_weekday(d)
    return today + timedelta(days=30)

def build_calendar_entries():
    """Return sorted list of (days_until, key, name, source, priority, release_date)."""
    today = datetime.now().date()
    entries = []
    for key, info in RELEASE_SCHEDULE.items():
        if info["rule"] == "daily":
            continue
        rd = _next_release_date(info["rule"], today)
        days = (rd - today).days
        entries.append((days, key, info["name"], info["source"], info["priority"], rd))
    entries.sort(key=lambda e: (PRIORITY_ORDER.get(e[4], 9), e[0]))
    return entries


# ── ECONOMIC REGIME & SECTOR PLAYBOOK DATA ────────────────────────────────────
# Framework: Invictus Research - four regimes on Growth × Inflation axes.
# Sector returns from the Invictus Equity Sector Backtest (stack-ranked avg annual returns).

REGIME_COLORS = {
    "Early Recovery": "1A6B3A",   # deep green   - growth ↑  inflation ↓
    "Reflation":      "154360",   # navy blue    - growth ↑  inflation ↑
    "Stagflation":    "A04000",   # burnt orange - growth ↓  inflation ↑
    "Deflation":      "7B241C",   # dark red     - growth ↓  inflation ↓
}

REGIME_EQUITY = {
    "Early Recovery": ("Excellent", "1E8449"),
    "Reflation":      ("Excellent", "1A5276"),
    "Stagflation":    ("Moderate",  "D4870A"),
    "Deflation":      ("Poor",      "B03A2E"),
}

REGIME_DESC = {
    "Early Recovery": "Growth rising  ·  Inflation subdued  ·  Best broad equity backdrop",
    "Reflation":      "Growth rising  ·  Inflation rising  ·  Commodities & cyclicals lead",
    "Stagflation":    "Growth falling  ·  Inflation elevated  ·  Defensives outperform",
    "Deflation":      "Growth falling  ·  Inflation subdued  ·  Capital preservation priority",
}

# Stack-ranked avg annual sector returns per regime (source: Invictus Research backtest)
REGIME_SECTORS = {
    "Early Recovery": [
        ("Technology",        10), ("Consumer Cyclical",  8), ("Communications",     8),
        ("Healthcare",         8), ("Industrials",         7), ("Materials",           6),
        ("Financials",         5), ("Energy",              0), ("Real Estate",         0),
        ("Consumer Staples",   0), ("Utilities",           0),
    ],
    "Reflation": [
        ("Financials",        13), ("Industrials",        12), ("Materials",          10),
        ("Energy",             8), ("Technology",          8), ("Consumer Cyclical",   8),
        ("Communications",     7), ("Real Estate",         7), ("Consumer Staples",    4),
        ("Healthcare",         3), ("Utilities",           2),
    ],
    "Stagflation": [
        ("Healthcare",         8), ("Real Estate",         6), ("Materials",           5),
        ("Energy",             5), ("Technology",          3), ("Industrials",         3),
        ("Utilities",          2), ("Consumer Staples",    1), ("Consumer Cyclical",   1),
        ("Communications",    -1), ("Financials",         -1),
    ],
    "Deflation": [
        ("Consumer Staples",   8), ("Healthcare",          3), ("Real Estate",         3),
        ("Utilities",          3), ("Consumer Cyclical",   0), ("Communications",     -1),
        ("Technology",        -2), ("Industrials",        -4), ("Materials",          -5),
        ("Financials",        -5), ("Energy",             -8),
    ],
}

# Style-factor rankings by monetary-policy stance (source: Invictus Research)
STYLE_FACTORS = {
    "Easier Policy": [
        "Small Cap > Large Cap",
        "Growth > Value",
        "Cyclical > Defensive",
        "High Valuation > Low Valuation",
        "High Beta > Low Beta",
        "Low Liquidity > High Liquidity",
    ],
    "Tighter Policy": [
        "Large Cap > Small Cap",
        "Value > Growth",
        "Defensive > Cyclical",
        "Low Valuation > High Valuation",
        "Low Beta > High Beta",
        "High Liquidity > Low Liquidity",
    ],
}


def classify_regime(data, dfm_factors=None):
    """
    Map current macro data onto one of four regimes:
      Early Recovery · Reflation · Stagflation · Deflation

    When dfm_factors is provided (dict with "growth" and "inflation" z-scores),
    the DFM factor signs drive the regime - statistically rigorous.
    Fallback: vote-based logic on raw indicator values.
    """
    if dfm_factors is not None:
        growth_up    = dfm_factors["growth"]    > 0
        inflation_up = dfm_factors["inflation"] > 0
    else:
        growth_keys = ["unemployment_rate", "initial_claims", "nonfarm_payrolls",
                       "adp", "ism_mfg", "ism_svc", "consumer_conf"]
        growth_votes = []
        for key in growth_keys:
            val = data.get(key, (None, []))[0]
            if val is None: continue
            p, _, _ = classify(key, val)
            if   p == "expansionary":   growth_votes.append(1)
            elif p == "contractionary": growth_votes.append(-1)
            else:                       growth_votes.append(0)
        growth_up = sum(growth_votes) > 0 if growth_votes else False

        # Inflation: positive score = inflation elevated, negative = subdued
        # weights: CPI/Core CPI most important (×2), CRB leads CPI (×2), PPI (×1)
        inf_score = 0
        checks = [
            ("cpi_yoy",      2.0,   1.5,  2),
            ("core_cpi_yoy", 2.0,   1.5,  2),
            ("ppi_mom",      0.2,   0.0,  1),
            ("crb_yoy",      5.0,  -5.0,  2),
        ]
        for key, hi, lo, w in checks:
            val = data.get(key, (None, []))[0]
            if val is None: continue
            if   val > hi: inf_score += w
            elif val < lo: inf_score -= w
        inflation_up = inf_score > 0

    if   growth_up and not inflation_up: return "Early Recovery"
    elif growth_up and     inflation_up: return "Reflation"
    elif not growth_up and inflation_up: return "Stagflation"
    else:                                return "Deflation"


def monetary_policy_stance(ff_hist):
    """
    Returns 'Easier Policy' or 'Tighter Policy'.
    Easier = Fed cutting (3-month avg < 12-month avg) OR rate ≤ 2.5% (below neutral).
    Tighter = Fed hiking or holding above neutral.
    """
    if not ff_hist: return "Tighter Policy"
    vals  = [v for _, v in ff_hist]
    avg3  = sum(vals[:3])  / min(3,  len(vals))
    avg12 = sum(vals[:12]) / min(12, len(vals))
    if avg3 < avg12 or vals[0] <= 2.5:
        return "Easier Policy"
    return "Tighter Policy"


# ── MARKET POSITIONING DATA ────────────────────────────────────────────────────
# Sources: CBOE (VIX), Chicago Fed (NFCI), St. Louis Fed (STLFSI4),
#          Fed Senior Loan Officer Survey (DRTSCILM).
# CFTC COT and AAII Sentiment are referenced manually - see positioning sheet.

POS_KEYS = ["vix", "nfci", "stlfsi", "lending", "yield_curve"]

POS_META = {
    "vix":     ("CBOE VIX  (Fear Index)",           "VIXCLS",   "index",  ">30 Elevated  |  15-30 Normal  |  <15 Complacent",  C["ch_vix"]),
    "nfci":    ("Chicago Financial Conditions",      "NFCI",     "index",  ">0 Tight  |  −0.5 to 0 Normal  |  <−0.5 Loose",    C["ch_nfci"]),
    "stlfsi":  ("St. Louis Financial Stress Index",  "STLFSI4",  "index",  ">0.5 High  |  −0.5 to 0.5 Normal  |  <−0.5 Low",   C["ch_fsi"]),
    "lending": ("Bank Lending Standards (C&I %net)", "DRTSCILM", "% net",  ">20 Tightening  |  −10 to 20 Normal  |  <−10 Easing", C["ch_lend"]),
    "yield_curve": ("10Y-2Y Yield Curve Spread", "T10Y2Y", "% spread", ">0.5% Normal  |  −0.5-0.5% Flat  |  <−0.5% Inverted", C["ch_yc"]),
}

POS_RULES = {
    "vix": [
        (lambda v: v > 40, "extreme_fear",  "Extreme fear - forced liquidation / capitulation risk",       "Contrarian Bullish - shorts extremely crowded"),
        (lambda v: v > 28, "elevated",      "Elevated fear - risk-off sentiment dominant",                 "Cautious - hedging demand elevated"),
        (lambda v: v < 13, "complacent",    "Extreme complacency - investors fully long, minimal hedges",  "Contrarian Bearish - longs very crowded"),
        (lambda v: v < 18, "calm",          "Low volatility - risk-on positioning favoured",               "Risk-On - benign conditions"),
        (lambda v: True,   "neutral",       "Normal volatility range - balanced sentiment",                "Neutral - no strong positioning signal"),
    ],
    "nfci": [
        (lambda v: v >  0.5, "tight",   "Tight financial conditions - credit stress elevated",      "Risk-Off - reduce risk exposure"),
        (lambda v: v < -0.5, "loose",   "Loose financial conditions - ample liquidity available",  "Risk-On - conditions supportive of risk assets"),
        (lambda v: True,     "neutral", "Financial conditions near long-run average",               "Neutral - no clear directional signal"),
    ],
    "stlfsi": [
        (lambda v: v >  0.5, "stress",  "Above-average financial stress - caution warranted",     "Risk-Off - stress elevated"),
        (lambda v: v < -0.5, "benign",  "Below-average financial stress - benign environment",    "Risk-On - stress subdued"),
        (lambda v: True,     "neutral", "Financial stress near long-run average",                 "Neutral - no clear directional signal"),
    ],
    "lending": [
        (lambda v: v > 20,  "tight",   "Banks tightening credit - economic headwinds building",   "Risk-Off - credit crunch signal"),
        (lambda v: v < -10, "easing",  "Banks easing credit - supportive of economic expansion",  "Risk-On - credit cycle supportive"),
        (lambda v: True,    "neutral", "Lending standards near normal",                           "Neutral - credit conditions OK"),
    ],
    "yield_curve": [
        (lambda v: v > 0.5,  "normal",   "Upward sloping yield curve - growth-supportive environment",                     "Risk-On - curve structure supports expansion"),
        (lambda v: v < -0.5, "inverted", "Inverted yield curve - historically precedes recession by 12-18 months",         "Risk-Off - recession warning signal"),
        (lambda v: True,     "flat",     "Flat yield curve - transition signal, watch for directional break",               "Neutral - caution warranted"),
    ],
}

POS_RISK_OFF = {"extreme_fear", "elevated", "tight", "stress", "inverted"}
POS_RISK_ON  = {"complacent", "calm", "loose", "benign", "easing", "normal"}


def classify_pos(key, val):
    if val is None:
        return "neutral", "No data available", "Neutral"
    for cond, signal, impact, action in POS_RULES.get(key, []):
        if cond(val): return signal, impact, action
    return "neutral", "N/A", "Neutral"

def pos_bg(signal):
    if signal in POS_RISK_OFF: return C["con"]
    if signal in POS_RISK_ON:  return C["exp"]
    return C["neu"]

def pos_lbl(signal):
    labels = {
        "extreme_fear": "EXTREME FEAR",  "elevated": "RISK-OFF",
        "complacent":   "COMPLACENT",    "calm":     "RISK-ON",
        "tight":        "TIGHT",         "loose":    "LOOSE",
        "stress":       "STRESSED",      "benign":   "BENIGN",
        "easing":       "EASING",
        "inverted": "INVERTED",  "normal": "NORMAL",  "flat": "FLAT",
    }
    return labels.get(signal, "NEUTRAL")

def overall_positioning(signals):
    risk_off = sum(1 for s in signals if s in POS_RISK_OFF)
    risk_on  = sum(1 for s in signals if s in POS_RISK_ON)
    if risk_off > risk_on:   return "RISK-OFF", C["con"]
    if risk_on  > risk_off:  return "RISK-ON",  C["exp"]
    return "NEUTRAL", C["neu"]

def fetch_positioning():
    print("  Fetching positioning indicators (VIX, NFCI, STLFSI4, DRTSCILM, T10Y2Y)...")
    fetch_limits = {"vix": 90, "nfci": 54, "stlfsi": 54, "lending": 10, "yield_curve": 300}
    result = {}
    for key in POS_KEYS:
        _, sid, *_ = POS_META[key]
        lim  = fetch_limits.get(key, 12)
        raw  = _fred(sid, lim)
        result[key] = (raw[0][1] if raw else None, raw)
    return result


# ── DATA FETCHING ─────────────────────────────────────────────────────────────

def fetch_all():
    print("  Fetching FRED series...")
    ism = read_manual_input()
    _spg_csv = _read_pmi_csv(CSV_SPG_SVC)
    print("  Fetching leading indicators (credit spreads, yield curve, JOLTS, permits, CFNAI)...")
    d = {
        "unemployment_rate": fetch_latest("UNRATE"),
        "initial_claims":    _scale_series(fetch_latest("IC4WSA"), 0.001),  # persons -> thousands
        "nonfarm_payrolls":  fetch_payrolls(),
        "adp":               fetch_adp(),
        "ahe_yoy":           fetch_yoy("CES0500000003"),
        "cpi_yoy":           fetch_yoy("CPIAUCSL"),
        "import_yoy":        fetch_yoy("IR"),
        "core_cpi_yoy":      fetch_yoy("CPILFESL"),
        "core_pce_yoy":      fetch_yoy("PCEPILFE"),
        "ppi_mom":           fetch_mom_pct("PPIACO"),
        "crb_yoy":           fetch_yoy("PALLFNFINDEXM"),   # IMF Global All-Commodity Price Index - CRB proxy
        "ism_mfg":           ism["ism_mfg"],
        "ism_svc":           ism["ism_svc"],
        "spg_svc":           (_spg_csv[0][1] if _spg_csv else None, _spg_csv[:12]),
        "chicago_pmi":       ism["chicago_pmi"],
        "indpro_mom":        fetch_mom_pct("INDPRO"),
        "retail_mom":        fetch_mom_pct("RSAFS"),
        "consumer_conf":     fetch_latest("UMCSENT"),
        # Leading indicators
        "hy_spread":         fetch_daily_latest("BAMLH0A0HYM2", 400),
        "ig_spread":         fetch_daily_latest("BAMLC0A0CM", 400),
        "t10y3m":            fetch_daily_latest("T10Y3M", 400),
        "building_permits":  fetch_level_yoy("PERMIT"),
        "jolts_openings":    fetch_level_yoy("JTSJOL"),
        "jolts_quits":       fetch_latest("JTSQUR"),
        "breakeven_10y":     fetch_daily_latest("T10YIE", 400),
        "cfnai":             fetch_latest("CFNAI"),
        "epu":               fetch_daily_latest("USEPUINDXD", 400),
        "dxy":               fetch_daily_latest("DTWEXBGS", 400),
        "housing_starts":    fetch_level_yoy("HOUST"),
    }
    return d


# ── DYNAMIC FACTOR MODEL (DFM) ────────────────────────────────────────────────
# Two-factor model: one latent GROWTH factor, one latent INFLATION factor.
# Growth panel: UNRATE, IC4WSA(→monthly), PAYEMS MoM, UMCSENT, ISM Mfg, ISM Svc, Chicago PMI.
# Inflation panel: CPI YoY, Core CPI YoY, PPI MoM, CRB YoY.
# PMIs now included via 121-month CSV history (2016-01 to 2026-02).
# statsmodels Kalman filter is used; scipy PCA is the automatic fallback.

def _weekly_to_monthly_mean(raw_weekly):
    """Aggregate IC4WSA weekly obs to monthly means. Returns [(YYYY-MM-01, mean), ...] newest-first."""
    from collections import defaultdict
    monthly = defaultdict(list)
    for date, val in raw_weekly:
        ym = date[:7]
        monthly[ym].append(val)
    result = [(f"{ym}-01", sum(vs) / len(vs)) for ym, vs in monthly.items()]
    result.sort(key=lambda x: x[0], reverse=True)
    return result


def _daily_to_monthly_mean(raw_daily):
    """Aggregate daily FRED obs to monthly means. Returns [(YYYY-MM-01, mean), ...] newest-first."""
    from collections import defaultdict
    monthly = defaultdict(list)
    for date, val in raw_daily:
        ym = date[:7]
        monthly[ym].append(val)
    result = [(f"{ym}-01", round(sum(vs) / len(vs), 4)) for ym, vs in monthly.items()]
    result.sort(key=lambda x: x[0], reverse=True)
    return result


def _yoy_extended(sid, limit=80):
    raw = _fred(sid, limit)
    if not raw: return None, []
    by_ym = {r[0][:7]: r[1] for r in raw}
    hist  = []
    for date, val in raw:
        ym     = date[:7]
        yr_ago = f"{int(ym[:4]) - 1}{ym[4:]}"
        if yr_ago in by_ym and len(hist) < limit - 12:
            hist.append((date, round((val - by_ym[yr_ago]) / by_ym[yr_ago] * 100, 2)))
    if not hist: return None, []
    return hist[0][1], hist

def _mom_extended(sid, limit=75):
    raw = _fred(sid, limit)
    if len(raw) < 2: return None, []
    hist = [(raw[i][0], round((raw[i][1] - raw[i+1][1]) / raw[i+1][1] * 100, 2))
            for i in range(min(limit - 1, len(raw) - 1))]
    return hist[0][1] if hist else None, hist

def _payrolls_extended(limit=75):
    raw = _fred("PAYEMS", limit)
    if len(raw) < 2: return None, []
    hist = [(raw[i][0], round(raw[i][1] - raw[i+1][1], 1))
            for i in range(min(limit - 1, len(raw) - 1))]
    return hist[0][1] if hist else None, hist

def _adp_extended(limit=280):
    """ADP private payrolls MoM change (thousands) - extended history for DFM. ADPMNUSNERSA is in persons, so /1000."""
    raw = _fred("ADPMNUSNERSA", limit)
    if len(raw) < 2: return None, []
    hist = [(raw[i][0], round((raw[i][1] - raw[i+1][1]) / 1000, 1))
            for i in range(min(limit - 1, len(raw) - 1))]
    return hist[0][1] if hist else None, hist


def fetch_dfm_history():
    """
    Fetch extended FRED history + CSV PMI history for DISPLAY DFM (120 months).
    Returns dict: key -> (latest_value, [(date, val), ...]) newest-first.
    """
    print("  Fetching extended history for display DFM (120-month panels)...")

    ic_weekly = _fred("IC4WSA", 700)
    ic_monthly = _weekly_to_monthly_mean(ic_weekly)

    unrate_raw  = _fred("UNRATE", 140)
    umcsent_raw = _fred("UMCSENT", 140)

    _mfg_csv = _read_pmi_csv(CSV_ISM_MFG)
    _svc_csv = _read_pmi_csv(CSV_ISM_SVC)
    _chi_csv = _read_pmi_csv(CSV_CHICAGO)
    _spg_csv = _read_pmi_csv(CSV_SPG_SVC)

    cfnai_raw   = _fred("CFNAI", 140)
    permit_yoy  = _yoy_extended("PERMIT", 160)
    jolts_raw   = _fred("JTSJOL", 140)
    be_daily    = _fred("T10YIE", 3000)
    be_monthly  = _daily_to_monthly_mean(be_daily) if be_daily else []
    m2_yoy      = _yoy_extended("M2SL", 160)
    pce_yoy     = _yoy_extended("PCEPILFE", 160)
    import_yoy  = _yoy_extended("IR", 160)
    indpro_mom  = _mom_extended("INDPRO", 140)
    housing_yoy = _yoy_extended("HOUST", 160)
    hy_daily    = _fred("BAMLH0A0HYM2", 3000)
    hy_monthly  = _daily_to_monthly_mean(hy_daily) if hy_daily else []
    ig_daily    = _fred("BAMLC0A0CM", 3000)
    ig_monthly  = _daily_to_monthly_mean(ig_daily) if ig_daily else []
    nfci_weekly = _fred("NFCI", 700)
    nfci_monthly = _weekly_to_monthly_mean(nfci_weekly)
    t3m_daily   = _fred("T10Y3M", 3000)
    t3m_monthly = _daily_to_monthly_mean(t3m_daily) if t3m_daily else []

    return {
        "unrate":        (unrate_raw[0][1] if unrate_raw else None,  unrate_raw),
        "ic4wsa_m":      (ic_monthly[0][1] if ic_monthly else None,  ic_monthly),
        "payems_m":      _payrolls_extended(140),
        "adp_m":         _adp_extended(140),
        "umcsent":       (umcsent_raw[0][1] if umcsent_raw else None, umcsent_raw),
        "cpi_yoy":       _yoy_extended("CPIAUCSL", 160),
        "core_yoy":      _yoy_extended("CPILFESL", 160),
        "ppi_mom":       _mom_extended("PPIACO", 140),
        "crb_yoy":       _yoy_extended("PALLFNFINDEXM", 160),
        "pce_yoy":       pce_yoy,
        "import_yoy":    import_yoy,
        "indpro_mom":    indpro_mom,
        "housing_yoy":   housing_yoy,
        "ism_mfg_csv":   (_mfg_csv[0][1] if _mfg_csv else None, _mfg_csv),
        "ism_svc_csv":   (_svc_csv[0][1] if _svc_csv else None, _svc_csv),
        "chicago_pmi_csv": (_chi_csv[0][1] if _chi_csv else None, _chi_csv),
        "spg_svc_csv":   (_spg_csv[0][1] if _spg_csv else None, _spg_csv),
        "cfnai":         (cfnai_raw[0][1] if cfnai_raw else None, cfnai_raw),
        "permit_yoy":    permit_yoy,
        "jolts":         (jolts_raw[0][1] if jolts_raw else None, jolts_raw),
        "breakeven":     (be_monthly[0][1] if be_monthly else None, be_monthly),
        "m2_yoy":        m2_yoy,
        "hy_oas":        (hy_monthly[0][1] if hy_monthly else None, hy_monthly),
        "ig_oas":        (ig_monthly[0][1] if ig_monthly else None, ig_monthly),
        "nfci_m":        (nfci_monthly[0][1] if nfci_monthly else None, nfci_monthly),
        "t10y3m":        (t3m_monthly[0][1] if t3m_monthly else None, t3m_monthly),
    }


def fetch_hmm_history():
    """
    Fetch LONG history (400+ months, FRED-only, NO ISM CSVs) for HMM training.
    Extended back to ~1993 for maximum regime diversity: 1990s expansion,
    dot-com bust, 2008 GFC, 2010s recovery, COVID crash, 2022 inflation spike.
    """
    print("  Fetching long history for HMM training (580+ months, FRED + Reuters)...")

    # Monthly series - fetch 620 obs to get ~50 years back to ~1976
    unrate_raw  = _fred("UNRATE", 620)
    umcsent_raw = _fred("UMCSENT", 620)
    cfnai_raw   = _fred("CFNAI", 620)
    jolts_raw   = _fred("JTSJOL", 620)

    ic_weekly = _fred("IC4WSA", 3500)
    ic_monthly = _weekly_to_monthly_mean(ic_weekly)

    # YoY and MoM series - fetch extra for the 12-month lookback
    payems_h = _payrolls_extended(620)
    permit_h = _yoy_extended("PERMIT", 640)
    cpi_h    = _yoy_extended("CPIAUCSL", 640)
    core_h   = _yoy_extended("CPILFESL", 640)
    ppi_h    = _mom_extended("PPIACO", 620)
    crb_h    = _yoy_extended("PALLFNFINDEXM", 640)
    m2_h     = _yoy_extended("M2SL", 640)
    pce_h    = _yoy_extended("PCEPILFE", 640)
    indpro_h = _mom_extended("INDPRO", 620)
    housing_h = _yoy_extended("HOUST", 640)

    # Reuters Commodities Index (daily CSV -> monthly YoY %)
    reuters_crb = _read_reuters_commodity()

    # Daily series - fetch 8500+ trading days for ~34 years
    be_daily  = _fred("T10YIE", 8500)
    be_monthly = _daily_to_monthly_mean(be_daily) if be_daily else []
    hy_daily  = _fred("BAMLH0A0HYM2", 8500)
    hy_monthly = _daily_to_monthly_mean(hy_daily) if hy_daily else []
    ig_daily  = _fred("BAMLC0A0CM", 8500)
    ig_monthly = _daily_to_monthly_mean(ig_daily) if ig_daily else []
    nfci_weekly = _fred("NFCI", 2500)
    nfci_monthly = _weekly_to_monthly_mean(nfci_weekly)
    t3m_daily = _fred("T10Y3M", 8500)
    t3m_monthly = _daily_to_monthly_mean(t3m_daily) if t3m_daily else []

    return {
        # Growth panel (FRED-only, no ISM CSVs)
        "unrate":   (unrate_raw[0][1] if unrate_raw else None, unrate_raw),
        "ic4wsa_m": (ic_monthly[0][1] if ic_monthly else None, ic_monthly),
        "payems_m": payems_h,
        "umcsent":  (umcsent_raw[0][1] if umcsent_raw else None, umcsent_raw),
        "cfnai":    (cfnai_raw[0][1] if cfnai_raw else None, cfnai_raw),
        "permits":  permit_h,
        "jolts":    (jolts_raw[0][1] if jolts_raw else None, jolts_raw),
        "indpro_m": indpro_h,
        "housing_yoy": housing_h,
        # Inflation panel
        "cpi_yoy":  cpi_h,
        "core_yoy": core_h,
        "ppi_mom":  ppi_h,
        "crb_yoy":  crb_h,
        "breakeven":(be_monthly[0][1] if be_monthly else None, be_monthly),
        "m2_yoy":   m2_h,
        "pce_yoy":  pce_h,
        "reuters_crb_yoy": (reuters_crb[0][1] if reuters_crb else None, reuters_crb),
        # Credit panel
        "hy_oas":   (hy_monthly[0][1] if hy_monthly else None, hy_monthly),
        "ig_oas":   (ig_monthly[0][1] if ig_monthly else None, ig_monthly),
        "nfci_m":   (nfci_monthly[0][1] if nfci_monthly else None, nfci_monthly),
        "t10y3m":   (t3m_monthly[0][1] if t3m_monthly else None, t3m_monthly),
    }


def _build_aligned_panel(series_list, n=120):
    """
    Inner-join series_list on shared YYYY-MM, return n most-recent complete months.
    series_list: [(name, [(date, val), ...]), ...]  (each list newest-first)
    Returns pd.DataFrame or None if < 36 complete months found.
    """
    by_ym = {}
    for name, hist in series_list:
        for date, val in hist:
            ym = date[:7]
            if ym not in by_ym:
                by_ym[ym] = {}
            by_ym[ym][name] = val

    names = [s[0] for s in series_list]
    complete = {ym: row for ym, row in by_ym.items() if all(k in row for k in names)}
    if len(complete) < 36:
        print(f"  [WARN] DFM panel has only {len(complete)} complete months - need >=36. Skipping.")
        return None

    sorted_months = sorted(complete.keys())[-n:]
    df = pd.DataFrame([complete[ym] for ym in sorted_months],
                      index=sorted_months, columns=names)
    return df


def _estimate_dfm(panel_df, anchor_col, factor_name, rolling_zscore=0):
    """
    Estimate a single latent factor from panel_df via Kalman filter (DFM) or PCA fallback.
    anchor_col: column name used for sign-normalisation (factor positively correlated with this).
    rolling_zscore: if >0, z-score each series against a trailing window of this many months
                    (instead of the full sample). Useful for long panels where the unconditional
                    mean shifts across eras (e.g., 1970s inflation vs 2020s).
    Returns dict with latest score, momentum, loadings, var_explained, top_contrib, method.
    Returns None on failure.
    """
    if panel_df is None or len(panel_df) < 36:
        return None
    try:
        if rolling_zscore > 0:
            # Rolling z-score: each month normalised against trailing window
            Z = panel_df.copy()
            for col in Z.columns:
                rm = Z[col].rolling(window=rolling_zscore, min_periods=36).mean()
                rs = Z[col].rolling(window=rolling_zscore, min_periods=36).std().replace(0, 1)
                Z[col] = (Z[col] - rm) / rs
            Z = Z.dropna()  # drop burn-in rows
            if len(Z) < 36:
                return None
        else:
            means = panel_df.mean()
            stds  = panel_df.std().replace(0, 1)
            Z     = (panel_df - means) / stds

        method     = "PCA"
        factor_raw = None
        loadings_v = None

        if _HAS_STATSMODELS:
            try:
                model      = _DFM(Z, k_factors=1, factor_order=2)
                result     = model.fit(disp=False, maxiter=200)
                # factors.filtered shape: (k_factors, nobs) in statsmodels 0.14+
                factor_raw = pd.Series(result.factors.filtered[0, :], index=Z.index)
                # Loadings: params named 'loading.f1.<col>'
                lp = {k.replace("loading.f1.", ""): v
                      for k, v in result.params.items()
                      if k.startswith("loading.f1.")}
                loadings_v = {col: float(lp.get(col, 0.0)) for col in Z.columns}
                # Idiosyncratic variance: params named 'sigma2.<col>'
                sp = {k.replace("sigma2.", ""): v
                      for k, v in result.params.items()
                      if k.startswith("sigma2.")}
                var_exp = {}
                for col in Z.columns:
                    l2   = loadings_v[col] ** 2
                    idio = max(float(sp.get(col, 1e-6)), 1e-6)
                    var_exp[col] = round(l2 / (l2 + idio) * 100, 1)
                method = "DFM"
            except Exception as e:
                print(f"  [WARN] DFM fit failed for {factor_name} ({e}) - falling back to PCA")
                factor_raw = None

        if factor_raw is None:
            cov        = np.cov(Z.values, rowvar=False)
            eigvals, eigvecs = np.linalg.eigh(cov)
            idx        = np.argmax(eigvals)
            lv         = eigvecs[:, idx]
            factor_raw = pd.Series(Z.values @ lv, index=Z.index)
            loadings_v = {col: float(lv[i]) for i, col in enumerate(Z.columns)}
            total_var  = float(eigvals.sum())
            ev         = float(eigvals[idx])
            var_exp    = {col: round(loadings_v[col]**2 * ev / max(total_var, 1e-9) * 100, 1)
                          for col in Z.columns}
            method     = "PCA"

        # Sign-normalise: factor should be positively correlated with anchor_col
        if anchor_col in Z.columns:
            r = float(np.corrcoef(factor_raw.values, Z[anchor_col].values)[0, 1])
            if r < 0:
                factor_raw = -factor_raw
                loadings_v = {k: -v for k, v in loadings_v.items()}

        # Standardise to z-score (mean=0, std=1) so ±0.5 thresholds are meaningful
        f_std = float(factor_raw.std())
        if f_std > 1e-9:
            factor_raw = (factor_raw - factor_raw.mean()) / f_std

        latest   = float(factor_raw.iloc[-1])
        momentum = float(factor_raw.iloc[-1] - factor_raw.iloc[-4]) if len(factor_raw) >= 4 else 0.0

        # Top contributor: loading × z-score of latest obs
        z_latest   = {col: float(Z[col].iloc[-1]) for col in Z.columns}
        contribs   = {col: loadings_v[col] * z_latest[col] for col in Z.columns}
        top_contrib = max(contribs.items(), key=lambda x: abs(x[1]))

        # Console diagnostics
        sorted_loads = sorted(loadings_v.items(), key=lambda x: abs(x[1]), reverse=True)
        load_str = " | ".join(f"{k.upper()} {v:+.2f}" for k, v in sorted_loads)
        vexp_str = " | ".join(f"{k.upper()} {var_exp[k]:.0f}%" for k, _ in sorted_loads)
        mom_dir  = "improving" if momentum > 0.1 else ("deteriorating" if momentum < -0.1 else "stable")
        print(f"\n  [DFM] {factor_name} factor - method: {method} - {len(panel_df)} months")
        print(f"    Loadings:  {load_str}")
        print(f"    Var expl:  {vexp_str}")
        print(f"    Momentum:  {momentum:+.2f} ({mom_dir})")

        # NBER backtest (growth factor only - correlation should be strongly negative)
        usrec = _fred("USREC", len(panel_df) + 5)
        if usrec:
            usrec_by_ym = {r[0][:7]: r[1] for r in usrec}
            aligned = [(factor_raw[ym], usrec_by_ym[ym])
                       for ym in factor_raw.index if ym in usrec_by_ym]
            if len(aligned) >= 10:
                f_vals = np.array([x[0] for x in aligned])
                r_vals = np.array([x[1] for x in aligned])
                nber_r = float(np.corrcoef(f_vals, r_vals)[0, 1])
                print(f"    NBER corr: {nber_r:.3f}  (negative = factor falls in recessions [OK])")

        return {
            "latest":        latest,
            "momentum":      momentum,
            "series":        factor_raw,
            "loadings":      loadings_v,
            "var_explained": var_exp,
            "top_contrib":   top_contrib,
            "method":        method,
        }
    except Exception as e:
        print(f"  [WARN] DFM estimation failed for {factor_name}: {e}")
        return None


def _dfm_signal(score):
    """Growth factor signal."""
    if score > 0.5:  return ("EXPANDING",   C["exp"])
    if score < -0.5: return ("CONTRACTING", C["con"])
    return ("NEUTRAL", C["neu"])


def _inf_signal(score):
    """Inflation factor signal (high inflation = red)."""
    if score > 0.5:  return ("ELEVATED", C["con"])
    if score < -0.5: return ("SUBDUED",  C["exp"])
    return ("MODERATE", C["neu"])


def _credit_signal(score):
    """Credit conditions factor signal (stress = red, loose = green)."""
    if score > 0.5:  return ("STRESSED",    C["con"])
    if score < -0.5: return ("SUPPORTIVE",  C["exp"])
    return ("NEUTRAL", C["neu"])


# ── HIDDEN MARKOV MODEL (HMM) REGIME DETECTION ──────────────────────────────
# Probabilistic regime classification using Gaussian HMM on DFM factor scores.
# Outputs probability distribution across 4 regimes, not binary classification.
# Built-in implementation using numpy/scipy - no external dependency required.

REGIME_NAMES_HMM = ["Early Recovery", "Reflation", "Stagflation", "Deflation"]


class _GaussianHMM:
    """Lightweight 4-state Gaussian HMM fitted via Baum-Welch (EM).
    Uses forward-backward algorithm for posterior state probabilities."""

    def __init__(self, n_states=4, n_dim=2, n_iter=100, seed=42):
        self.K = n_states
        self.D = n_dim
        self.n_iter = n_iter
        self.rng = np.random.RandomState(seed)
        self.pi = None       # initial state probs (K,)
        self.A  = None       # transition matrix (K, K)
        self.means = None    # emission means (K, D)
        self.covs  = None    # emission covariances (K, D, D)

    def _init_params(self, X):
        T = len(X)
        self.pi = np.ones(self.K) / self.K
        # Transition matrix: regimes are persistent (90% stay, 10% transition)
        self.A  = np.ones((self.K, self.K)) / self.K * 0.1 + np.eye(self.K) * 0.9
        # INFORMED INITIALIZATION: place state means at the 4 economic quadrants.
        # Data is z-scored (mean=0, std=1), so ±0.7 captures the regime structure.
        # State 0: Early Recovery  (growth+, inflation-)
        # State 1: Reflation      (growth+, inflation+)
        # State 2: Stagflation    (growth-, inflation+)
        # State 3: Deflation      (growth-, inflation-)
        g_std = max(X[:, 0].std(), 0.5)
        i_std = max(X[:, 1].std(), 0.5)
        spread = 0.7
        self.means = np.array([
            [+spread * g_std, -spread * i_std],  # Early Recovery
            [+spread * g_std, +spread * i_std],  # Reflation
            [-spread * g_std, +spread * i_std],  # Stagflation
            [-spread * g_std, -spread * i_std],  # Deflation
        ])
        # Initialize covariances from data in each quadrant
        self.covs = np.zeros((self.K, self.D, self.D))
        for k in range(self.K):
            # Select observations nearest to this state's mean
            dists = np.sum((X - self.means[k]) ** 2, axis=1)
            nearest = np.argsort(dists)[:max(T // self.K, 10)]
            self.covs[k] = np.cov(X[nearest].T) + np.eye(self.D) * 0.1

    def _emission_probs(self, X):
        """Compute B[t, k] = P(x_t | state=k) for all t, k."""
        T = len(X)
        B = np.zeros((T, self.K))
        for k in range(self.K):
            try:
                B[:, k] = _mvn.pdf(X, mean=self.means[k], cov=self.covs[k], allow_singular=True)
            except Exception:
                B[:, k] = 1e-300
        B = np.maximum(B, 1e-300)  # floor to avoid log(0)
        return B

    def _forward(self, B):
        T, K = B.shape
        alpha = np.zeros((T, K))
        alpha[0] = self.pi * B[0]
        scale = np.zeros(T)
        scale[0] = max(alpha[0].sum(), 1e-300)
        alpha[0] /= scale[0]
        for t in range(1, T):
            alpha[t] = (alpha[t-1] @ self.A) * B[t]
            s = alpha[t].sum()
            scale[t] = max(s, 1e-300)
            alpha[t] /= scale[t]
        return alpha, scale

    def _backward(self, B, scale):
        T, K = B.shape
        beta = np.zeros((T, K))
        beta[-1] = 1.0
        for t in range(T-2, -1, -1):
            beta[t] = (self.A @ (B[t+1] * beta[t+1]))
            beta[t] /= max(scale[t+1], 1e-300)
        return beta

    def fit(self, X):
        self._init_params(X)
        T = len(X)
        prev_ll = -np.inf
        for iteration in range(self.n_iter):
            B = self._emission_probs(X)
            alpha, scale = self._forward(B)
            beta = self._backward(B, scale)
            ll = np.sum(np.log(np.maximum(scale, 1e-300)))
            if abs(ll - prev_ll) < 1e-4:
                break
            prev_ll = ll

            # Posterior state probabilities
            gamma = alpha * beta
            g_sum = gamma.sum(axis=1, keepdims=True)
            g_sum = np.maximum(g_sum, 1e-300)
            gamma /= g_sum

            # Transition posteriors
            xi = np.zeros((self.K, self.K))
            for t in range(T-1):
                num = np.outer(alpha[t], B[t+1] * beta[t+1]) * self.A
                s = max(num.sum(), 1e-300)
                xi += num / s

            # M-step: update parameters
            self.pi = gamma[0] / max(gamma[0].sum(), 1e-300)
            for k in range(self.K):
                row_sum = max(xi[k].sum(), 1e-300)
                self.A[k] = xi[k] / row_sum
            # Floor transition probs at 1% so no state becomes absorbing
            self.A = np.maximum(self.A, 0.01)
            self.A /= self.A.sum(axis=1, keepdims=True)

            for k in range(self.K):
                w = gamma[:, k]
                w_sum = max(w.sum(), 1e-300)
                self.means[k] = (w[:, None] * X).sum(axis=0) / w_sum
                diff = X - self.means[k]
                self.covs[k] = (diff.T @ (diff * w[:, None])) / w_sum
                self.covs[k] += np.eye(self.D) * 0.01  # regularise (prevent spike collapse)

        return self

    def predict_proba(self, X):
        """Smoothed probabilities (forward-backward). Uses all data including future."""
        B = self._emission_probs(X)
        alpha, scale = self._forward(B)
        beta = self._backward(B, scale)
        gamma = alpha * beta
        g_sum = gamma.sum(axis=1, keepdims=True)
        g_sum = np.maximum(g_sum, 1e-300)
        gamma /= g_sum
        return gamma

    def predict_filtered(self, X):
        """Filtered probabilities (forward-only). No look-ahead bias."""
        B = self._emission_probs(X)
        alpha, _ = self._forward(B)
        return alpha

    def score(self, X):
        B = self._emission_probs(X)
        _, scale = self._forward(B)
        return np.sum(np.log(np.maximum(scale, 1e-300)))


def estimate_hmm_regime(growth_series, inflation_series, n_states=4):
    """
    Fit a 4-state Gaussian HMM to (growth, inflation) factor pair.
    States are initialised at the 4 economic quadrants:
      State 0 = Early Recovery (growth+, inflation-)
      State 1 = Reflation     (growth+, inflation+)
      State 2 = Stagflation   (growth-, inflation+)
      State 3 = Deflation     (growth-, inflation-)
    Returns dict with regime_name, probabilities, transition_matrix, or None on failure.
    """
    if growth_series is None or inflation_series is None:
        return None
    try:
        # Align on shared dates
        shared = sorted(set(growth_series.index) & set(inflation_series.index))
        if len(shared) < 48:
            print(f"  [WARN] HMM: only {len(shared)} aligned months - need >=48. Skipping.")
            return None

        X = np.column_stack([
            growth_series.loc[shared].values,
            inflation_series.loc[shared].values,
        ])

        # State labels are FIXED by the informed initialisation order
        # (the HMM init places means at the 4 quadrants in this order)
        STATE_LABELS = ["Early Recovery", "Reflation", "Stagflation", "Deflation"]

        best_model = None
        best_score = -np.inf
        # Multiple seeds to avoid poor local optima
        for seed in range(6):
            model = _GaussianHMM(n_states=n_states, n_dim=2, n_iter=150, seed=seed)
            model.fit(X)
            score = model.score(X)
            if score > best_score:
                best_score = score
                best_model = model

        if best_model is None:
            return None

        probs = best_model.predict_proba(X)
        filtered_probs = best_model.predict_filtered(X)
        current_probs = probs[-1]  # latest observation

        # After EM, verify state means still map to the correct quadrants.
        # Re-label states by their FINAL mean positions (EM may have moved them).
        means = best_model.means
        final_labels = {}
        used = set()
        # For each quadrant, find the state whose mean is closest to that quadrant
        quadrant_targets = {
            "Early Recovery": np.array([+1, -1]),
            "Reflation":      np.array([+1, +1]),
            "Stagflation":    np.array([-1, +1]),
            "Deflation":      np.array([-1, -1]),
        }
        for regime, target in quadrant_targets.items():
            # Cosine-like: which state mean has the right SIGN pattern?
            best_state = None
            best_align = -np.inf
            for si in range(n_states):
                if si in used:
                    continue
                # Alignment = dot product with target (measures quadrant match)
                alignment = np.dot(means[si], target)
                if alignment > best_align:
                    best_align = alignment
                    best_state = si
            if best_state is not None:
                final_labels[best_state] = regime
                used.add(best_state)

        # Build probability dict by regime name
        regime_probs = {r: 0.0 for r in REGIME_NAMES_HMM}
        for si, prob in enumerate(current_probs):
            regime_probs[final_labels.get(si, "Deflation")] += prob

        # Validate labels against fitted means - add economic descriptions
        state_descriptions = {}
        for si in range(n_states):
            lbl = final_labels.get(si, "?")
            g, i = means[si]
            if lbl == "Early Recovery" and g < -0.3:
                desc = "Trough/Recovery phase (growth rebuilding from low base)"
            elif lbl == "Deflation" and g > 0.1:
                desc = "Low-inflation expansion (growth positive, prices subdued)"
            else:
                desc = REGIME_DESC.get(lbl, "")
            state_descriptions[lbl] = f"g={g:+.1f} i={i:+.1f}"

        # Find primary regime
        primary = max(regime_probs, key=regime_probs.get)
        confidence = regime_probs[primary]

        # Transition warning: second-highest probability > 30%
        sorted_probs = sorted(regime_probs.values(), reverse=True)
        transition_risk = sorted_probs[1] > 0.30 if len(sorted_probs) > 1 else False

        # Console output with state mean positions for transparency
        print(f"\n  [HMM] Regime detection ({len(shared)} months, quadrant-initialised):")
        print(f"    State means after EM fitting:")
        for si in range(n_states):
            lbl = final_labels.get(si, "?")
            g, i = means[si]
            print(f"      {lbl:18s}  growth={g:+.2f}  inflation={i:+.2f}")
        print(f"    Current probabilities:")
        for r in REGIME_NAMES_HMM:
            bar = "#" * int(regime_probs[r] * 30)
            print(f"      {r:18s} {regime_probs[r]*100:5.1f}%  {bar}")
        if transition_risk:
            print(f"    !!  TRANSITION RISK -- second regime > 30%")

        # Build full regime history (all months, for timeline chart)
        # Build smoothed history (forward-backward, for regime timeline chart)
        regime_history = {"dates": [], "Early Recovery": [], "Reflation": [], "Stagflation": [], "Deflation": []}
        # Build filtered history (forward-only, no look-ahead bias, for backtest)
        regime_history_filtered = {"dates": [], "Early Recovery": [], "Reflation": [], "Stagflation": [], "Deflation": []}
        for t_idx, date in enumerate(shared):
            ym = date[:7]
            regime_history["dates"].append(ym)
            regime_history_filtered["dates"].append(ym)
            for hist_dict, prob_matrix in [(regime_history, probs), (regime_history_filtered, filtered_probs)]:
                month_probs = {r: 0.0 for r in REGIME_NAMES_HMM}
                for si, p in enumerate(prob_matrix[t_idx]):
                    month_probs[final_labels.get(si, "Deflation")] += p
                for r in REGIME_NAMES_HMM:
                    hist_dict[r].append(round(month_probs[r] * 100, 1))

        # Regime stability (Shannon entropy of probability distribution)
        import math
        entropy = -sum(p * math.log2(max(p, 1e-10)) for p in regime_probs.values())
        stability = round(1 - entropy / 2.0, 3)  # 1 = certain, 0 = uniform

        return {
            "regime":          primary,
            "confidence":      confidence,
            "probabilities":   regime_probs,
            "transition_risk": transition_risk,
            "stability":       stability,
            "transition":      best_model.A,
            "state_desc":      state_descriptions,
            "regime_history":  regime_history,
            "regime_history_filtered": regime_history_filtered,
        }
    except Exception as e:
        print(f"  [WARN] HMM estimation failed: {e}")
        return None


def _momentum_arrow(m):
    """3-month momentum arrow."""
    if m > 0.15:  return "▲"
    if m < -0.15: return "▼"
    return "→"


# ── DIRECTIONAL FORECAST & MOMENTUM ──────────────────────────────────────────
# Rate-of-change analysis for leading indicators: are conditions improving or deteriorating?

LEADING_KEYS = [k for k, v in INDICATOR_TYPE.items() if v == "Leading"]

def compute_leading_momentum(data):
    """
    For each leading indicator, compare current value vs 3-month-ago value.
    Returns list of (key, direction, pct_change) where direction is 'improving'/'deteriorating'/'stable'.
    Also returns aggregate counts.
    """
    results = []
    for key in LEADING_KEYS:
        val, hist = data.get(key, (None, []))
        if val is None or len(hist) < 4:
            continue
        current = hist[0][1]
        ago_3m  = hist[3][1] if len(hist) > 3 else hist[-1][1]
        if ago_3m == 0:
            results.append((key, "stable", 0.0))
            continue
        change = current - ago_3m

        # Determine direction based on whether movement is expansionary or contractionary
        # For most indicators, rising = improving. But for spreads/EPU, rising = deteriorating.
        inverted = key in ("hy_spread", "ig_spread", "epu")  # rising is BAD for these
        if inverted:
            change = -change  # flip so positive = improving

        # Threshold for "significant" change: >5% of 3M-ago value for levels, >0.3 for indices
        threshold = abs(ago_3m) * 0.05 if abs(ago_3m) > 10 else 0.3
        if change > threshold:
            direction = "improving"
        elif change < -threshold:
            direction = "deteriorating"
        else:
            direction = "stable"
        results.append((key, direction, change))

    improving    = sum(1 for _, d, _ in results if d == "improving")
    deteriorating = sum(1 for _, d, _ in results if d == "deteriorating")
    stable       = sum(1 for _, d, _ in results if d == "stable")
    return results, improving, deteriorating, stable


def directional_forecast(data, dfm_results=None):
    """
    Synthesize leading indicator momentum into a 3-6 month directional forecast.
    Returns (outlook_label, outlook_color, summary_text).
    """
    results, improving, deteriorating, stable = compute_leading_momentum(data)
    total = improving + deteriorating + stable
    if total == 0:
        return "INSUFFICIENT DATA", C["neu"], "Not enough leading indicator history"

    # DFM growth momentum as additional signal
    dfm_mom = 0.0
    if dfm_results and dfm_results.get("growth"):
        dfm_mom = dfm_results["growth"].get("momentum", 0.0)

    # Decision logic
    pct_deteriorating = deteriorating / total
    pct_improving     = improving / total

    if pct_deteriorating >= 0.6 or (pct_deteriorating >= 0.4 and dfm_mom < -0.3):
        label = "DETERIORATING"
        color = C["con"]
    elif pct_improving >= 0.6 or (pct_improving >= 0.4 and dfm_mom > 0.3):
        label = "IMPROVING"
        color = C["exp"]
    elif pct_deteriorating > pct_improving:
        label = "WEAKENING"
        color = C["neu"]
    elif pct_improving > pct_deteriorating:
        label = "STABILISING"
        color = C["stabilise"]
    else:
        label = "MIXED"
        color = C["neu"]

    summary = (f"{improving} improving  |  {stable} stable  |  {deteriorating} deteriorating  "
               f"out of {total} leading indicators")
    return label, color, summary


def compute_sahm_rule(unrate_hist):
    """
    Sahm Rule: 3-month avg unemployment minus 12-month min.
    >= 0.5 has preceded every US recession since 1970. Source: Federal Reserve.
    """
    vals = [v for _, v in unrate_hist[:12]]
    if len(vals) < 12: return None
    return round(sum(vals[:3]) / 3 - min(vals), 2)


# ── REGIME BACKTEST ──────────────────────────────────────────────────────────
# Uses the long-history DFM factors to classify historical regimes,
# then computes forward S&P 500 returns by regime as validation.

def backtest_regimes(growth_series, inflation_series, hmm_result=None):
    """
    Backtest regime classification against equity market forward returns.
    Returns dict: {regime: {"months": N, "avg_6m": X, "avg_12m": Y}} or None.
    """
    if growth_series is None or inflation_series is None:
        return None
    try:
        # Fetch SP500 monthly data
        sp_raw = _fred("SPASTT01USM661N", 500)  # S&P 500 OECD monthly index
        if len(sp_raw) < 24:
            print("  [WARN] Insufficient SP500 data for backtest")
            return None
        sp_monthly = _daily_to_monthly_mean(sp_raw) if any(sp_raw[i][0][:7] == sp_raw[i+1][0][:7]
                                                           for i in range(min(5, len(sp_raw)-1))) else sp_raw
        sp_by_ym = {d[:7]: v for d, v in sp_monthly}

        # Align growth, inflation, and SP500 on shared months
        shared = sorted(set(growth_series.index) & set(inflation_series.index) & set(sp_by_ym.keys()))
        if len(shared) < 24:
            print(f"  [WARN] Only {len(shared)} aligned months for backtest - need >=24")
            return None

        # Build HMM filtered regime lookup if available
        hmm_regime_lookup = {}
        if hmm_result and hmm_result.get("regime_history_filtered"):
            rh = hmm_result["regime_history_filtered"]
            for idx, ym in enumerate(rh["dates"]):
                probs = {r: rh[r][idx] for r in REGIME_NAMES_HMM}
                hmm_regime_lookup[ym] = max(probs, key=probs.get)

        # Classify regime for each month
        regime_returns = {r: {"fwd_6m": [], "fwd_12m": []} for r in REGIME_NAMES_HMM}
        for i, ym in enumerate(shared):
            # Use HMM filtered classification if available, else fall back to sign-based
            if ym in hmm_regime_lookup:
                regime = hmm_regime_lookup[ym]
            else:
                g = growth_series.loc[ym]
                inf = inflation_series.loc[ym]
                if   g > 0 and inf <= 0: regime = "Early Recovery"
                elif g > 0 and inf >  0: regime = "Reflation"
                elif g <= 0 and inf > 0: regime = "Stagflation"
                else:                    regime = "Deflation"

            sp_now = sp_by_ym.get(ym)
            if sp_now is None or sp_now == 0:
                continue

            # 6-month forward return
            fwd_6m_ym = shared[i + 6] if i + 6 < len(shared) else None
            if fwd_6m_ym and fwd_6m_ym in sp_by_ym:
                ret_6m = (sp_by_ym[fwd_6m_ym] - sp_now) / sp_now * 100
                regime_returns[regime]["fwd_6m"].append(ret_6m)

            # 12-month forward return
            fwd_12m_ym = shared[i + 12] if i + 12 < len(shared) else None
            if fwd_12m_ym and fwd_12m_ym in sp_by_ym:
                ret_12m = (sp_by_ym[fwd_12m_ym] - sp_now) / sp_now * 100
                regime_returns[regime]["fwd_12m"].append(ret_12m)

        # Summarize
        result = {}
        print(f"\n  [BACKTEST] Regime vs equity market forward returns ({len(shared)} months):")
        print(f"    {'Regime':18s}  {'Months':>6s}  {'Avg 6M':>8s}  {'Avg 12M':>8s}")
        for regime in REGIME_NAMES_HMM:
            r6  = regime_returns[regime]["fwd_6m"]
            r12 = regime_returns[regime]["fwd_12m"]
            avg6  = sum(r6) / len(r6) if r6 else 0.0
            avg12 = sum(r12) / len(r12) if r12 else 0.0
            n = max(len(r6), len(r12))
            result[regime] = {"months": n, "avg_6m": round(avg6, 1), "avg_12m": round(avg12, 1)}
            print(f"    {regime:18s}  {n:6d}  {avg6:+7.1f}%  {avg12:+7.1f}%")

        return result
    except Exception as e:
        print(f"  [WARN] Backtest failed: {e}")
        return None


# ── CHART DATA SHEET  (hidden - all charts pull their data from here) ─────────

CD = {k: 1 + i * 3 for i, k in enumerate([
    "unemployment_rate", "initial_claims", "nonfarm_payrolls",
    "cpi_yoy", "core_cpi_yoy", "ppi_mom", "crb_yoy",
    "ism_mfg", "ism_svc", "chicago_pmi", "consumer_conf",
    # Leading indicators
    "hy_spread", "ig_spread", "t10y3m", "building_permits",
    "jolts_openings", "jolts_quits", "breakeven_10y", "cfnai", "epu"])}
CD["fedfunds"] = max(CD.values()) + 3
# Positioning indicators - stored in the same hidden chart-data sheet
for _pk in POS_KEYS:
    CD[_pk] = max(CD.values()) + 3

def _date_lbl(d):
    """YYYY-MM[-DD] → MM-YY for chart x-axis (e.g. '2026-03-01' → '03-26')."""
    return f"{d[5:7]}-{d[2:4]}"

def build_chart_data(wb, data, ff_hist, pos_data=None):
    if "_Chart Data" in wb.sheetnames:
        del wb["_Chart Data"]
    ws = wb.create_sheet("_Chart Data")
    ws.sheet_state = "veryHidden"
    for key, (val, hist) in data.items():
        col = CD[key]
        _, name, _, _, _, _, y_lbl = METRIC_INFO[key]
        ws.cell(1, col,     f"{name} - Date")
        ws.cell(1, col + 1, f"{name} ({y_lbl})")
        for i, (date, v) in enumerate(reversed(hist[:12]), 2):
            ws.cell(i, col,     _date_lbl(date))
            ws.cell(i, col + 1, v)
    col = CD["fedfunds"]
    ws.cell(1, col,     "Fed Funds Rate - Date")
    ws.cell(1, col + 1, "Fed Funds Rate (%)")
    for i, (date, v) in enumerate(reversed(ff_hist[:48]), 2):
        ws.cell(i, col,     _date_lbl(date))
        ws.cell(i, col + 1, v)
    # Positioning indicators - used by the Market Positioning sheet charts
    if pos_data:
        for pk in POS_KEYS:
            col       = CD[pk]
            name, *_  = POS_META[pk]
            _, hist_p = pos_data.get(pk, (None, []))
            ws.cell(1, col,     f"{name} - Date")
            ws.cell(1, col + 1, name)
            # Yield curve is daily - use up to 260 obs (≈1 year); others weekly/quarterly use 52
            n_pos = 260 if pk == "yield_curve" else 52
            for i, (date, v) in enumerate(reversed(hist_p[:n_pos]), 2):
                ws.cell(i, col,     _date_lbl(date))
                ws.cell(i, col + 1, v)

    # PMI chart override: use 24 months from CSV instead of 12-month manual-input history
    for pmi_key, csv_path in CSV_PMI_MAP.items():
        col = CD[pmi_key]
        _, name, _, _, _, _, y_lbl = METRIC_INFO[pmi_key]
        csv_hist = _read_pmi_csv(csv_path)
        ws.cell(1, col,     f"{name} - Date")
        ws.cell(1, col + 1, f"{name} ({y_lbl})")
        for i, (date, v) in enumerate(reversed(csv_hist[:24]), 2):
            ws.cell(i, col,     _date_lbl(date))
            ws.cell(i, col + 1, v)

    return ws

def _make_chart(title, y_label, ch_ws, col, n_rows, line_color,
                width=24.0, height=13.0, bar=False, y_fmt="#,##0.0"):
    # builds a chart - somehow this works, don't question it
    chart = BarChart() if bar else LineChart()
    chart.title  = title
    chart.style  = 2
    chart.width  = width
    chart.height = height

    data_ref = Reference(ch_ws, min_col=col + 1, min_row=1, max_row=n_rows + 1)
    chart.add_data(data_ref, titles_from_data=True)
    cat_ref = Reference(ch_ws, min_col=col, min_row=2, max_row=n_rows + 1)
    chart.set_categories(cat_ref)

    chart.x_axis.delete      = False
    chart.y_axis.delete      = False
    chart.x_axis.tickLblPos  = "low"

    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines = ChartLines()

    chart.y_axis.numFmt  = y_fmt
    chart.y_axis.title   = y_label

    if chart.series:
        s = chart.series[0]
        if bar:
            s.graphicalProperties.solidFill = line_color
        else:
            s.graphicalProperties.line.solidFill = line_color
            s.graphicalProperties.line.width      = 22000
            s.smooth = True

    chart.legend = None
    return chart


# ── EXECUTIVE SUMMARY DASHBOARD ───────────────────────────────────────────────

def build_dashboard(wb, data, ff_hist, ch_ws, pos_data=None, dfm_results=None,
                    hmm_result=None, forecast=None, backtest=None, conviction=None):
    global _TEMPLATE_MODE
    _prev_tm = _TEMPLATE_MODE
    _TEMPLATE_MODE = False
    try:
        _build_dashboard_inner(wb, data, ff_hist, ch_ws, pos_data, dfm_results,
                               hmm_result, forecast, backtest, conviction)
    finally:
        _TEMPLATE_MODE = _prev_tm


def _build_dashboard_inner(wb, data, ff_hist, ch_ws, pos_data=None, dfm_results=None,
                           hmm_result=None, forecast=None, backtest=None, conviction=None):
    if "Executive Summary" in wb.sheetnames:
        del wb["Executive Summary"]
    ws = wb.create_sheet("Executive Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF0000"
    set_widths(ws, [4, 22, 30, 16, 11, 44.3, 34, 22, 20, 45.7, 14, 14])
    LCOL = 12

    # == SECTION A: TITLE (rows 1-2) ==========================================
    rh(ws, 1, 48)
    M(ws, 1, 1, 1, LCOL, "US MACROECONOMIC INDICATORS DASHBOARD",
      bg=C["title"], fc="FFFFFF", bold=True, size=20)
    rh(ws, 2, 18)
    M(ws, 2, 1, 2, LCOL,
      f"Last Refreshed: {datetime.now().strftime('%d %B %Y  |  %H:%M')}     "
      "|     Data: FRED API     |     ISM/Chicago PMI: Auto-extracted from PDFs",
      bg=C["sub"], fc="BDD7EE", size=9)

    # == SECTION B: THE HEADLINE (rows 4-7) ====================================
    has_dfm = dfm_results and dfm_results.get("growth") and dfm_results.get("inflation")
    dfm_factors_for_regime = None
    if has_dfm:
        dfm_factors_for_regime = {"growth": dfm_results["growth"]["latest"],
                                  "inflation": dfm_results["inflation"]["latest"]}
    regime     = classify_regime(data, dfm_factors=dfm_factors_for_regime)
    stance     = monetary_policy_stance(ff_hist)
    reg_color  = REGIME_COLORS[regime]
    style_fac  = STYLE_FACTORS[stance]
    sectors    = REGIME_SECTORS[regime]
    stance_col = C["exp"] if stance == "Easier Policy" else C["con"]
    if hmm_result:
        regime = hmm_result["regime"]; reg_color = REGIME_COLORS[regime]; sectors = REGIME_SECTORS[regime]

    w_label, w_score, w_lead, w_coin, w_lag = weighted_signal(data)
    f_label = forecast[0] if forecast else "N/A"
    f_color = forecast[1] if forecast else C["neu"]

    rh(ws, 3, 6)
    rh(ws, 4, 50)
    hmm_conf_txt = f"  {hmm_result['confidence']*100:.0f}%" if hmm_result else ""
    signal_M(ws, 4, 1, 4, 5, f"{regime.upper()}{hmm_conf_txt}",
             bg=reg_color, fc="FFFFFF", bold=True, size=22)
    signal_M(ws, 4, 6, 4, 8, f"OUTLOOK:  {f_label}",
             bg=f_color, fc="FFFFFF", bold=True, size=14)
    signal_M(ws, 4, 9, 4, LCOL, f"SCORE:  {w_score:+.2f}  {p_lbl(w_label)}",
             bg=p_bg(w_label), fc="FFFFFF", bold=True, size=14)

    if hmm_result:
        rh(ws, 5, 24)
        probs = hmm_result["probabilities"]
        cols_pr = LCOL // 4
        for ri, rname in enumerate(REGIME_NAMES_HMM):
            p = probs.get(rname, 0.0)
            cs = 1 + ri * cols_pr; ce = cs + cols_pr - 1
            signal_M(ws, 5, cs, 5, ce, f"{rname}  {p*100:.0f}%",
                     bg=REGIME_COLORS[rname], fc="FFFFFF", bold=True, size=9)
    else:
        rh(ws, 5, 5)

    rh(ws, 6, 24)
    growth_keys_r = ["unemployment_rate","initial_claims","nonfarm_payrolls","ism_mfg","ism_svc","consumer_conf"]
    growth_ps_r = [classify(k, data.get(k,(None,[]))[0])[0] for k in growth_keys_r if data.get(k,(None,[]))[0] is not None]
    growth_sig_r = overall_signal(growth_ps_r)
    growth_lbl_r = ("\u25b2 GROWTH UP" if growth_sig_r=="expansionary" else "\u25bc GROWTH DOWN" if growth_sig_r=="contractionary" else "\u25c6 GROWTH FLAT")
    cpi_now = data.get("cpi_yoy",(None,[]))[0]; crb_now = data.get("crb_yoy",(None,[]))[0]
    inf_up = (cpi_now is not None and cpi_now > 2.5) or (crb_now is not None and crb_now > 10.0)
    inf_dn = (cpi_now is not None and cpi_now < 1.5) and (crb_now is None or crb_now < -10.0)
    inf_lbl = ("\u25b2 INFLATION UP" if inf_up else "\u25bc INFLATION DOWN" if inf_dn else "\u25c6 INFLATION FLAT")
    inf_col = C["con"] if inf_up else (C["exp"] if inf_dn else C["neu"])
    unrate_hist_r = data.get("unemployment_rate",(None,[]))[1]
    sahm = compute_sahm_rule(unrate_hist_r)
    sahm_lbl = f"SAHM {sahm:.2f}" if sahm is not None else "SAHM N/A"
    sahm_col = C["con"] if sahm and sahm >= 0.5 else C["neu"] if sahm and sahm >= 0.3 else C["exp"] if sahm else C["neu"]
    signal_M(ws, 6, 1, 6, 3, growth_lbl_r, bg=p_bg(growth_sig_r), fc="FFFFFF", bold=True, size=9)
    signal_M(ws, 6, 4, 6, 6, inf_lbl, bg=inf_col, fc="FFFFFF", bold=True, size=9)
    stance_arrow = "\u25bc" if stance == "Easier Policy" else "\u25b2"
    signal_M(ws, 6, 7, 6, 9, f"{stance_arrow} {stance.upper()}", bg=stance_col, fc="FFFFFF", bold=True, size=9)
    signal_M(ws, 6, 10, 6, LCOL, sahm_lbl, bg=sahm_col, fc="FFFFFF", bold=True, size=9)

    rh(ws, 7, 22)
    for tier, val, cs, ce in [("Leading",w_lead,1,3),("Coincident",w_coin,4,6),("Lagging",w_lag,7,9)]:
        tbg = p_bg("expansionary" if val > 0.1 else ("contractionary" if val < -0.1 else "neutral"))
        signal_M(ws, 7, cs, 7, ce, f"{tier}: {val:+.2f}", bg=tbg, fc="FFFFFF", bold=True, size=9)
    if pos_data:
        pos_signals = [classify_pos(k, pos_data.get(k,(None,[]))[0])[0] for k in POS_KEYS]
        ov_lbl, ov_col = overall_positioning(pos_signals)
        signal_M(ws, 7, 10, 7, LCOL, f"POS: {ov_lbl}", bg=ov_col, fc="FFFFFF", bold=True, size=9)

    # == SECTION B2: METHODOLOGY RANKING ========================================
    cr = 8
    rh(ws, cr, 6); cr += 1

    # Dynamic readings for methodology table
    hmm_reading = f"{regime.upper()} {hmm_result['confidence']*100:.0f}%" if hmm_result else "Unavailable"
    score_reading = f"{w_score:+.2f} {p_lbl(w_label)}"
    dfm_reading = (f"Growth {dfm_results['growth']['latest']:+.2f}, Inflation {dfm_results['inflation']['latest']:+.2f}"
                   if has_dfm else "Unavailable")
    fc_reading = f_label
    cv_reading = conviction[0] if conviction else "N/A"

    meth_rows = [
        ("1", "HMM Regime Detection", hmm_reading, "HIGH",
         "4-state Hidden Markov Model on growth/inflation DFM factors (220 months). Primary signal.",
         "Overfits to GFC/COVID; may misclassify novel regimes. Trust when >80%.",
         "Primary. >80% = act on regime. <60% = treat as TRANSITION."),
        ("2", "Weighted Macro Score", score_reading, "HIGH",
         "29 indicators weighted by timing: leading 2-3x, coincident 1.5-2x, lagging 1x.",
         "Leading indicators dominate; can front-run false signals.",
         "Best single number. >+0.3 = bullish. <-0.3 = bearish."),
        ("3", "DFM Factor Model", dfm_reading, "MEDIUM",
         "3 latent factors (Growth, Inflation, Credit) via Kalman filter on 15+ series.",
         "Kalman unstable with <36 months; PCA fallback less precise.",
         "Cross-ref with HMM. Momentum arrows more useful than levels."),
        ("4", "Directional Forecast", fc_reading, "MEDIUM",
         "3-6 month outlook from 9 leading indicators' 3-month momentum.",
         "Pure extrapolation - cannot predict shocks or policy pivots.",
         "Trend confirmation only. Most useful when it agrees with HMM."),
        ("5", "Conviction Signal", cv_reading, "SUPPORTING",
         "Combines HMM confidence with forecast direction for allocation sizing.",
         "Derivative - inherits all upstream weaknesses.",
         "TRANSITION is the most actionable signal (raise cash)."),
    ]

    rh(ws, cr, 22)
    signal_M(ws, cr, 1, cr, LCOL,
             "  METHODOLOGY RANKING  --  Ordered by Predictive Importance",
             bg=C["col_hdr"], fc="FFFFFF", bold=True, size=10, h="left")
    cr += 1
    rh(ws, cr, 20)
    for j, h in enumerate(["#", "Method", "Reading", "Rank", "Description", "Failure Modes",
                            "How to Interpret", "", "", "", "", ""], 1):
        W(ws, cr, j, h, bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=8,
          h="left" if j in (5, 6, 7) else "center", bs="medium")
    cr += 1
    meth_start = cr
    for i, (num, mname, reading, rank, desc, fail, interp) in enumerate(meth_rows):
        rbg = C["row0"] if i % 2 == 0 else C["row1"]
        rh(ws, cr, 36)
        W(ws, cr, 1, num, bg=rbg, fc=C["dgray"], size=9)
        W(ws, cr, 2, mname, bg=rbg, bold=True, size=9, h="left")
        rank_bg = C["exp"] if rank in ("HIGH",) else (C["neu"] if rank == "MEDIUM" else C["dgray"])
        W(ws, cr, 3, reading, bg=rbg, bold=True, size=9)
        W(ws, cr, 4, rank, bg=rank_bg, fc="FFFFFF", bold=True, size=8)
        W(ws, cr, 5, desc, bg=rbg, size=8, h="left", wrap=True, fc=C["dgray"])
        W(ws, cr, 6, fail, bg=rbg, size=8, h="left", wrap=True, fc=C["dgray"])
        W(ws, cr, 7, interp, bg=rbg, size=8, h="left", wrap=True, fc=C["dgray"])
        for cc in range(8, LCOL + 1):
            W(ws, cr, cc, "", bg=rbg)
        cr += 1
    box_border(ws, meth_start, 1, cr - 1, LCOL)

    # == SECTION C: INDICATOR TABLE ============================================
    rh(ws, cr, 6); cr += 1
    rh(ws, cr, 22)
    M(ws, cr, 1, cr, 2, "SIGNAL LEGEND", bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=9)
    for i, (lbl, p) in enumerate([("EXPANSIONARY","exp"),("NEUTRAL","neu"),("CONTRACTIONARY","con")]):
        signal_M(ws, cr, 3+i*3, cr, 5+i*3, lbl, bg=C[p], fc="FFFFFF", bold=True, size=9)
    signal_M(ws, cr, 12, cr, 12, "WT", bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=8)
    cr += 1

    rh(ws, cr, 28)
    hdrs = ["","Category","Metric","Latest Value","Unit","Economic Impact","Likely Policy Response","Economic Signal","Macro Signal","Thresholds","Data Date","Wt"]
    for j, h in enumerate(hdrs, 1):
        W(ws, cr, j, h, bg=C["col_hdr"], fc="FFFFFF", bold=True, size=9,
          h="left" if j in (3,6,7,10) else "center", wrap=True, bs="medium")
    cr += 1
    if not _TEMPLATE_MODE:
        ws.freeze_panes = f"B{cr}"

    ind_start = cr
    n_metrics = len(METRICS)
    pressures = []
    for i, (key, name, cat, val_fmt, thresh, _, _) in enumerate(METRICS):
        r = ind_start + i
        val, hist = data.get(key, (None, []))
        pressure, impact, fed = classify(key, val)
        pressures.append(pressure)
        bg = C["row0"] if i % 2 == 0 else C["row1"]
        pending = (key in PDF_KEYS and val is None)
        rh(ws, r, 38)
        W(ws, r, 1, i+1, bg=bg, fc=C["dgray"], size=9)
        cb = ws.cell(row=r, column=2, value=cat)
        cb.fill = _fill(CAT_COLOR[cat]); cb.font = _font(bold=True, color="FFFFFF", size=8)
        cb.alignment = _align(h="center", wrap=True); cb.border = _border_all()
        W(ws, r, 3, name, bg=bg, bold=True, size=10, h="left")
        if key in METRIC_TOOLTIPS:
            ws.cell(row=r, column=3).comment = Comment(METRIC_TOOLTIPS[key], "Economic Tracker")
        if pending:
            cd = ws.cell(row=r, column=4, value="ADD PDF")
            cd.fill = _fill(C["input"]); cd.font = _font(bold=True, color=C["input_fc"], size=9)
            cd.alignment = _align(); cd.border = _border_all()
        else:
            cd = ws.cell(row=r, column=4, value=val)
            cd.fill = _fill(p_bg(pressure)); cd.font = _font(bold=True, color="FFFFFF", size=14)
            cd.alignment = _align(); cd.number_format = val_fmt; cd.border = _border_all()
        W(ws, r, 5, UNIT_MAP.get(key,""), bg=bg, fc=C["dgray"], size=8, italic=True)
        W(ws, r, 6, impact, bg=bg, size=9, h="left", wrap=True)
        W(ws, r, 7, fed, bg=bg, size=9, h="left", wrap=True)
        pressure_cell(ws, r, 8, pressure, pending=pending, size=9)
        signal_cell(ws, r, 9, pressure, pending=pending, size=9)
        W(ws, r, 10, thresh, bg=bg, size=8, h="left", wrap=True, italic=True, fc=C["dgray"])
        date_s = hist[0][0] if hist else ("PDF Not Found" if key in PDF_KEYS else "N/A")
        W(ws, r, 11, date_s, bg=bg, size=8, fc=C["dgray"])
        wt = INDICATOR_WEIGHTS.get(key, 1.0)
        wt_str = f"{wt:.0f}x" if wt == int(wt) else f"{wt:.1f}x"
        itype = INDICATOR_TYPE.get(key, "Lagging")
        W(ws, r, 12, wt_str, bg=TYPE_COLOR.get(itype, C["dgray"]), fc="FFFFFF", bold=True, size=8)

    cr = ind_start + n_metrics

    # == SECTION D: DFM FACTORS (1 row) ========================================
    rh(ws, cr, 6); cr += 1
    rh(ws, cr, 30)
    if has_dfm:
        g = dfm_results["growth"]; inf = dfm_results["inflation"]
        g_score = g["latest"]; g_label, g_bg = _dfm_signal(g_score)
        i_score = inf["latest"]; i_label, i_bg = _inf_signal(i_score)
        g_arr = _momentum_arrow(g["momentum"]); i_arr = _momentum_arrow(inf["momentum"])
        signal_M(ws, cr, 1, cr, 2, "DFM FACTORS", bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=9)
        has_credit = dfm_results.get("credit") is not None
        if has_credit:
            cred = dfm_results["credit"]; c_score = cred["latest"]; c_label, c_bg = _credit_signal(c_score)
            signal_M(ws, cr, 3, cr, 5, f"GROWTH {g_score:+.2f} {g_arr} {g_label}", bg=g_bg, fc="FFFFFF", bold=True, size=9)
            signal_M(ws, cr, 6, cr, 7, f"CREDIT {c_score:+.2f} {_momentum_arrow(cred['momentum'])} {c_label}", bg=c_bg, fc="FFFFFF", bold=True, size=9)
            signal_M(ws, cr, 8, cr, 10, f"INFLATION {i_score:+.2f} {i_arr} {i_label}", bg=i_bg, fc="FFFFFF", bold=True, size=9)
        else:
            signal_M(ws, cr, 3, cr, 6, f"GROWTH {g_score:+.2f} {g_arr} {g_label}", bg=g_bg, fc="FFFFFF", bold=True, size=9)
            signal_M(ws, cr, 7, cr, 10, f"INFLATION {i_score:+.2f} {i_arr} {i_label}", bg=i_bg, fc="FFFFFF", bold=True, size=9)
        dfm_reg = classify_regime(data, dfm_factors=dfm_factors_for_regime)
        signal_M(ws, cr, 11, cr, LCOL, f"DFM: {dfm_reg.upper()}", bg=REGIME_COLORS[dfm_reg], fc="FFFFFF", bold=True, size=8)
    else:
        signal_M(ws, cr, 1, cr, LCOL, "DFM FACTORS -- Unavailable", bg=C["input"], fc=C["input_fc"], bold=True, size=9)
    cr += 1

    # == SECTION E: SECTOR PLAYBOOK ============================================
    rh(ws, cr, 6); cr += 1
    rh(ws, cr, 22)
    signal_M(ws, cr, 1, cr, LCOL,
             f"  SECTOR PLAYBOOK  --  '{regime}' Regime     |     "
             f"STYLE: {stance.upper()} -- " + " > ".join(f.split(" > ")[0] for f in style_fac[:3]),
             bg=C["col_hdr"], fc="FFFFFF", bold=True, size=9, h="left")
    cr += 1
    sp_hdr = cr
    rh(ws, cr, 20)
    signal_M(ws, cr, 1, cr, 1, "#", bg=reg_color, fc="FFFFFF", bold=True, size=9)
    signal_M(ws, cr, 2, cr, 5, "SECTOR", bg=reg_color, fc="FFFFFF", bold=True, size=9)
    signal_M(ws, cr, 6, cr, 8, "AVG RETURN", bg=reg_color, fc="FFFFFF", bold=True, size=9)
    signal_M(ws, cr, 9, cr, LCOL, "SIGNAL", bg=reg_color, fc="FFFFFF", bold=True, size=9)
    cr += 1
    for i, (sector, ret) in enumerate(sectors):
        rh(ws, cr, 18)
        rbg = C["row0"] if i % 2 == 0 else C["row1"]
        ret_col = C["exp"] if ret > 5 else (C["ret_pos"] if ret > 0 else (C["neu"] if ret == 0 else C["con"]))
        sig_lbl = "OVERWEIGHT" if ret > 2 else ("NEUTRAL" if ret >= 0 else "UNDERWEIGHT")
        sig_col = C["exp"] if ret > 2 else (C["neu"] if ret >= 0 else C["con"])
        W(ws, cr, 1, i+1, bg=rbg, size=8, fc=C["dgray"])
        signal_M(ws, cr, 2, cr, 5, sector, bg=rbg, fc="000000", bold=(i<3), size=9, h="left")
        signal_M(ws, cr, 6, cr, 8, f"{'+' if ret>=0 else ''}{ret}%", bg=ret_col, fc="FFFFFF", bold=True, size=10)
        signal_M(ws, cr, 9, cr, LCOL, sig_lbl, bg=sig_col, fc="FFFFFF", bold=True, size=9)
        cr += 1
    box_border(ws, sp_hdr, 1, cr-1, LCOL)

    # Backtest (compact 1-row)
    if backtest:
        rh(ws, cr, 6); cr += 1
        rh(ws, cr, 20)
        signal_M(ws, cr, 1, cr, 4, "BACKTEST (12M fwd)", bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=8)
        for ri, rname in enumerate(REGIME_NAMES_HMM):
            bt = backtest.get(rname, {"months":0,"avg_6m":0.0,"avg_12m":0.0})
            cs = 5 + ri*2; ce = cs + 1
            r12 = bt["avg_12m"]; is_cur = (rname == regime)
            signal_M(ws, cr, cs, cr, ce, f"{rname} {r12:+.0f}%",
                     bg=REGIME_COLORS[rname] if is_cur else (C["exp"] if r12>0 else C["con"]),
                     fc="FFFFFF", bold=is_cur, size=8)
        cr += 1

    # Conviction (compact 1 row)
    if conviction:
        cv_label, cv_color, cv_desc = conviction
        rh(ws, cr, 6); cr += 1
        rh(ws, cr, 20)
        signal_M(ws, cr, 1, cr, LCOL,
                 f"  CONVICTION: {cv_label}  --  {cv_desc}",
                 bg=cv_color, fc="FFFFFF", bold=True, size=9, h="left")
        cr += 1

    # == SECTION H: ECONOMIC CALENDAR ==========================================
    cal = build_calendar_entries()
    if cal:
        rh(ws, cr, 6); cr += 1
        rh(ws, cr, 22)
        signal_M(ws, cr, 1, cr, LCOL,
                 "  UPCOMING DATA RELEASES  --  PMI requires manual PDF drop",
                 bg=C["col_hdr"], fc="FFFFFF", bold=True, size=10, h="left")
        cr += 1
        rh(ws, cr, 20)
        cal_hdrs = ["", "Priority", "Indicator", "Source", "Next Release", "Days", "", "", "", "", "", ""]
        for j, h in enumerate(cal_hdrs, 1):
            W(ws, cr, j, h, bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=8,
              h="center", bs="medium")
        cr += 1
        for i, (days, key, name, source, priority, rd) in enumerate(cal[:12]):
            rbg = C["input"] if source == "PDF" else (C["row0"] if i % 2 == 0 else C["row1"])
            rh(ws, cr, 18)
            W(ws, cr, 1, i + 1, bg=rbg, fc=C["dgray"], size=8)
            pri_bg = C["con"] if priority == "HIGH" else (C["neu"] if priority == "MEDIUM" else C["dgray"])
            signal_M(ws, cr, 2, cr, 2, priority, bg=pri_bg, fc="FFFFFF", bold=True, size=8)
            signal_M(ws, cr, 3, cr, 5, name, bg=rbg, fc="000000", bold=(source == "PDF"), size=9, h="left")
            W(ws, cr, 6, source, bg=rbg, fc=C["dgray"], size=8)
            signal_M(ws, cr, 7, cr, 8, rd.strftime("%d %b %Y"), bg=rbg, fc="000000", size=9)
            days_bg = C["con"] if days <= 3 else (C["neu"] if days <= 7 else rbg)
            signal_M(ws, cr, 9, cr, 10, f"{days}d", bg=days_bg, fc="FFFFFF" if days <= 7 else "000000", bold=True, size=9)
            action = "ACTION: ADD PDF" if source == "PDF" else "Auto-updated"
            signal_M(ws, cr, 11, cr, LCOL, action, bg=C["input"] if source == "PDF" else rbg,
                     fc=C["input_fc"] if source == "PDF" else C["dgray"], bold=(source == "PDF"), size=8)
            cr += 1
        box_border(ws, cr - len(cal[:12]), 1, cr - 1, LCOL)

    print("  Built: Executive Summary")


# ── CATEGORY SHEETS  (Job Market / Inflation / Economic Activities) ────────────

def build_category_sheet(wb, sheet_name, category, data, cat_col, ch_ws):
    if _TEMPLATE_MODE and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Clear merged cells from previous run - they conflict with new data writes
        for mg in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mg))
    else:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = CAT_TAB[category]
        set_widths(ws, [2, 26, 16, 24, 22, 58])

    LAST = 6

    rh(ws, 1, 44)
    M(ws, 1, 1, 1, LAST,
      f"  {category.upper()}  -  US ECONOMIC INDICATORS",
      bg=cat_col, fc="FFFFFF", bold=True, size=16)
    rh(ws, 2, 18)
    M(ws, 2, 1, 2, LAST,
      f"  Refreshed {datetime.now().strftime('%d %B %Y, %H:%M')}     |     "
      "Source: FRED API  |  ISM & Chicago PMI: Auto-extracted from PDF Reports",
      bg=C["sub"], fc="BDD7EE", size=9)
    rh(ws, 3, 10)
    if not _TEMPLATE_MODE:
        ws.freeze_panes = "A4"
    cr = 4

    cat_metrics = [(k, n, vfmt, t, ch_c, y_lbl)
                   for k, n, c, vfmt, t, ch_c, y_lbl in METRICS if c == category]

    for key, name, val_fmt, thresh, ch_c, y_lbl in cat_metrics:
        val, hist    = data.get(key, (None, []))
        pressure, impact, fed = classify(key, val)
        pending = (key in PDF_KEYS and val is None)

        rh(ws, cr, 32)
        M(ws, cr, 1, cr, LAST, f"  {name.upper()}",
          bg=cat_col, fc="FFFFFF", bold=True, size=13, h="left")
        cr += 1

        rh(ws, cr, 30); rh(ws, cr+1, 30)
        if pending:
            signal_M(ws, cr, 2, cr+1, 3, "ADD PDF TO REPORTS FOLDER",
                     bg=C["input"], fc=C["input_fc"], bold=True, size=11)
        else:
            if not _TEMPLATE_MODE:
                ws.merge_cells(start_row=cr, start_column=2, end_row=cr+1, end_column=3)
            vb = ws.cell(row=cr, column=2, value=val)
            vb.fill = _fill(p_bg(pressure)); vb.font = _font(bold=True, color="FFFFFF", size=22)
            vb.alignment = _align(); vb.number_format = val_fmt
            if not _TEMPLATE_MODE:
                box_border(ws, cr, 2, cr+1, 3)

        signal_M(ws, cr, 4, cr+1, 4,
                 p_lbl(pressure) if not pending else "PENDING",
                 bg=p_bg(pressure) if not pending else C["input"],
                 fc="FFFFFF"      if not pending else C["input_fc"],
                 bold=True, size=13)

        W(ws, cr,   5, "Economic Impact",       bg=cat_col, fc="FFFFFF", bold=True, size=9, h="left")
        W(ws, cr,   6, impact, bg=C["row1"], size=9, h="left", wrap=True)
        W(ws, cr+1, 5, "Likely Policy Response", bg=cat_col, fc="FFFFFF", bold=True, size=9, h="left")
        W(ws, cr+1, 6, fed,    bg=C["row0"], size=9, h="left", wrap=True)
        cr += 2

        rh(ws, cr, 22)
        M(ws, cr, 2, cr, LAST,
          f"  Thresholds  |  {thresh}",
          bg=C["thresh_bg"], fc=C["thresh_fc"], size=9, h="left", bold=False)
        cr += 1

        rh(ws, cr, 22)
        for col_idx, lbl in [(2, "Date"), (3, "Value"),
                             (4, "Signal"), (5, "Pressure"), (6, "Assessment")]:
            W(ws, cr, col_idx, lbl,
              bg=cat_col, fc="FFFFFF", bold=True, size=9, bs="medium",
              h="left" if col_idx == 6 else "center")
        cr += 1

        hist_d = list(hist[:12])
        if hist_d:
            for hi, (hdate, hval) in enumerate(hist_d):
                hp, ctx, _ = classify(key, hval)
                rbg = C["row0"] if hi % 2 == 0 else C["row1"]
                rh(ws, cr, 20)

                W(ws, cr, 2, hdate, bg=rbg, size=9, fc=C["dgray"])
                hc = ws.cell(row=cr, column=3, value=hval)
                hc.fill = _fill(rbg); hc.font = _font(size=9, bold=True)
                hc.alignment = _align(); hc.number_format = val_fmt
                hc.border = _border_all()

                signal_cell(  ws, cr, 4, hp, pending=False, size=8)
                pressure_cell(ws, cr, 5, hp, pending=False, size=8)

                W(ws, cr, 6, ctx, bg=rbg, size=8, h="left", wrap=True,
                  fc=C["dgray"], italic=True)
                cr += 1
            box_border(ws, cr - len(hist_d), 2, cr - 1, 6)
        else:
            rh(ws, cr, 22)
            M(ws, cr, 2, cr, LAST,
              "No data - drop the PDF into 'Reports' and re-run" if pending
              else "No data available - check FRED API key",
              bg="FFF8DC", fc="7D5A00", size=9, italic=True)
            cr += 1

        rh(ws, cr, 16); cr += 2

    rh(ws, cr, 12); cr += 1
    # PMI metrics use 24 months of CSV history; FRED metrics use 12 months
    has_pmi = any(k in CSV_PMI_MAP for k, *_ in cat_metrics)
    chart_hdr = ("  HISTORICAL TREND CHARTS - Last 24 Months  (PMI series)  |  Last 12 Months  (FRED series)"
                 if has_pmi else
                 "  HISTORICAL TREND CHARTS - Last 12 Months")
    rh(ws, cr, 34)
    M(ws, cr, 1, cr, LAST, chart_hdr,
      bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=12, h="left")
    cr += 1

    # Chart index - clickable list of charts
    index_start = cr
    rh(ws, cr, 20)
    M(ws, cr, 2, cr, LAST, "  CHART INDEX - Click a metric to jump to its chart",
      bg=C["lgray"], fc=C["dgray"], bold=True, size=9, h="left", italic=True)
    cr += 1
    chart_index_rows = {}  # key -> row where chart anchor will be
    chart_spacing = 22
    chart_anchor = cr + len(cat_metrics) + 1  # after the index rows + spacer
    for ci, (key, name, *_rest) in enumerate(cat_metrics):
        chart_index_rows[key] = chart_anchor + ci * chart_spacing
    for key, name, val_fmt, thresh, ch_c, y_lbl in cat_metrics:
        rh(ws, cr, 18)
        link_cell = ws.cell(row=cr, column=2, value=f"  ▸ {name}")
        link_cell.font = _font(color=C["link"], size=9, bold=False)
        link_cell.alignment = _align(h="left")
        link_cell.hyperlink = f"#'{sheet_name}'!A{chart_index_rows[key]}"
        cr += 1
    rh(ws, cr, 8); cr += 1

    # Always regenerate charts so references point to the current _Chart Data sheet
    if hasattr(ws, '_charts'):
        ws._charts = []

    # Charts with back-links
    for key, name, val_fmt, thresh, ch_c, y_lbl in cat_metrics:
        _, hist = data.get(key, (None, []))
        n = 24 if key in CSV_PMI_MAP else len(hist[:12])
        use_bar = (key == "nonfarm_payrolls")
        # Label row with back-to-index link
        rh(ws, cr, 20)
        label = ws.cell(row=cr, column=2, value=f"  {name}")
        label.font = _font(bold=True, size=10, color=cat_col)
        label.alignment = _align(h="left")
        back = ws.cell(row=cr, column=LAST, value="↑ Back to Index")
        back.font = _font(color=C["link"], size=8)
        back.hyperlink = f"#'{sheet_name}'!A{index_start}"
        cr += 1
        ws.add_chart(
            _make_chart(
                name, y_lbl, ch_ws, CD[key], n if n >= 2 else 2,
                ch_c, width=22.0, height=11.0, bar=use_bar,
                y_fmt=CHART_Y_FMT.get(key, "#,##0.0")
            ),
            f"B{cr}")
        cr += chart_spacing - 1

    print(f"  Built: {sheet_name}")


# ── GUIDE & SOURCES SHEET ─────────────────────────────────────────────────────

def build_positioning(wb, pos_data, ch_ws):
    """
    Market Positioning sheet - shows live FRED-sourced positioning/sentiment
    indicators plus reference links for CFTC COT, AAII, and margin debt.
    """
    global _TEMPLATE_MODE
    _prev_tm = _TEMPLATE_MODE
    _TEMPLATE_MODE = False
    try:
        _build_positioning_inner(wb, pos_data, ch_ws)
    finally:
        _TEMPLATE_MODE = _prev_tm


def _build_positioning_inner(wb, pos_data, ch_ws):
    if "Market Positioning" in wb.sheetnames:
        del wb["Market Positioning"]
    ws = wb.create_sheet("Market Positioning")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "640394"
    set_widths(ws, [2, 30, 14, 10, 34, 46, 30])

    LAST = 7

    # ── Header ────────────────────────────────────────────────────────────────
    rh(ws, 1, 48)
    M(ws, 1, 1, 1, LAST, "  MARKET POSITIONING & SENTIMENT DASHBOARD",
      bg=C["title"], fc="FFFFFF", bold=True, size=18, h="left")
    rh(ws, 2, 18)
    M(ws, 2, 1, 2, LAST,
      f"  Last Refreshed: {datetime.now().strftime('%d %B %Y  |  %H:%M')}     "
      "|     Sources: FRED API (CBOE · Chicago Fed · St. Louis Fed · Federal Reserve SLOOS)     "
      "|     Manual: CFTC COT · AAII · Margin Debt",
      bg=C["sub"], fc="BDD7EE", size=9)
    rh(ws, 3, 10)

    # ── Why Positioning Matters ───────────────────────────────────────────────
    rh(ws, 4, 28)
    M(ws, 4, 2, 4, LAST,
      "  WHY ANALYSE POSITIONING AND SENTIMENT?",
      bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=11, h="left")
    rh(ws, 5, 70)
    M(ws, 5, 2, 5, LAST,
      '"Non-linear moves in asset prices are largely driven by just two things: liquidity and '
      'positioning.  As investors we want to be on the right side of those moves.  Typically they '
      'occur when the Street is positioned disproportionately on one side of the trade.  When '
      'positioning is really stretched in one direction, even small counter-consensus surprises can '
      'cause big moves in price.  Think of it like a rubber band - the further it\'s stretched, the '
      'more violently it will snap-back on a catalyst."',
      bg=C["note_bg"], fc=C["dgray"], size=9, italic=True, h="left", wrap=True)
    rh(ws, 6, 18)
    M(ws, 6, 2, 6, LAST,
      "  Sources used: FRED (VIX · NFCI · STLFSI · SLOOS)  |  CFTC Commitment of Traders  "
      "|  AAII Investor Sentiment Survey  |  Fed Flow of Funds (Z.1)  |  Options Markets",
      bg=C["lgray"], fc=C["dgray"], size=9, italic=True, h="left")
    rh(ws, 7, 8)

    # ── Live FRED Indicators ──────────────────────────────────────────────────
    rh(ws, 8, 24)
    M(ws, 8, 1, 8, LAST,
      "  LIVE MARKET POSITIONING INDICATORS",
      bg=C["col_hdr"], fc="FFFFFF", bold=True, size=12, h="left")

    rh(ws, 9, 26)
    for ci, hdr in enumerate(["#", "Indicator", "Latest Value", "Unit",
                               "Condition / Signal", "Market Implication", "FRED Series"], 1):
        W(ws, 9, ci, hdr, bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=9,
          h="left" if ci in (2, 5, 6) else "center", bs="medium")

    pos_signals = []
    for i, pk in enumerate(POS_KEYS):
        r   = 10 + i
        rh(ws, r, 36)
        rbg = C["row0"] if i % 2 == 0 else C["row1"]
        name, sid, unit, thresh, ch_col = POS_META[pk]
        pval, _ = pos_data.get(pk, (None, []))
        ps, impact, action = classify_pos(pk, pval)
        pos_signals.append(ps)

        W(ws, r, 1, i + 1, bg=rbg, fc=C["dgray"], size=9)
        W(ws, r, 2, name,  bg=rbg, bold=True, size=9, h="left")

        # Value cell - coloured by signal
        vc = ws.cell(row=r, column=3,
                     value=round(pval, 3) if pval is not None else "N/A")
        vc.fill = _fill(pos_bg(ps)); vc.font = _font(bold=True, color="FFFFFF", size=12)
        vc.alignment = _align(); vc.border = _border_all()
        if pval is not None: vc.number_format = CHART_Y_FMT.get(pk, "0.00")

        W(ws, r, 4, unit,   bg=rbg, fc=C["dgray"], size=8, italic=True)
        W(ws, r, 5, impact, bg=rbg, size=9, h="left", wrap=True)
        W(ws, r, 6, action, bg=rbg, size=9, h="left", wrap=True)
        W(ws, r, 7, sid,    bg=rbg, size=8, fc=C["dgray"], italic=True)
    box_border(ws, 9, 1, 14, LAST)

    # ── Overall Positioning Signal ────────────────────────────────────────────
    rh(ws, 15, 8)
    ov_lbl, ov_col = overall_positioning(pos_signals)
    rh(ws, 16, 46)
    signal_M(ws, 16, 1, 16, LAST,
             f"OVERALL POSITIONING SIGNAL:   {ov_lbl}"
             f"     |     Risk-Off: {sum(1 for s in pos_signals if s in POS_RISK_OFF)}"
             f"   Neutral: {sum(1 for s in pos_signals if s not in POS_RISK_OFF | POS_RISK_ON)}"
             f"   Risk-On: {sum(1 for s in pos_signals if s in POS_RISK_ON)}",
             bg=ov_col, fc="FFFFFF", bold=True, size=14)
    rh(ws, 17, 8)

    # ── Manual Reference Sources ──────────────────────────────────────────────
    rh(ws, 18, 24)
    M(ws, 18, 1, 18, LAST,
      "  MANUAL REFERENCE SOURCES - Refresh these alongside the FRED data",
      bg=C["col_hdr"], fc="FFFFFF", bold=True, size=11, h="left")

    manual_sources = [
        ("CFTC Commitment of Traders",
         "Net speculative positioning for S&P 500, Treasuries, Gold, Oil, and FX futures. "
         "Released every Friday at 3:30 PM ET. Use 'Traders in Financial Futures' (TFF) report.",
         "publicreporting.cftc.gov / cftc.gov/MarketReports/CommitmentsofTraders"),
        ("AAII Investor Sentiment Survey",
         "Weekly bull/bear survey of individual investors. Bull-Bear spread historically mean-reverts. "
         "Extremes (>40% bulls or <20% bulls) are contrarian signals. Released every Thursday.",
         "aaii.com/sentimentsurvey"),
        ("FINRA Margin Debt",
         "Monthly total margin debt outstanding on NYSE. Rising margin with rising prices = "
         "leverage-driven rally (fragile). Falling margin + falling prices = forced deleveraging.",
         "finra.org/rules-guidance/key-topics/margin-accounts/margin-statistics"),
        ("Fed Flow of Funds  (Z.1)",
         "Quarterly breakdown of household equity allocations as % of total financial assets. "
         "Equity allocation >40% has historically coincided with poor forward 10-year returns.",
         "federalreserve.gov/releases/z1/  |  FRED series: BOGZ1FL153064476Q"),
    ]

    for i, (src, desc, link) in enumerate(manual_sources):
        r   = 19 + i
        rh(ws, r, 54)
        rbg = C["row0"] if i % 2 == 0 else C["row1"]
        W(ws, r, 2, src,  bg=C["lgray"], bold=True, size=9, h="left", v="top")
        W(ws, r, 3, "Manual",  bg=rbg, fc=C["dgray"], size=8, italic=True, v="top")
        W(ws, r, 4, " - ",       bg=rbg, fc=C["dgray"], size=8, v="top")
        c_desc = ws.cell(row=r, column=5, value=desc)
        c_desc.fill = _fill(rbg); c_desc.font = _font(size=9, italic=True)
        c_desc.alignment = _align(h="left", v="top", wrap=True); c_desc.border = _border_all()
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6) if not _TEMPLATE_MODE else None
        W(ws, r, 7, link, bg=rbg, fc="1155CC", size=8, h="left", v="top", italic=True)
    box_border(ws, 18, 1, 22, LAST)

    # ── Historical Charts ─────────────────────────────────────────────────────
    rh(ws, 23, 8)
    rh(ws, 24, 28)
    M(ws, 24, 1, 24, LAST, "  HISTORICAL TREND CHARTS - Positioning Indicators",
      bg=C["chart_hdr"], fc="FFFFFF", bold=True, size=12, h="left")

    # Chart index
    pos_cr = 25
    rh(ws, pos_cr, 20)
    M(ws, pos_cr, 2, pos_cr, LAST, "  CHART INDEX - Click to jump",
      bg=C["lgray"], fc=C["dgray"], bold=True, size=9, h="left", italic=True)
    pos_cr += 1
    pos_index_start = pos_cr
    pos_chart_spacing = 22
    pos_chart_base = pos_cr + len(POS_KEYS) + 1
    pos_chart_rows = {}
    for ci, pk in enumerate(POS_KEYS):
        pos_chart_rows[pk] = pos_chart_base + ci * pos_chart_spacing
    for pk in POS_KEYS:
        rh(ws, pos_cr, 18)
        name_p = POS_META[pk][0]
        link_cell = ws.cell(row=pos_cr, column=2, value=f"  ▸ {name_p}")
        link_cell.font = _font(color=C["link"], size=9)
        link_cell.alignment = _align(h="left")
        link_cell.hyperlink = f"#'Market Positioning'!A{pos_chart_rows[pk]}"
        pos_cr += 1
    rh(ws, pos_cr, 8); pos_cr += 1

    for pk in POS_KEYS:
        _, _, _, _, ch_col = POS_META[pk]
        _, hist_p = pos_data.get(pk, (None, []))
        n_limit = 260 if pk == "yield_curve" else 52
        n = len(hist_p[:n_limit])
        name_p = POS_META[pk][0]
        rh(ws, pos_cr, 20)
        label = ws.cell(row=pos_cr, column=2, value=f"  {name_p}")
        label.font = _font(bold=True, size=10, color=C["title"])
        label.alignment = _align(h="left")
        back = ws.cell(row=pos_cr, column=LAST, value="↑ Back to Index")
        back.font = _font(color=C["link"], size=8)
        back.hyperlink = f"#'Market Positioning'!A{pos_index_start}"
        pos_cr += 1
        if not _TEMPLATE_MODE and n >= 2:
            ws.add_chart(
                _make_chart(name_p, "Value", ch_ws, CD[pk], n, ch_col,
                            width=22.0, height=11.0,
                            y_fmt=CHART_Y_FMT.get(pk, "0.00")),
                f"B{pos_cr}")
        pos_cr += pos_chart_spacing - 1

    print("  Built: Market Positioning")


def build_notes(wb):
    if "Guide & Sources" in wb.sheetnames:
        del wb["Guide & Sources"]
    ws = wb.create_sheet("Guide & Sources")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["dgray"]
    set_widths(ws, [2, 28, 66, 2])
    rh(ws, 1, 40)
    M(ws, 1, 1, 1, 4, "  GUIDE, DATA SOURCES & INTERPRETATION",
      bg=C["title"], fc="FFFFFF", bold=True, size=14, h="left")

    sections = [
        ("HOW TO REFRESH", C["col_hdr"], [
            ("Monthly workflow",
             "1. Replace PDFs in 'Reports/' with the latest month's reports.  "
             "2. Run: py -3 economic_tracker.py  from the project folder.  "
             "3. The script extracts PMI values, updates the data log and regenerates "
             "US_Economic_Tracker.xlsx."),
            ("Automation",
             "See READ-ME.txt for full instructions on automating via Windows Task Scheduler, "
             "macOS launchd, or cron."),
            ("Portability",
             "This project is fully self-contained - all paths are relative to the script.  "
             "Copy the entire folder to any machine with Python installed and it will run "
             "without modification."),
        ]),
        ("PDF NAMING CONVENTIONS", C["ea"], [
            ("ISM Manufacturing PMI",
             "Released 1st business day of month.  Source: ismworld.org.  "
             "Filename must contain 'manuf'  (e.g. ISM Manufacturing.pdf)"),
            ("ISM Non-Manufacturing (Services) PMI",
             "Released 3rd business day of month.  Source: ismworld.org.  "
             "Filename must contain 'serv' or 'non-manuf'  (e.g. ISM Services.pdf)"),
            ("Chicago PMI  (MNI Chicago Business Barometer)",
             "Released last business day of month.  Source: MNI Indicators.  "
             "Filename must contain 'mni' or 'barometer'  "
             "(e.g. 20260228.MNI_Chicago_Bus_Barometer_Press_2026_02.pdf)"),
            ("Manual override",
             "If auto-extraction fails, open ISM_Manual_Input.xlsx and enter the value "
             "manually.  The script reads this file on every run."),
        ]),
        ("SIGNAL INTERPRETATION", C["jm"], [
            ("EXPANSIONARY  (green)",
             "The metric is consistent with a growing, healthy economy."),
            ("NEUTRAL  (amber)",
             "The metric is within the target or expected range.  No clear policy signal."),
            ("CONTRACTIONARY  (red)",
             "The metric signals economic weakness or above-target inflation."),
            ("Overall Assessment",
             "The overall signal requires a true majority (50%+).  If 50% or more of known "
             "metrics are contractionary the reading is CONTRACTIONARY; if 50% or more are "
             "expansionary it is EXPANSIONARY.  When neither side reaches 50%, the reading "
             "shows as MIXED with the leaning direction in parentheses - e.g. MIXED "
             "(CONTRACTIONARY).  Neutral metrics count toward the total denominator but do "
             "not break ties."),
            ("Chicago PMI  (Regional)",
             "The Chicago PMI is a regional survey covering the Midwest manufacturing sector.  "
             "It is included in the overall signal count but should be interpreted with "
             "slightly less weight than the national ISM headline indicators."),
            ("DFM Factor Scores  (3-factor model in v8)",
             "Three-factor Dynamic Factor Model estimated via Kalman filter (statsmodels).  "
             "Growth factor: UNRATE, IC4WSA, PAYEMS, UMCSENT, ISM Mfg/Svc, Chicago PMI + CFNAI, "
             "building permits, JOLTS (where history allows).  "
             "Inflation factor: CPI, Core CPI, PPI, CRB, 10Y breakeven, M2 growth.  "
             "Credit factor (NEW v8): HY OAS, IG OAS, NFCI, 10Y−3M yield curve.  "
             "z-score > +0.5 = EXPANDING/ELEVATED/STRESSED; < −0.5 = CONTRACTING/SUBDUED/SUPPORTIVE.  "
             "▲▼→ arrows show 3-month momentum.  PCA fallback if statsmodels is unavailable."),
            ("HMM Probabilistic Regime Detection  (v8)",
             "4-state Gaussian Hidden Markov Model fitted to the (growth, inflation) DFM factor "
             "time series.  Outputs probability distribution across all four regimes - e.g. "
             "'Reflation 62% | Stagflation 24%'.  Detects regime transitions before they are "
             "obvious in raw data.  Built-in implementation using numpy/scipy - no additional "
             "packages required.  Transition Risk flag appears when the second-highest regime "
             "exceeds 30% probability."),
            ("Weighted Macro Score  (v8)",
             "Replaces simple majority-vote as the primary macro assessment.  Each indicator "
             "is weighted by its predictive timing: Leading indicators (credit spreads, yield "
             "curve, building permits, JOLTS, CFNAI, breakevens) get 2-3x weight.  Coincident "
             "indicators (ISM, claims, payrolls) get 1.5-2x.  Lagging indicators (unemployment, "
             "CPI, PPI, consumer sentiment) get 1x.  Output: normalised score from −1.0 to +1.0 "
             "with breakdown by tier.  The unweighted vote is preserved as a secondary reference."),
            ("Sahm Rule  (regime sub-indicators)",
             "Federal Reserve recession indicator developed by Claudia Sahm.  "
             "Computed as 3-month average unemployment rate minus the 12-month minimum.  "
             "A reading of 0.50 or above has preceded every US recession since 1970 with no false "
             "positives.  Green < 0.30 (low risk) | Amber 0.30-0.49 (elevated) | Red ≥ 0.50 (recession risk).  "
             "Uses existing UNRATE history - no additional FRED call required."),
        ]),
        ("FRED DATA SOURCES", C["inf"], [
            ("Unemployment Rate",     "FRED: UNRATE         |  Bureau of Labor Statistics  |  Monthly"),
            ("Initial Jobless Claims","FRED: IC4WSA (4-wk avg) |  Dept of Labor           |  Weekly"),
            ("Nonfarm Payrolls",      "FRED: PAYEMS (MoM change)  |  BLS                  |  Monthly"),
            ("CPI YoY %",             "FRED: CPIAUCSL (YoY % computed)  |  BLS             |  Monthly"),
            ("Core CPI YoY %",        "FRED: CPILFESL (YoY % computed)  |  BLS             |  Monthly"),
            ("PPI - All Commodities MoM %",
             "FRED: PPIACO (MoM % computed)  |  BLS  |  Monthly  "
             "(All-commodity producer price index - not the finished-goods headline PPI)"),
            ("CRB Commodity Index YoY %",
             "FRED: PALLFNFINDEXM (IMF Primary Commodity Price Index - used as CRB proxy)  |  IMF  |  Monthly  "
             ">+10% = contractionary (inflationary pipeline)  |  <−5% = contractionary (deflationary)  |  else neutral.  "
             "Also feeds inflation panel of DFM."),
            ("Consumer Sentiment (UoM)", "FRED: UMCSENT  |  University of Michigan Consumer Sentiment Index  |  Monthly"),
            ("Federal Funds Rate",    "FRED: FEDFUNDS  |  Federal Reserve                   |  Monthly  (chart only)"),
            ("10Y-2Y Yield Curve Spread",
             "FRED: T10Y2Y  |  Federal Reserve Board  |  Daily (latest observation used).  "
             "> +0.5% = NORMAL (upward-sloping, growth-supportive)  |  −0.5% to +0.5% = FLAT (caution)  |  "
             "< −0.5% = INVERTED (classic recession warning, historically 12-18 months leading).  "
             "Shown on Market Positioning sheet with 1-year chart (~260 daily observations)."),
        ]),
        ("LEADING INDICATORS (v8)", C["li"], [
            ("HY Credit Spread (OAS)",
             "FRED: BAMLH0A0HYM2  |  ICE BofA US High Yield Index OAS  |  Daily (aggregated to monthly).  "
             ">500bp = credit stress (risk-off)  |  <350bp = tight spreads (risk-on).  "
             "Leading indicator: widens before every recession as investors flee to quality."),
            ("IG Credit Spread (OAS)",
             "FRED: BAMLC0A0CM  |  ICE BofA US Corporate Index OAS  |  Daily (aggregated to monthly).  "
             ">150bp = stress  |  <100bp = benign.  Investment-grade spreads filter out HY-specific noise."),
            ("10Y-3M Yield Curve",
             "FRED: T10Y3M  |  10-Year minus 3-Month Treasury spread  |  Daily.  "
             "More predictive than T10Y2Y - every US recession since 1960 preceded by inversion of this curve.  "
             "<0 = inverted (recession warning)  |  >1.5% = steep (growth-supportive)."),
            ("Building Permits (YoY %)",
             "FRED: PERMIT  |  New Privately-Owned Housing Units Authorized  |  Monthly.  "
             "12-month leading indicator - predicted 8 of last 9 US recessions.  "
             "<−10% YoY = contractionary  |  >+5% = expansionary."),
            ("JOLTS Job Openings (YoY %)",
             "FRED: JTSJOL  |  Job Openings: Total Nonfarm  |  Monthly.  "
             "Leading indicator of labour demand - falls before unemployment rate rises.  "
             "<−15% YoY = weakening demand  |  >+10% = strong demand."),
            ("JOLTS Quits Rate (%)",
             "FRED: JTSQUR  |  Quits: Total Nonfarm  |  Monthly.  "
             "Worker confidence indicator - people quit when they're confident of finding better jobs.  "
             "<2.0% = low confidence (cooling)  |  >2.5% = high confidence."),
            ("10Y Breakeven Inflation",
             "FRED: T10YIE  |  10-Year Breakeven Inflation Rate  |  Daily.  "
             "Market-implied inflation expectations (TIPS vs nominal).  Forward-looking unlike CPI.  "
             ">2.8% = elevated expectations  |  <1.8% = deflation risk  |  1.8-2.8% = anchored."),
            ("Chicago Fed National Activity Index",
             "FRED: CFNAI  |  Weighted average of 85 monthly indicators  |  Monthly.  "
             "Best single nowcast proxy for GDP growth.  Zero = long-run trend.  "
             "<−0.7 = recession signal  |  >+0.2 = above-trend growth."),
            ("Economic Policy Uncertainty (EPU)",
             "FRED: USEPUINDXD  |  Baker-Bloom-Davis EPU Index  |  Daily (aggregated to monthly).  "
             "Captures policy/geopolitical uncertainty from newspaper coverage + tax code expiration + "
             "economic forecaster disagreement.  >200 = extreme crisis-level (GFC, COVID, trade wars)  |  "
             "130-200 = elevated (growth headwind)  |  <80 = low/stable (supports growth)."),
        ]),
        ("DASHBOARD CONCEPTS (v8+)", C["col_hdr"], [
            ("Weighted Macro Score",
             "Primary overall signal. Each indicator is weighted by predictive timing: "
             "Leading indicators (credit spreads, yield curve, JOLTS, permits, CFNAI, EPU) get 2-3x weight.  "
             "Coincident indicators (ISM, claims, payrolls) get 1.5-2x.  Lagging (unemployment, CPI, "
             "consumer sentiment) get 1x.  Score ranges from -1.0 (hard contractionary) to +1.0 (expansionary).  "
             "The weight column (Wt) shows each indicator's multiplier.  Colours: teal = leading, "
             "blue = coincident, purple = lagging."),
            ("Sector Playbook Signals",
             "OVERWEIGHT: historical avg annual return > +2% in this regime -- tilt toward this sector.  "
             "NEUTRAL: return between 0% and +2% -- hold market weight.  "
             "UNDERWEIGHT: negative historical return -- reduce exposure.  "
             "Top 3 sectors are bolded.  Source: Invictus Research Equity Sector Backtest."),
            ("Conviction Signal",
             "Combines HMM regime confidence with the directional forecast to suggest allocation intensity.  "
             "HIGH = full regime tilt recommended (HMM > 80% and forecast directional).  "
             "MEDIUM = half tilt (regime clear but outlook mixed).  "
             "LOW = minimal tilt (regime uncertain).  TRANSITION = reduce risk, raise cash."),
            ("Backtest (12M Forward)",
             "Historical validation using S&amp;P 500 (FRED: SPASTT01USM661N).  For each month in the "
             "DFM factor history, the regime is classified and the 12-month forward equity return is computed.  "
             "Results are averaged by regime.  Small sample sizes -- treat as directional guidance, not forecasts."),
            ("Positioning (RISK-ON / RISK-OFF)",
             "Composite of 5 market indicators: VIX (fear), NFCI (financial conditions), STLFSI (financial "
             "stress), Bank Lending Standards (credit availability), 10Y-2Y Yield Curve (term structure).  "
             "Majority vote: more risk-off signals = RISK-OFF overall.  See Market Positioning sheet for detail."),
        ]),
    ]

    r = 3
    for sec_title, sec_bg, rows in sections:
        rh(ws, r, 28)
        M(ws, r, 2, r, 3, f"  {sec_title}",
          bg=sec_bg, fc="FFFFFF", bold=True, size=11, h="left")
        r += 1
        for lbl, desc in rows:
            rh(ws, r, 44)
            W(ws, r, 2, lbl,  bg=C["lgray"],   bold=True, size=9, h="left", v="top")
            W(ws, r, 3, desc, bg=C["note_bg"], size=9,    h="left", v="top", wrap=True)
            box_border(ws, r, 2, r, 3)
            r += 1
        rh(ws, r, 8); r += 1

    print("  Built: Guide & Sources")


# ── STATE TRACKING & ALERTS ───────────────────────────────────────────────────
# Persists regime/score between runs. Fires alerts when conditions change.

STATE_FILE = os.path.join(BASE_DIR, ".tracker_state.json")


def load_previous_state():
    if os.path.isfile(STATE_FILE):
        try:
            with open(STATE_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_current_state(regime, confidence, score, forecast_label, leading_score):
    state = {
        "timestamp":    datetime.now().isoformat(),
        "regime":       regime,
        "confidence":   round(confidence, 3),
        "weighted_score": round(score, 3),
        "forecast":     forecast_label,
        "leading_score": round(leading_score, 3),
    }
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2)
    except Exception as e:
        print(f"  [WARN] Could not save state: {e}")


def check_alerts(old, new_regime, new_confidence, new_score, new_forecast):
    """Compare previous state to current. Return list of alert messages."""
    alerts = []
    prev_regime   = old.get("regime")
    prev_score    = old.get("weighted_score")
    prev_forecast = old.get("forecast")

    if prev_regime and prev_regime != new_regime:
        msg = f"REGIME CHANGE: {prev_regime} -> {new_regime} ({new_confidence*100:.0f}% confidence)"
        alerts.append(msg)
        log_alert("regime_change", msg)
        send_email_alert("Regime Change", msg)

    if prev_score is not None and abs(new_score - prev_score) > 0.15:
        direction = "improved" if new_score > prev_score else "deteriorated"
        msg = f"MACRO SCORE {direction}: {prev_score:+.2f} -> {new_score:+.2f}"
        alerts.append(msg)
        log_alert("score_shift", msg)
        send_email_alert("Score Shift", msg)

    if prev_forecast and prev_forecast != new_forecast:
        msg = f"OUTLOOK CHANGE: {prev_forecast} -> {new_forecast}"
        alerts.append(msg)
        log_alert("outlook_change", msg)
        send_email_alert("Outlook Change", msg)

    return alerts


ALERT_LOG = os.path.join(DATA_DIR, "alert_log.json")

def log_alert(alert_type, message, details=None):
    """Append alert to persistent JSON log (kept on disk for dashboard)."""
    entry = {"timestamp": datetime.now().isoformat(), "type": alert_type, "message": message}
    if details:
        entry["details"] = details
    log = []
    if os.path.isfile(ALERT_LOG):
        try:
            with open(ALERT_LOG, encoding="utf-8") as f:
                log = json.load(f)
        except Exception:
            pass
    log.insert(0, entry)
    log = log[:50]  # keep last 50
    try:
        with open(ALERT_LOG, "w", encoding="utf-8") as f:
            json.dump(log, f, indent=2)
    except Exception:
        pass

def send_email_alert(subject, body):
    """Send alert via Outlook SMTP. Silently skips if not configured."""
    sender = os.environ.get("ALERT_EMAIL_FROM")
    password = os.environ.get("ALERT_EMAIL_PASSWORD")
    recipient = os.environ.get("ALERT_EMAIL_TO")
    if not all([sender, password, recipient]):
        return
    try:
        import smtplib
        from email.mime.text import MIMEText
        msg = MIMEText(body)
        msg["Subject"] = f"[US Econ Tracker] {subject}"
        msg["From"] = sender
        msg["To"] = recipient
        with smtplib.SMTP("smtp-mail.outlook.com", 587) as s:
            s.starttls()
            s.login(sender, password)
            s.send_message(msg)
        print(f"  [EMAIL] Alert sent to {recipient}")
    except Exception as e:
        print(f"  [WARN] Email alert failed: {e}")

def send_notification(title, message):
    """Try Windows toast + email, fall back to console."""
    print(f"\n  *** ALERT: {message} ***")
    log_alert("notification", message)
    send_email_alert(title, message)
    try:
        from plyer import notification
        notification.notify(title=title, message=message, timeout=15)
    except Exception:
        pass


# ── CONVICTION SIGNAL ─────────────────────────────────────────────────────────

def compute_conviction(hmm_result, forecast):
    """Combine HMM confidence + forecast direction into a conviction score."""
    if hmm_result is None:
        return "UNKNOWN", C["neu"], "HMM unavailable"

    conf = hmm_result.get("confidence", 0)
    regime = hmm_result.get("regime", "")
    stability = hmm_result.get("stability", 1.0)
    f_label = forecast[0] if forecast else "MIXED"

    # Stability override: if regime probabilities are spread, force TRANSITION
    if stability < 0.5:
        return "TRANSITION", C["warn"], "Regime uncertain (stability {:.0f}%) - reduce all tilts, raise cash".format(stability * 100)

    # Conflict check: HMM says growth-positive regime but leading indicators weakening
    if conf >= 0.8 and f_label in ("DETERIORATING", "WEAKENING") and regime in ("Early Recovery", "Reflation"):
        return "LOW", C["warn"], "Regime looks solid but leading indicators weakening - reduce tilt"
    elif conf >= 0.8 and f_label in ("DETERIORATING", "WEAKENING"):
        return "HIGH", C["con"], "Full regime allocation recommended"
    elif conf >= 0.8 and f_label in ("IMPROVING", "STABILISING"):
        return "HIGH", C["exp"], "Full regime allocation recommended"
    elif conf >= 0.8:
        return "MEDIUM", C["neu"], "Half tilt - regime clear but outlook mixed"
    elif conf >= 0.5:
        return "LOW", C["neu"], "Minimal tilt - regime uncertain, watch closely"
    else:
        return "TRANSITION", C["warn"], "Reduce all tilts, raise cash - regime in flux"


# ── MAIN ──────────────────────────────────────────────────────────────────────

def _validate_score_params():
    """Cross-validate SCORE_PARAMS against historical S&P 500 forward returns."""
    import csv as csv_mod
    print("=" * 60)
    print("  SCORE_PARAMS Cross-Validation")
    print("=" * 60)

    # Fetch extended history for all indicators
    print("\n  Fetching extended indicator history...")
    dfm_hist = fetch_dfm_history()

    # Fetch S&P 500 returns
    sp_raw = _fred("SPASTT01USM661N", 400)
    if not sp_raw or len(sp_raw) < 24:
        print("  [ERROR] Insufficient S&P 500 data"); return
    sp_by_ym = {d[:7]: v for d, v in sp_raw}
    sp_months = sorted(sp_by_ym.keys())
    # 3-month forward return for each month
    sp_fwd3 = {}
    for i, ym in enumerate(sp_months):
        if i + 3 < len(sp_months):
            fwd_ym = sp_months[i + 3]
            sp_fwd3[ym] = (sp_by_ym[fwd_ym] - sp_by_ym[ym]) / sp_by_ym[ym] * 100

    # For each FRED indicator, get 120-month history
    print("  Fetching individual indicator histories...")
    ind_history = {}
    fetch_map = {
        "unemployment_rate": ("UNRATE", "latest"), "initial_claims": ("IC4WSA", "latest"),
        "nonfarm_payrolls": ("PAYEMS", "payrolls"), "cpi_yoy": ("CPIAUCSL", "yoy"),
        "core_cpi_yoy": ("CPILFESL", "yoy"), "core_pce_yoy": ("PCEPILFE", "yoy"),
        "ppi_mom": ("PPIACO", "mom"), "crb_yoy": ("PALLFNFINDEXM", "yoy"),
        "consumer_conf": ("UMCSENT", "latest"), "cfnai": ("CFNAI", "latest"),
    }
    for key, (sid, ftype) in fetch_map.items():
        if ftype == "latest":
            raw = _fred(sid, 140)
            if raw:
                ind_history[key] = {d[:7]: v for d, v in raw}
        elif ftype == "yoy":
            _, hist = _yoy_extended(sid, 160) if True else (None, [])
            # Use the _yoy_extended which returns (val, [(date, val),...])
            result = _yoy_extended(sid, 160)
            if result and result[1]:
                ind_history[key] = {d[:7]: v for d, v in result[1]}
        elif ftype == "mom":
            result = _mom_extended(sid, 140)
            if result and result[1]:
                ind_history[key] = {d[:7]: v for d, v in result[1]}
        elif ftype == "payrolls":
            result = _payrolls_extended(140)
            if result and result[1]:
                ind_history[key] = {d[:7]: v for d, v in result[1]}

    # Sweep params for each indicator
    print(f"\n  Validating {len(ind_history)} indicators against S&P 500 3-month forward returns...\n")
    results = []
    for key, hist_dict in ind_history.items():
        params = SCORE_PARAMS.get(key)
        if not params:
            continue
        mid_orig = params["mid"]
        scale_orig = params["scale"]

        best_corr = -999
        best_mid = mid_orig
        best_scale = scale_orig

        # Sweep: 5 midpoints x 5 scales = 25 combos
        for mid_mult in [0.8, 0.9, 1.0, 1.1, 1.2]:
            for scale_mult in [0.7, 0.85, 1.0, 1.15, 1.3]:
                test_mid = mid_orig * mid_mult
                test_scale = scale_orig * scale_mult
                if test_scale <= 0:
                    continue

                scores = []
                fwd_returns = []
                for ym in sorted(hist_dict.keys()):
                    if ym not in sp_fwd3:
                        continue
                    val = hist_dict[ym]
                    if params.get("goldilocks"):
                        s = -abs(val - test_mid) / test_scale
                    elif params.get("invert"):
                        s = (test_mid - val) / test_scale
                    else:
                        s = (val - test_mid) / test_scale
                    s = max(-1.0, min(1.0, s))
                    scores.append(s)
                    fwd_returns.append(sp_fwd3[ym])

                if len(scores) < 24:
                    continue
                # Correlation
                import numpy as np
                corr = float(np.corrcoef(scores, fwd_returns)[0, 1])
                if corr > best_corr:
                    best_corr = corr
                    best_mid = test_mid
                    best_scale = test_scale

        # Current params correlation
        curr_scores = []
        curr_fwd = []
        for ym in sorted(hist_dict.keys()):
            if ym not in sp_fwd3:
                continue
            curr_scores.append(continuous_score(key, hist_dict[ym]))
            curr_fwd.append(sp_fwd3[ym])
        import numpy as np
        curr_corr = float(np.corrcoef(curr_scores, curr_fwd)[0, 1]) if len(curr_scores) >= 24 else 0

        diff = "OK" if abs(best_corr - curr_corr) < 0.05 else "IMPROVE"
        results.append({
            "key": key, "curr_mid": mid_orig, "curr_scale": scale_orig, "curr_corr": round(curr_corr, 3),
            "best_mid": round(best_mid, 2), "best_scale": round(best_scale, 2), "best_corr": round(best_corr, 3),
            "verdict": diff,
        })
        print(f"  {key:20s}  curr_corr={curr_corr:+.3f}  best_corr={best_corr:+.3f}  "
              f"mid: {mid_orig}->{round(best_mid,2)}  scale: {scale_orig}->{round(best_scale,2)}  [{diff}]")

    # Save to CSV
    out_path = os.path.join(DATA_DIR, "score_param_validation.csv")
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv_mod.DictWriter(f, fieldnames=["key","curr_mid","curr_scale","curr_corr","best_mid","best_scale","best_corr","verdict"])
        w.writeheader()
        w.writerows(results)
    print(f"\n  Results saved to {out_path}")
    print(f"  {sum(1 for r in results if r['verdict'] == 'OK')}/{len(results)} params within 0.05 of optimal")


def _run_walk_forward():
    """Full walk-forward backtest: retrain HMM each month with expanding window."""
    print("=" * 60)
    print("  Walk-Forward Backtest (this takes 10-20 minutes)")
    print("=" * 60)

    cache_path = os.path.join(DATA_DIR, "walk_forward_cache.json")

    # Fetch all long history
    print("\n  Step 1: Fetching extended history...")
    hmm_hist = fetch_hmm_history()

    # Build full panels
    hmm_g_series = []
    for name, key in [("unrate","unrate"),("ic4wsa","ic4wsa_m"),("payems","payems_m"),
                      ("umcsent","umcsent"),("cfnai","cfnai"),("permits","permits"),
                      ("indpro","indpro_m"),("housing","housing_yoy")]:
        hist = hmm_hist.get(key, (None, []))[1]
        if hist and len(hist) >= 36:
            hmm_g_series.append((name, hist))
    hmm_i_series = []
    for name, key in [("cpi","cpi_yoy"),("core","core_yoy"),("ppi","ppi_mom"),("reuters_crb","reuters_crb_yoy")]:
        hist = hmm_hist.get(key, (None, []))[1]
        if hist and len(hist) >= 36:
            hmm_i_series.append((name, hist))

    full_g_panel = _build_aligned_panel(hmm_g_series, n=580)
    full_i_panel = _build_aligned_panel(hmm_i_series, n=580)
    if full_g_panel is None or full_i_panel is None:
        print("  [ERROR] Insufficient panel data"); return

    n_months = len(full_g_panel)
    print(f"  Growth panel: {full_g_panel.shape}, Inflation panel: {full_i_panel.shape}")

    # Fetch S&P 500 for returns
    sp_raw = _fred("SPASTT01USM661N", 500)
    sp_by_ym = {d[:7]: v for d, v in sp_raw} if sp_raw else {}
    sp_months = sorted(sp_by_ym.keys())
    sp_rets = {}
    for i in range(len(sp_months) - 1):
        sp_rets[sp_months[i+1]] = round((sp_by_ym[sp_months[i+1]] - sp_by_ym[sp_months[i]]) / sp_by_ym[sp_months[i]] * 100, 2)

    # Walk forward: start with first 120 months, add one each iteration
    min_window = 120
    results = []
    g_dates = list(full_g_panel.index)
    i_dates = list(full_i_panel.index)
    shared_dates = sorted(set(g_dates) & set(i_dates))

    import numpy as np
    print(f"\n  Step 2: Running walk-forward ({n_months - min_window} iterations)...")

    for t in range(min_window, len(shared_dates)):
        window_dates = shared_dates[:t+1]
        ym = window_dates[-1]

        # Build sub-panels
        g_sub = full_g_panel.loc[full_g_panel.index.isin(window_dates)]
        i_sub = full_i_panel.loc[full_i_panel.index.isin(window_dates)]

        if len(g_sub) < 48 or len(i_sub) < 48:
            continue

        try:
            # Estimate DFM factors on expanding window using Kalman filter (same as main model)
            g_dfm = _estimate_dfm(g_sub, anchor_col="payems", factor_name=None, rolling_zscore=120)
            i_dfm = _estimate_dfm(i_sub, anchor_col="cpi", factor_name=None, rolling_zscore=120)
            if not g_dfm or not i_dfm:
                continue
            g_vals = g_dfm["series"].values
            i_vals = i_dfm["series"].values
            # Align on shared index
            shared_idx = sorted(set(g_dfm["series"].index) & set(i_dfm["series"].index))
            if len(shared_idx) < 48:
                continue
            g_aligned = g_dfm["series"].loc[shared_idx].values
            i_aligned = i_dfm["series"].loc[shared_idx].values

            # Fit HMM on expanding window
            X = np.column_stack([g_aligned, i_aligned])
            model = _GaussianHMM(n_states=4, n_dim=2, n_iter=80, seed=0)
            model.fit(X)

            # Forward-only probability for the last month
            alpha = model.predict_filtered(X)
            current_probs = alpha[-1]

            # Relabel states
            means = model.means
            quadrant_targets = {
                "Early Recovery": np.array([+1, -1]), "Reflation": np.array([+1, +1]),
                "Stagflation": np.array([-1, +1]), "Deflation": np.array([-1, -1]),
            }
            final_labels = {}
            used = set()
            for regime, target in quadrant_targets.items():
                best_state, best_align = None, -np.inf
                for si in range(4):
                    if si in used: continue
                    alignment = np.dot(means[si], target)
                    if alignment > best_align:
                        best_align = alignment
                        best_state = si
                if best_state is not None:
                    final_labels[best_state] = regime
                    used.add(best_state)

            regime_probs = {r: 0.0 for r in REGIME_NAMES_HMM}
            for si, p in enumerate(current_probs):
                regime_probs[final_labels.get(si, "Deflation")] += p
            dominant = max(regime_probs, key=regime_probs.get)

            sp_ret = sp_rets.get(ym, None)
            results.append({
                "date": ym,
                "regime": dominant,
                "probs": {r: round(float(v) * 100, 1) for r, v in regime_probs.items()},
                "sp500_return": sp_ret,
            })

            if (t - min_window) % 25 == 0:
                print(f"    {ym}  ({t - min_window + 1}/{len(shared_dates) - min_window})  {dominant}")

        except Exception as e:
            continue

    # Save cache
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2)
    print(f"\n  Walk-forward complete: {len(results)} months classified")
    print(f"  Results saved to {cache_path}")

    # Summary stats
    import math
    regime_buckets = {r: [] for r in REGIME_NAMES_HMM}
    for r in results:
        if r["sp500_return"] is not None:
            regime_buckets[r["regime"]].append(r["sp500_return"])
    print(f"\n  {'Regime':20s}  {'N':>4s}  {'Avg':>7s}  {'Sharpe':>7s}  {'Hit':>5s}")
    for regime in REGIME_NAMES_HMM:
        rets = regime_buckets[regime]
        n = len(rets)
        if n < 2: continue
        avg = sum(rets) / n
        std = (sum((x - avg)**2 for x in rets) / (n-1)) ** 0.5
        sharpe = avg / std * math.sqrt(12) if std > 0 else 0
        hit = sum(1 for x in rets if x > 0) / n * 100
        print(f"  {regime:20s}  {n:4d}  {avg:+6.2f}%  {sharpe:+6.2f}  {hit:4.0f}%")


def main():
    print("=" * 60)
    print("  US Economic Indicators Tracker")
    print(f"  {BASE_DIR}")
    print("=" * 60)

    print("\nStep 0: Checking for PMI data imports...")
    import_pmi_csvs()

    print("\nStep 1: Parsing PDF reports...")
    mfg_date, mfg_val, svc_date, svc_val, chi_date, chi_val, spg_date, spg_val = parse_reports_folder()

    print("\nStep 2: Updating data log...")
    update_manual_input(mfg_date, mfg_val, svc_date, svc_val, chi_date, chi_val, spg_date, spg_val)

    print("\nStep 2b: Updating CSV history...")
    update_pmi_csvs(mfg_date, mfg_val, svc_date, svc_val, chi_date, chi_val, spg_date, spg_val)

    # Fetch order: largest FRED pulls first to prime the cache, so subsequent
    # calls for the same series (with smaller limits) are instant cache hits.
    # This eliminates ~31 redundant HTTP requests per run.

    # ── HMM regime detection (using LONG-HISTORY panels) ─────────────────────
    print("\nStep 3: Fetching long-history FRED data for HMM...")
    hmm_result = None
    hmm_g = None
    hmm_i = None
    try:
        hmm_hist = fetch_hmm_history()
        # Build FRED-only panels with maximum history (no ISM CSV constraint)
        # Use whatever series are available (some may fail due to FRED API issues)
        hmm_g_series = []
        for name, key in [("unrate","unrate"),("ic4wsa","ic4wsa_m"),("payems","payems_m"),
                          ("umcsent","umcsent"),("cfnai","cfnai"),("permits","permits"),
                          ("indpro","indpro_m"),("housing","housing_yoy")]:
            hist = hmm_hist.get(key, (None, []))[1]
            if hist and len(hist) >= 36:
                hmm_g_series.append((name, hist))
        hmm_g_panel = _build_aligned_panel(hmm_g_series, n=700) if len(hmm_g_series) >= 3 else None
        hmm_i_series = []
        for name, key in [("cpi","cpi_yoy"),("core","core_yoy"),("ppi","ppi_mom")]:
            hist = hmm_hist.get(key, (None, []))[1]
            if hist and len(hist) >= 36:
                hmm_i_series.append((name, hist))
        hmm_i_panel = _build_aligned_panel(hmm_i_series, n=700) if len(hmm_i_series) >= 2 else None
        hmm_g = _estimate_dfm(hmm_g_panel, anchor_col="payems", factor_name="HMM-Growth", rolling_zscore=120) if hmm_g_panel is not None else None
        hmm_i = _estimate_dfm(hmm_i_panel, anchor_col="cpi",    factor_name="HMM-Inflation", rolling_zscore=120)
        if hmm_g and hmm_i:
            hmm_result = estimate_hmm_regime(hmm_g["series"], hmm_i["series"])
    except Exception as e:
        print(f"  [WARN] HMM long-history estimation failed: {e}")
    if hmm_result is None:
        print("  HMM unavailable - using DFM binary classification as fallback")

    print("\nStep 3b: Estimating DFM factors (growth + inflation + credit)...")
    dfm_results = {"growth": None, "inflation": None, "credit": None}
    try:
        dfm_hist = fetch_dfm_history()

        # Growth panel - original 7 series + leading indicators where available
        g_series = [
            ("unrate",      dfm_hist["unrate"][1]),
            ("ic4wsa",      dfm_hist["ic4wsa_m"][1]),
            ("payems",      dfm_hist["payems_m"][1]),
            ("umcsent",     dfm_hist["umcsent"][1]),
            ("ism_mfg",     dfm_hist["ism_mfg_csv"][1]),
            ("ism_svc",     dfm_hist["ism_svc_csv"][1]),
            ("chicago_pmi", dfm_hist["chicago_pmi_csv"][1]),
        ]
        # Add leading indicators to growth panel if history is sufficient
        if dfm_hist.get("cfnai") and len(dfm_hist["cfnai"][1]) >= 36:
            g_series.append(("cfnai", dfm_hist["cfnai"][1]))
        if dfm_hist.get("permit_yoy") and len(dfm_hist["permit_yoy"][1]) >= 36:
            g_series.append(("permits", dfm_hist["permit_yoy"][1]))
        if dfm_hist.get("jolts") and len(dfm_hist["jolts"][1]) >= 36:
            g_series.append(("jolts", dfm_hist["jolts"][1]))
        if dfm_hist.get("adp_m") and len(dfm_hist["adp_m"][1]) >= 36:
            g_series.append(("adp", dfm_hist["adp_m"][1]))
        if dfm_hist.get("indpro_mom") and len(dfm_hist["indpro_mom"][1]) >= 36:
            g_series.append(("indpro", dfm_hist["indpro_mom"][1]))
        if dfm_hist.get("housing_yoy") and len(dfm_hist["housing_yoy"][1]) >= 36:
            g_series.append(("housing", dfm_hist["housing_yoy"][1]))
        g_panel = _build_aligned_panel(g_series)

        # Inflation panel - original 4 + breakeven + M2
        i_series = [
            ("cpi",  dfm_hist["cpi_yoy"][1]),
            ("core", dfm_hist["core_yoy"][1]),
            ("ppi",  dfm_hist["ppi_mom"][1]),
            ("crb",  dfm_hist["crb_yoy"][1]),
        ]
        if dfm_hist.get("breakeven") and len(dfm_hist["breakeven"][1]) >= 36:
            i_series.append(("breakeven", dfm_hist["breakeven"][1]))
        if dfm_hist.get("m2_yoy") and len(dfm_hist["m2_yoy"][1]) >= 36:
            i_series.append(("m2", dfm_hist["m2_yoy"][1]))
        if dfm_hist.get("pce_yoy") and len(dfm_hist["pce_yoy"][1]) >= 36:
            i_series.append(("pce", dfm_hist["pce_yoy"][1]))
        if dfm_hist.get("import_yoy") and len(dfm_hist["import_yoy"][1]) >= 36:
            i_series.append(("imports", dfm_hist["import_yoy"][1]))
        i_panel = _build_aligned_panel(i_series)

        # Credit conditions panel - NEW in v8
        c_series = []
        if dfm_hist.get("hy_oas") and len(dfm_hist["hy_oas"][1]) >= 36:
            c_series.append(("hy_oas", dfm_hist["hy_oas"][1]))
        if dfm_hist.get("ig_oas") and len(dfm_hist["ig_oas"][1]) >= 36:
            c_series.append(("ig_oas", dfm_hist["ig_oas"][1]))
        if dfm_hist.get("nfci_m") and len(dfm_hist["nfci_m"][1]) >= 36:
            c_series.append(("nfci", dfm_hist["nfci_m"][1]))
        if dfm_hist.get("t10y3m") and len(dfm_hist["t10y3m"][1]) >= 36:
            c_series.append(("t10y3m", dfm_hist["t10y3m"][1]))

        c_panel = _build_aligned_panel(c_series) if len(c_series) >= 2 else None

        dfm_results = {
            "growth":    _estimate_dfm(g_panel,  anchor_col="payems", factor_name="Growth"),
            "inflation": _estimate_dfm(i_panel,  anchor_col="cpi",    factor_name="Inflation"),
            "credit":    _estimate_dfm(c_panel,  anchor_col="nfci",   factor_name="Credit") if c_panel is not None else None,
        }
    except Exception as e:
        print(f"  [WARN] DFM estimation failed: {e}")

    # Now fetch indicator data - most series already cached from HMM/DFM pulls above
    print("\nStep 3c: Fetching indicator data (macro + leading + positioning)...")
    data     = fetch_all()
    ff_hist  = fetch_fedfunds(48)
    pos_data = fetch_positioning()

    # Cross-validate: compare HMM regime vs vote-based regime
    vote_regime = classify_regime(data)
    if hmm_result:
        hmm_regime = hmm_result["regime"]
        hmm_result["vote_regime"] = vote_regime
        if hmm_regime != vote_regime:
            print(f"\n  [CROSS-CHECK] HMM: '{hmm_regime}' ({hmm_result['confidence']*100:.0f}%)  vs  Indicators: '{vote_regime}'")
            print(f"    HMM uses {len(hmm_result.get('probabilities',{}))} states trained on 220-month DFM factors")
            print(f"    Vote uses current indicator values with threshold rules")
        print(f"  HMM regime: {hmm_result['regime']}  |  Vote regime: {vote_regime}")
    else:
        print(f"  Vote-based regime: {vote_regime}")

    # ── Weighted signal summary ───────────────────────────────────────────────
    w_label, w_score, w_lead, w_coin, w_lag = weighted_signal(data)
    print(f"\n  [WEIGHTED] Macro Score: {w_score:+.2f}  ({w_label.upper()})")
    print(f"    Leading: {w_lead:+.2f}  |  Coincident: {w_coin:+.2f}  |  Lagging: {w_lag:+.2f}")

    # ── Directional forecast ──────────────────────────────────────────────────
    print("\nStep 3d: Computing directional forecast...")
    forecast = directional_forecast(data, dfm_results)
    f_label, f_color, f_summary = forecast
    print(f"  3-6 Month Outlook: {f_label}")
    print(f"    {f_summary}")

    # ── Regime backtest ───────────────────────────────────────────────────────
    print("\nStep 3e: Running regime backtest vs S&P 500...")
    backtest_result = None
    if hmm_g and hmm_i:
        backtest_result = backtest_regimes(hmm_g["series"], hmm_i["series"], hmm_result)
    elif dfm_results.get("growth") and dfm_results.get("inflation"):
        backtest_result = backtest_regimes(
            dfm_results["growth"]["series"], dfm_results["inflation"]["series"])

    # ── State tracking & alerts ─────────────────────────────────────────────
    current_regime = hmm_result["regime"] if hmm_result else vote_regime
    current_conf   = hmm_result["confidence"] if hmm_result else 0.0
    prev_state = load_previous_state()
    alerts = check_alerts(prev_state, current_regime, current_conf, w_score, f_label)
    for a in alerts:
        send_notification("US Economic Tracker", a)
    if not alerts and prev_state.get("regime"):
        print("\n  [STATE] No significant changes since last run")
    save_current_state(current_regime, current_conf, w_score, f_label, w_lead)

    # ── Conviction signal ─────────────────────────────────────────────────────
    conviction_label, conviction_color, conviction_desc = compute_conviction(hmm_result, forecast)
    print(f"\n  [CONVICTION] {conviction_label} - {conviction_desc}")

    print("\nStep 4: Building workbook...")
    global _TEMPLATE_MODE
    # Template mode: if a v8 template exists, load it and update data only  - 
    # preserving the user's custom formatting, column widths, row heights, colours.
    # The dashboard and positioning sheets always rebuild from scratch (they
    # force _TEMPLATE_MODE = False internally), but category sheets (Job Market,
    # Inflation, Economic Activities, Leading Indicators) will preserve formatting.
    if os.path.isfile(TEMPLATE_FILE):
        try:
            wb = openpyxl.load_workbook(TEMPLATE_FILE)
            _TEMPLATE_MODE = True
            print(f"  Template loaded: {os.path.basename(TEMPLATE_FILE)} - category sheet formatting preserved")
        except Exception as e:
            print(f"  [WARN] Template load failed ({e}) - building from scratch")
            _TEMPLATE_MODE = False
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
    else:
        _TEMPLATE_MODE = False
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        print("  No template found - building from scratch (first run)")

    ch_ws = build_chart_data(wb, data, ff_hist, pos_data)
    conviction = (conviction_label, conviction_color, conviction_desc)
    build_dashboard(wb, data, ff_hist, ch_ws, pos_data, dfm_results=dfm_results,
                    hmm_result=hmm_result, forecast=forecast, backtest=backtest_result,
                    conviction=conviction)
    build_category_sheet(wb, "Job Market",          "Job Market",          data, C["jm"],  ch_ws)
    build_category_sheet(wb, "Inflation",            "Inflation",           data, C["inf"], ch_ws)
    build_category_sheet(wb, "Economic Activities",  "Economic Activities", data, C["ea"],  ch_ws)
    build_category_sheet(wb, "Leading Indicators",   "Leading Indicators",  data, C["li"],  ch_ws)
    build_positioning(wb, pos_data, ch_ws)
    build_notes(wb)

    try:
        wb.save(OUTPUT_FILE)
        try:
            wb.save(TEMPLATE_FILE)
        except PermissionError:
            pass   # template save is best-effort
        print(f"\n{'='*60}")
        print(f"  Done.  Saved: {os.path.basename(OUTPUT_FILE)}")
        print(f"{'='*60}\n")
    except PermissionError:
        print(f"\n{'='*60}")
        print(f"  ERROR: Cannot save - close {os.path.basename(OUTPUT_FILE)} in Excel and re-run.")
        print(f"{'='*60}\n")

def _daemon_loop(time_str):
    """
    Self-scheduling daemon: runs main() daily at HH:MM.
    If the run fails, retries every 3 hours (up to 3 attempts).
    No external scheduler needed - just leave this process running.
    """
    hh, mm = int(time_str[:2]), int(time_str[3:])
    print(f"\n  [DAEMON] Starting - target time: {time_str} daily")
    print(f"  [DAEMON] Press Ctrl+C to stop\n")
    while True:
        now = datetime.now()
        target = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
        if target <= now:
            from datetime import timedelta
            target += timedelta(days=1)
        wait = (target - datetime.now()).total_seconds()
        print(f"  [DAEMON] Next run: {target.strftime('%Y-%m-%d %H:%M')}  "
              f"(sleeping {wait/3600:.1f}h)")
        try:
            while (target - datetime.now()).total_seconds() > 0:
                time.sleep(min(60, max(0, (target - datetime.now()).total_seconds())))
        except KeyboardInterrupt:
            print("\n  [DAEMON] Stopped by user."); return

        success = False
        for attempt in range(1, 4):
            try:
                print(f"\n  [DAEMON] Run attempt {attempt}/3 at {datetime.now().strftime('%H:%M')}")
                main()
                success = True
                print(f"  [DAEMON] Run completed successfully.")
                break
            except Exception as e:
                print(f"  [DAEMON] Attempt {attempt} failed: {e}")
                if attempt < 3:
                    print(f"  [DAEMON] Retrying in 3 hours...")
                    try:
                        time.sleep(10800)
                    except KeyboardInterrupt:
                        print("\n  [DAEMON] Stopped by user."); return
        if not success:
            print(f"  [DAEMON] All 3 attempts failed. Will try again tomorrow.")


def _schedule_task(time_str):
    """Create a Windows Task Scheduler task to run daily at the given time."""
    if not re.match(r'^\d{2}:\d{2}$', time_str):
        print(f"  ERROR: Invalid time format '{time_str}' - expected HH:MM (e.g. 07:00)")
        return
    import subprocess, sys
    script = os.path.abspath(__file__)
    python = sys.executable
    tr = f'"{python}" "{script}"'
    cmd = ['schtasks', '/create', '/tn', 'USEconomicTracker',
           '/tr', tr, '/sc', 'DAILY', '/st', time_str, '/f']
    print(f"  Creating scheduled task: daily at {time_str}")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode == 0:
        print(f"  SUCCESS - task 'US Economic Tracker' created. View in taskschd.msc")
    else:
        print(f"  FAILED - {result.stderr.strip()}")
        print(f"  Try running as Administrator if you get 'Access Denied'")


def _unschedule_task():
    """Remove the scheduled task."""
    import subprocess
    result = subprocess.run(['schtasks', '/delete', '/tn', 'USEconomicTracker', '/f'],
                          capture_output=True, text=True)
    if result.returncode == 0:
        print("  Scheduled task removed.")
    else:
        print(f"  {result.stderr.strip()}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="US Economic Indicators Tracker")
    parser.add_argument("--schedule", metavar="HH:MM",
                        help="Create a daily Windows Task Scheduler task at the given time")
    parser.add_argument("--unschedule", action="store_true",
                        help="Remove the scheduled task")
    parser.add_argument("--daemon", metavar="HH:MM",
                        help="Run as a self-scheduling daemon: daily at HH:MM, retries every 3h on failure")
    parser.add_argument("--offline", action="store_true",
                        help="Run without network - use cached FRED data only")
    parser.add_argument("--clear-cache", action="store_true",
                        help="Delete the FRED cache file and exit")
    parser.add_argument("--validate-params", action="store_true",
                        help="Cross-validate SCORE_PARAMS against historical S&P 500 returns")
    parser.add_argument("--walk-forward", action="store_true",
                        help="Run full walk-forward backtest (slow, ~15 min) and cache results")
    args = parser.parse_args()

    if args.validate_params:
        _validate_score_params()
    elif args.walk_forward:
        _run_walk_forward()
    elif args.clear_cache:
        if os.path.exists(CACHE_FILE):
            os.remove(CACHE_FILE)
            print(f"  Cache cleared: {CACHE_FILE}")
        else:
            print("  No cache file found.")
    elif args.daemon:
        if not re.match(r'^\d{2}:\d{2}$', args.daemon):
            print(f"  ERROR: Invalid time format '{args.daemon}' - expected HH:MM")
        else:
            _daemon_loop(args.daemon)
    elif args.schedule:
        _schedule_task(args.schedule)
    elif args.unschedule:
        _unschedule_task()
    else:
        if args.offline:
            globals()['_OFFLINE_MODE'] = True
            print("  [OFFLINE MODE] Using cached data only - no network requests")
        main()
